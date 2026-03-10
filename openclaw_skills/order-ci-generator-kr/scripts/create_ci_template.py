"""
创建CI模板文件 - 标准xlsx格式
参照原CI模板风格
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 创建新工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

# ============================================================
# 样式定义 - 参照原CI模板风格
# ============================================================

# 标题样式
title_font = Font(name='Arial', size=16, bold=True)
title_alignment = Alignment(horizontal='center', vertical='center')

# 标签样式（左侧灰色标签）
label_font = Font(name='Arial', size=10, bold=True)
label_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
label_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# 内容样式
content_font = Font(name='Arial', size=10)
content_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 表头样式
header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# 数据行样式
data_font = Font(name='Arial', size=9)
data_alignment = Alignment(horizontal='center', vertical='center')

# Total行样式
total_font = Font(name='Arial', size=10, bold=True)
total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# Section标题样式
section_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
section_fill = PatternFill(start_color='2E5FA3', end_color='2E5FA3', fill_type='solid')

# 边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# 设置列宽
# ============================================================
col_widths = [20, 18, 28, 22, 12, 15, 18]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ============================================================
# Row 1: 标题
# ============================================================
ws['A1'] = 'COMMERCIAL INVOICE'
ws['A1'].font = title_font
ws['A1'].alignment = title_alignment
ws.merge_cells('A1:G1')
ws.row_dimensions[1].height = 30

# ============================================================
# Row 2: Invoice No. / Date
# ============================================================
ws['A2'] = 'Invoice No.'
ws['A2'].font = label_font
ws['A2'].fill = label_fill
ws['A2'].alignment = label_alignment
ws['A2'].border = thin_border

ws['B2'] = 'UNI2026012101'
ws['B2'].font = content_font
ws['B2'].alignment = content_alignment
ws.merge_cells('B2:D2')

ws['E2'] = 'Invoice Date'
ws['E2'].font = label_font
ws['E2'].fill = label_fill
ws['E2'].alignment = label_alignment
ws['E2'].border = thin_border

ws['F2'] = '2026-01-21'
ws['F2'].font = content_font
ws['F2'].alignment = content_alignment
ws.merge_cells('F2:G2')

ws.row_dimensions[2].height = 22

# ============================================================
# Row 3: Country of Origin
# ============================================================
ws['A3'] = 'Country of Origin'
ws['A3'].font = label_font
ws['A3'].fill = label_fill
ws['A3'].alignment = label_alignment
ws['A3'].border = thin_border

ws['B3'] = 'CHINA'
ws['B3'].font = Font(name='Arial', size=10, bold=True)
ws['B3'].alignment = content_alignment
ws.merge_cells('B3:G3')

ws.row_dimensions[3].height = 22

# ============================================================
# Row 4: Shipper Information 标题
# ============================================================
ws['A4'] = 'Shipper Information (发货人)'
ws['A4'].font = section_font
ws['A4'].fill = section_fill
ws['A4'].alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('A4:G4')
ws.row_dimensions[4].height = 25

# ============================================================
# Row 5: Company
# ============================================================
ws['A5'] = 'Company\n公司'
ws['A5'].font = label_font
ws['A5'].fill = label_fill
ws['A5'].alignment = label_alignment
ws['A5'].border = thin_border

ws['B5'] = 'HONG KONG UNICORN TECHNOLOGY LIMITED'
ws['B5'].font = content_font
ws['B5'].alignment = content_alignment
ws.merge_cells('B5:G5')

ws.row_dimensions[5].height = 35

# ============================================================
# Row 6: Contact / TEL
# ============================================================
ws['A6'] = 'Contact\n联系人'
ws['A6'].font = label_font
ws['A6'].fill = label_fill
ws['A6'].alignment = label_alignment
ws['A6'].border = thin_border

ws['B6'] = 'Joy Kim'
ws['B6'].font = content_font
ws['B6'].alignment = content_alignment
ws.merge_cells('B6:D6')

ws['E6'] = 'TEL\n电话'
ws['E6'].font = label_font
ws['E6'].fill = label_fill
ws['E6'].alignment = label_alignment
ws['E6'].border = thin_border

ws['F6'] = '+86 177-2781-0735'
ws['F6'].font = content_font
ws['F6'].alignment = content_alignment
ws.merge_cells('F6:G6')

ws.row_dimensions[6].height = 35

# ============================================================
# Row 7: Address
# ============================================================
ws['A7'] = 'Address\n详细地址'
ws['A7'].font = label_font
ws['A7'].fill = label_fill
ws['A7'].alignment = label_alignment
ws['A7'].border = thin_border

ws['B7'] = 'Room 701, Unit 108B, 7/F, TowerB New Mandarin Plaza 14 Science Museum Road, Tsim Sha Tsui, Kowloon, Hong Kong'
ws['B7'].font = content_font
ws['B7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.merge_cells('B7:G7')

ws.row_dimensions[7].height = 45

# ============================================================
# Row 8: Consignee Information 标题
# ============================================================
ws['A8'] = 'Consignee Information (收货人)'
ws['A8'].font = section_font
ws['A8'].fill = section_fill
ws['A8'].alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('A8:G8')
ws.row_dimensions[8].height = 25

# ============================================================
# Row 9: Company / Contact (动态填充)
# ============================================================
ws['A9'] = 'Company\n公司'
ws['A9'].font = label_font
ws['A9'].fill = label_fill
ws['A9'].alignment = label_alignment
ws['A9'].border = thin_border

ws['B9'] = '[客户公司名]'  # 动态填充客户公司名
ws['B9'].font = content_font
ws['B9'].alignment = content_alignment
ws.merge_cells('B9:D9')

ws['E9'] = 'Contact\n联系人'
ws['E9'].font = label_font
ws['E9'].fill = label_fill
ws['E9'].alignment = label_alignment
ws['E9'].border = thin_border

ws['F9'] = '[联系人]'  # 动态填充联系人
ws['F9'].font = content_font
ws['F9'].alignment = content_alignment
ws.merge_cells('F9:G9')

ws.row_dimensions[9].height = 35

# ============================================================
# Row 10: Country / TEL (动态填充)
# ============================================================
ws['A10'] = 'Country\n目的国'
ws['A10'].font = label_font
ws['A10'].fill = label_fill
ws['A10'].alignment = label_alignment
ws['A10'].border = thin_border

ws['B10'] = '[国家]'  # 动态填充国家
ws['B10'].font = content_font
ws['B10'].alignment = content_alignment
ws.merge_cells('B10:D10')

ws['E10'] = 'TEL\n电话'
ws['E10'].font = label_font
ws['E10'].fill = label_fill
ws['E10'].alignment = label_alignment
ws['E10'].border = thin_border

ws['F10'] = '[电话]'  # 动态填充电话
ws['F10'].font = content_font
ws['F10'].alignment = content_alignment
ws.merge_cells('F10:G10')

ws.row_dimensions[10].height = 35

# ============================================================
# Row 11: Address (动态填充)
# ============================================================
ws['A11'] = 'Address\n详细地址'
ws['A11'].font = label_font
ws['A11'].fill = label_fill
ws['A11'].alignment = label_alignment
ws['A11'].border = thin_border

ws['B11'] = '[地址]'  # 动态填充地址
ws['B11'].font = content_font
ws['B11'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.merge_cells('B11:G11')

ws.row_dimensions[11].height = 45

# ============================================================
# Row 12: 表头
# ============================================================
headers = ['NO.', 'DESCRIPTION OF GOODS\n品类', 'Part No\n型号', 'HS Code\n海关编码', 'Qty (pcs)\n数量', 'Unit Price(KRW)\n单价', 'Total (KRW)\n总价']
for col, header in enumerate(headers, 1):
    cell = ws.cell(12, col)
    cell.value = header
    cell.font = header_font
    cell.alignment = header_alignment
    cell.fill = header_fill
    cell.border = thin_border

ws.row_dimensions[12].height = 40

# ============================================================
# Row 13-15: 数据行模板
# ============================================================
template_data = [
    [1, '集成电路/IC', 'TSX-3225 24.0000MF10Z-C3', '8542399000', 2000, 176, 352000],
    [2, '集成电路/IC', 'CL31A106KBHNNNE', '8542399000', 2000, 13, 26000],
    [3, '集成电路/IC', 'MAKK2016T3R3M', '8542399000', 3000, 33, 99000],
]

for row_idx, row_data in enumerate(template_data, 13):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        # 数字格式
        if col_idx in [5, 6, 7] and isinstance(value, (int, float)):
            cell.number_format = '#,##0'
    ws.row_dimensions[row_idx].height = 22

# ============================================================
# Row 16: Total Quantity
# ============================================================
ws['A16'] = ''
ws['D16'] = 'Total Quantity:'
ws['D16'].font = total_font
ws['D16'].alignment = Alignment(horizontal='right', vertical='center')
ws['D16'].border = thin_border
ws.merge_cells('A16:C16')

ws['E16'] = 7000
ws['E16'].font = total_font
ws['E16'].alignment = Alignment(horizontal='center', vertical='center')
ws['E16'].fill = total_fill
ws['E16'].border = thin_border
ws['E16'].number_format = '#,##0'

# F16, G16 空白但加边框
ws['F16'].border = thin_border
ws['G16'].border = thin_border

ws.row_dimensions[16].height = 25

# ============================================================
# Row 17: Total Invoice Amount
# ============================================================
ws['A17'] = 'Total invoice amount:'
ws['A17'].font = total_font
ws['A17'].alignment = Alignment(horizontal='right', vertical='center')
ws['A17'].fill = total_fill
ws['A17'].border = thin_border
ws.merge_cells('A17:F17')

ws['G17'] = 477000
ws['G17'].font = Font(name='Arial', size=10, bold=True, color='FF0000')
ws['G17'].alignment = Alignment(horizontal='center', vertical='center')
ws['G17'].fill = total_fill
ws['G17'].border = thin_border
ws['G17'].number_format = '#,##0'

ws.row_dimensions[17].height = 25

# ============================================================
# Row 18: Signature
# ============================================================
ws['A18'] = "SHIPPER'S SIGNATURE & STAMP:"
ws['A18'].font = Font(name='Arial', size=10, bold=True)
ws['A18'].alignment = Alignment(horizontal='left', vertical='top')
ws.merge_cells('A18:G18')
ws.row_dimensions[18].height = 100

# ============================================================
# 保存模板
# ============================================================
output_path = 'openclaw_skills/order-ci-generator-kr/template/CI_template.xlsx'
wb.save(output_path)
print(f'模板已创建: {output_path}')