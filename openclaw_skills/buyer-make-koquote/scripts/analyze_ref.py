import openpyxl
from openpyxl.utils import get_column_letter
import re

# Reference template
path = r'E:\1_业务\1_unicorn\1_贸易业务\1_报价\1_한국\1_TaeJu solusion\유니콘_전자부품견적서_{yyyymmddhhmm}.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

with open(r'e:\WorkPlace\7_AI_APP\UniUltraOpenPlatFormCls\openclaw_skills\buyer-make-koquote\scripts\analysis_output.txt', 'w', encoding='utf-8') as f:
    f.write("=== REF TEMPLATE ROW CONTENT R1-R40 ===\n")
    for r in range(1, 41):
        cells = []
        for c in range(1, 14):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = repr(v)
                if len(val_str) > 50:
                    val_str = val_str[:50] + "..."
                cells.append(get_column_letter(c) + "=" + val_str)
        h = ws.row_dimensions[r].height
        status = "EMPTY" if not cells else " | ".join(cells)
        f.write("  R%02d (h=%s): %s\n" % (r, h, status))

    f.write("\n=== CURRENT PROJECT TEMPLATE R1-R25 ===\n")
    wb2 = openpyxl.load_workbook(r'e:\WorkPlace\7_AI_APP\UniUltraOpenPlatFormCls\openclaw_skills\buyer-make-koquote\template\유니콘_전자부품견적서_template.xlsx')
    ws2 = wb2.active
    for r in range(1, 26):
        cells = []
        for c in range(1, 14):
            v = ws2.cell(r, c).value
            if v is not None:
                val_str = repr(v)
                if len(val_str) > 50:
                    val_str = val_str[:50] + "..."
                cells.append(get_column_letter(c) + "=" + val_str)
        h = ws2.row_dimensions[r].height
        status = "EMPTY" if not cells else " | ".join(cells)
        f.write("  R%02d (h=%s): %s\n" % (r, h, status))

    f.write("\n=== REF TEMPLATE MERGED (R15+) ===\n")
    for mc in sorted(str(m) for m in ws.merged_cells.ranges):
        nums = re.findall(r'\d+', mc)
        if nums and int(nums[0]) >= 15:
            f.write("  " + mc + "\n")

    f.write("\n=== CURRENT TEMPLATE MERGED (R15+) ===\n")
    for mc in sorted(str(m) for m in ws2.merged_cells.ranges):
        nums = re.findall(r'\d+', mc)
        if nums and int(nums[0]) >= 15:
            f.write("  " + mc + "\n")

    # Check if ref template has tail rows
    f.write("\n=== REF TEMPLATE - Check R37+ ===\n")
    max_r = ws.max_row
    f.write("  max_row = %d\n" % max_r)
    for r in range(37, min(max_r + 1, 45)):
        cells = []
        for c in range(1, 14):
            v = ws.cell(r, c).value
            if v is not None:
                val_str = repr(v)
                if len(val_str) > 50:
                    val_str = val_str[:50] + "..."
                cells.append(get_column_letter(c) + "=" + val_str)
        h = ws.row_dimensions[r].height
        status = "EMPTY" if not cells else " | ".join(cells)
        f.write("  R%02d (h=%s): %s\n" % (r, h, status))

print("Done, output written to analysis_output.txt")
