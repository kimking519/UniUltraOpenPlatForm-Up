from fastapi import FastAPI, Request, Form, Depends, HTTPException, Response, Cookie, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from contextlib import asynccontextmanager
import asyncio
import threading
from Sills.base import init_db, get_db_connection, get_exchange_rates
from Sills.db_daily import get_daily_list, add_daily, update_daily
from Sills.db_emp import get_emp_list, add_employee, batch_import_text, verify_login, change_password, update_employee, delete_employee
from Sills.db_vendor import add_vendor, batch_import_vendor_text, update_vendor, delete_vendor
from Sills.db_cli import get_cli_list, add_cli, batch_import_cli_text, update_cli, delete_cli
from Sills.db_quote import get_quote_list, add_quote, batch_import_quote_text, batch_import_quote_from_rows, delete_quote, update_quote, batch_delete_quote, batch_copy_quote, batch_add_quotes
from Sills.db_offer import get_offer_list, add_offer, batch_import_offer_text, update_offer, delete_offer, batch_delete_offer, batch_convert_from_quote
from Sills.db_order import get_order_list, add_order, update_order_status, update_order, delete_order, batch_import_order, batch_delete_order, batch_convert_from_offer, get_order_by_id
from Sills.db_buy import get_buy_list, add_buy, update_buy_node, update_buy, delete_buy, batch_import_buy, batch_delete_buy, batch_convert_from_order
from Sills.db_order_manager import (
    get_manager_list, get_manager_by_id, add_manager, update_manager, delete_manager,
    add_offer_to_manager, remove_offer_from_manager, get_manager_offers,
    get_available_offers_for_manager, batch_import_manager, batch_import_manager_from_rows, batch_delete_managers,
    add_attachment, get_attachments, delete_attachment
)
from Sills.db_mail import (
    get_mail_list, get_mail_by_id, save_email, delete_email,
    create_mail_relation, get_mail_relations, remove_mail_relation,
    get_mail_config, is_sync_locked, get_sync_progress,
    # 多账户管理
    get_all_mail_accounts, get_mail_account_by_id, add_mail_account,
    update_mail_account, switch_current_account, delete_mail_account,
    # 同步间隔设置
    get_sync_interval, set_sync_interval,
    # 文件夹管理
    get_folders, get_folder_by_id, add_folder, update_folder, delete_folder,
    get_mail_count_by_folder, get_mails_by_folder,
    # 过滤规则管理
    get_filter_rules, add_filter_rule, update_filter_rule, delete_filter_rule,
    # 自动分类
    auto_classify_emails, classify_mails
)
from Sills.mail_service import sync_inbox, sync_inbox_async, send_email_now
from Sills.ai_service import intent_recognizer, smart_replier
from Sills.db_config import is_postgresql, is_sqlite, get_pg_config, get_sqlite_path
from Sills.db_bank_transaction import (
    get_transaction_list, get_transaction_by_id, add_transaction,
    batch_import_transactions, update_transaction, delete_transaction,
    batch_delete_by_batch, batch_delete_selected, get_batch_list, update_matched_status
)
from Sills.db_bank_ledger import (
    get_ledger_by_transaction, get_ledger_by_manager, create_ledger,
    delete_ledger, update_ledger, set_primary_ledger, get_ledger_summary,
    validate_allocation_amount
)
from utils.price_engine import PriceEngine

from typing import Optional
import uvicorn
import shutil
import os
import platform
import io
import json
import base64
import urllib.request
import urllib.error
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl


def get_backup_root():
    """获取备份根目录"""
    is_windows = platform.system() == "Windows"
    return r"E:\WorkPlace\1_AIemployee\备份目录" if is_windows else "/home/kim/workspace/DbBackup"


def start_auto_backup():
    """启动自动备份定时任务"""
    import threading
    import time

    def backup_loop():
        while True:
            try:
                # 每24小时备份一次
                time.sleep(24 * 60 * 60)
                # 备份逻辑可以在这里添加
            except Exception as e:
                print(f"[Backup] Error: {e}")

    backup_thread = threading.Thread(target=backup_loop, daemon=True)
    backup_thread.start()


@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期管理"""
    # Startup
    init_db()
    start_auto_backup()
    yield
    # Shutdown (如果需要清理资源，可以在这里添加)


app = FastAPI(lifespan=lifespan)

# 内部服务API密钥（用于skill调用绕过认证）
INTERNAL_API_KEY = os.environ.get("INTERNAL_API_KEY", "dev-local-key")

# 开发模式：跳过认证校验（生产环境设为False）
SKIP_AUTH = os.environ.get("SKIP_AUTH", "true").lower() == "true"

# 自定义401异常处理器 - API请求返回JSON格式
@app.exception_handler(HTTPException)
async def custom_http_exception_handler(request: Request, exc: HTTPException):
    # API请求返回JSON格式错误
    if request.url.path.startswith('/api/') and exc.status_code == 401:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            {"success": False, "message": exc.detail},
            status_code=401
        )
    # 重定向类型的异常（301, 302, 303, 307, 308）需要返回RedirectResponse
    if exc.status_code in (301, 302, 303, 307, 308) and exc.headers and "Location" in exc.headers:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=exc.headers["Location"], status_code=exc.status_code)
    # 其他异常按默认处理
    raise exc

# Add session middleware
app.add_middleware(SessionMiddleware, secret_key="uni_platform_secret_key_2026")

# Mount static files and templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

def get_server_env():
    """检测服务器环境（包括WSL）"""
    system = platform.system()

    if system == "Linux":
        # 检测是否是WSL
        try:
            with open("/proc/version", "r") as f:
                version = f.read().lower()
                if "microsoft" in version or "wsl" in version:
                    return "WSL"
        except:
            pass
        return "Linux"
    elif system == "Windows":
        return "Windows"
    elif system == "Darwin":
        return "macOS"
    else:
        return system

def do_backup():
    """执行备份操作（内部函数，无权限检查）- 支持 SQLite 和 PostgreSQL"""
    import subprocess

    backup_root = get_backup_root()
    date_str = datetime.now().strftime("%Y%m%d")  # 只精确到天，每天覆盖
    backup_dir = os.path.join(backup_root, f"backup_{date_str}")

    # 确保备份根目录存在
    if not os.path.exists(backup_root):
        os.makedirs(backup_root, exist_ok=True)

    # If exists, delete and recreate
    if os.path.exists(backup_dir):
        shutil.rmtree(backup_dir)
    os.makedirs(backup_dir, exist_ok=True)

    project_root = os.path.dirname(os.path.abspath(__file__))
    backup_count = 0

    # 根据数据库类型选择备份方式
    if is_postgresql():
        # PostgreSQL 备份：使用 pg_dump
        pg_config = get_pg_config()
        dump_file = os.path.join(backup_dir, "uni_platform.sql")

        # 构建 pg_dump 命令
        cmd = [
            "pg_dump",
            "-h", pg_config['host'],
            "-p", str(pg_config['port']),
            "-U", pg_config['user'],
            "-d", pg_config['database'],
            "-F", "p",  # plain SQL format
            "-f", dump_file
        ]

        # 设置环境变量传递密码
        env = os.environ.copy()
        env["PGPASSWORD"] = pg_config['password']

        try:
            result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=300)
            if result.returncode == 0:
                backup_count = 1
                print(f"[备份] PostgreSQL 备份成功: {dump_file}")
            else:
                raise Exception(f"pg_dump 失败: {result.stderr}")
        except FileNotFoundError:
            raise Exception("pg_dump 命令未找到，请确保 PostgreSQL 客户端已安装并添加到 PATH")
    else:
        # SQLite 备份：复制文件
        db_files = [f for f in os.listdir(project_root) if f.endswith(".db")]
        for db_file in db_files:
            src = os.path.join(project_root, db_file)
            dst = os.path.join(backup_dir, db_file)
            shutil.copy2(src, dst)
        backup_count = len(db_files)

    # Copy static directory
    static_src = os.path.join(project_root, "static")
    if os.path.exists(static_src):
        static_dst = os.path.join(backup_dir, "static")
        shutil.copytree(static_src, static_dst, dirs_exist_ok=True)

    return backup_count, backup_dir

def cleanup_old_backups(backup_root, days=3):
    """清理超过指定天数的备份目录"""
    if not os.path.exists(backup_root):
        return 0

    deleted_count = 0
    cutoff_time = datetime.now().timestamp() - (days * 24 * 60 * 60)

    for item in os.listdir(backup_root):
        if item.startswith("backup_"):
            item_path = os.path.join(backup_root, item)
            if os.path.isdir(item_path):
                try:
                    # 获取目录修改时间
                    mtime = os.path.getmtime(item_path)
                    if mtime < cutoff_time:
                        shutil.rmtree(item_path)
                        deleted_count += 1
                        print(f"[备份清理] 已删除过期备份: {item}")
                except Exception as e:
                    print(f"[备份清理] 删除 {item} 失败: {str(e)}")

    return deleted_count

def auto_backup_task():
    """自动备份定时任务"""
    while True:
        try:
            count, backup_dir = do_backup()
            print(f"[自动备份] 成功备份 {count} 个数据库文件到 {backup_dir}")

            # 清理超过3天的备份
            backup_root = get_backup_root()
            deleted = cleanup_old_backups(backup_root, days=3)
            if deleted > 0:
                print(f"[自动备份] 已清理 {deleted} 个过期备份")
        except Exception as e:
            print(f"[自动备份] 失败: {str(e)}")
        # 每30分钟执行一次
        threading.Event().wait(1800)

def start_auto_backup():
    """启动自动备份线程"""
    backup_thread = threading.Thread(target=auto_backup_task, daemon=True)
    backup_thread.start()
    print("[系统] 自动备份服务已启动，每30分钟执行一次")
    # 注意：自动备份线程会立即执行第一次备份，不需要在这里重复执行

async def get_current_user(request: Request, emp_id: str = Cookie(None), rule: str = Cookie(None), account: str = Cookie(None)):
    # 0. 开发模式：跳过认证
    if SKIP_AUTH:
        return {"emp_id": "000", "rule": "3", "account": "Dev-Mode"}

    # 1. 检查内部服务API密钥（优先）- 用于skill调用
    internal_key = request.headers.get("X-Internal-API-Key")
    if internal_key == INTERNAL_API_KEY:
        return {"emp_id": "000", "rule": "3", "account": "Internal-Service"}

    # 2. 正常Cookie认证流程
    if not emp_id or not rule:
        return None
    return {"emp_id": emp_id, "rule": rule, "account": account}

async def login_required(request: Request, current_user: dict = Depends(get_current_user)):
    if not current_user:
        # 判断是否是API请求（路径以/api/开头）
        if request.url.path.startswith('/api/'):
            # API请求抛出401异常，由异常处理器返回JSON
            raise HTTPException(
                status_code=401,
                detail="登录已过期，请重新登录"
            )
        # 页面请求返回HTML重定向
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    return current_user

@app.get("/", response_class=HTMLResponse)
async def index(request: Request, current_user: dict = Depends(get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
        
    with get_db_connection() as conn:
        cli_count = conn.execute("SELECT COUNT(*) FROM uni_cli").fetchone()[0]
        emp_count = conn.execute("SELECT COUNT(*) FROM uni_emp").fetchone()[0]
        order_sum = conn.execute("SELECT COALESCE(SUM(paid_amount), 0) FROM uni_order").fetchone()[0]
        
    return templates.TemplateResponse("dashboard.html", {
        "request": request, 
        "active_page": "dashboard",
        "current_user": current_user,
        "stats": {
            "cli_count": cli_count,
            "emp_count": emp_count,
            "order_sum": order_sum
        }
    })

@app.get("/favicon.ico")
async def favicon():
    """返回空的 favicon 以避免 404 错误"""
    from fastapi.responses import Response
    # 返回空的 1x1 透明 ICO
    empty_ico = bytes([
        0x00, 0x00, 0x01, 0x00, 0x01, 0x00, 0x01, 0x01, 0x00, 0x00, 0x01, 0x00,
        0x18, 0x00, 0x30, 0x00, 0x00, 0x00, 0x16, 0x00, 0x00, 0x00, 0x28, 0x00,
        0x00, 0x00, 0x01, 0x00, 0x00, 0x00, 0x02, 0x00, 0x00, 0x00, 0x01, 0x00,
        0x18, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
        0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00,
        0x00, 0x00, 0xFF, 0xFF, 0xFF, 0x00, 0x00, 0x00
    ])
    return Response(content=empty_ico, media_type="image/x-icon")

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = "", account: str = ""):
    return templates.TemplateResponse("login.html", {"request": request, "error": error, "account": account})

@app.post("/login")
async def login(response: Response, account: str = Form(...), password: str = Form(...)):
    if account == "Admin" and password == "uni519":
        # System init backdoor, just in case
        response = RedirectResponse(url="/", status_code=303)
        response.set_cookie(key="emp_id", value="000")
        response.set_cookie(key="account", value="Admin")
        response.set_cookie(key="rule", value="3")
        return response

    ok, user, msg = verify_login(account, password)
    if not ok:
        # Redirect back to login with error message and retained account
        return RedirectResponse(url=f"/login?error={msg}&account={account}", status_code=303)
    
    # Check if first time login (password is default 12345)
    from Sills.db_emp import hash_password
    if user['password'] == hash_password('12345'):
        response = RedirectResponse(url="/change_password", status_code=303)
        response.set_cookie(key="emp_id", value=user['emp_id'])
        response.set_cookie(key="account", value=user['account'])
        response.set_cookie(key="rule", value=user['rule'])
        return response

    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(key="emp_id", value=str(user['emp_id']))
    response.set_cookie(key="account", value=str(user['account']))
    response.set_cookie(key="rule", value=str(user['rule']))
    return response

@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("emp_id")
    response.delete_cookie("rule")
    response.delete_cookie("account")
    return response

@app.get("/change_password", response_class=HTMLResponse)
async def change_pwd_page(request: Request, current_user: dict = Depends(get_current_user), error: str = ""):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("change_pwd.html", {"request": request, "current_user": current_user, "error": error})

@app.post("/change_password")
async def change_pwd_post(new_password: str = Form(...), confirm_password: str = Form(...), current_user: dict = Depends(get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    
    if new_password == '12345':
        return RedirectResponse(url="/change_password?error=新密码不能为12345", status_code=303)
    if new_password != confirm_password:
        return RedirectResponse(url="/change_password?error=两次输入的密码不一致", status_code=303)
        
    change_password(current_user['emp_id'], new_password)
    return RedirectResponse(url="/", status_code=303)

# Placeholder routes for modules 1-8
@app.get("/daily", response_class=HTMLResponse)
async def daily_page(request: Request, page: int = 1, current_user: dict = Depends(login_required)):
    items, total = get_daily_list(page=page)
    return templates.TemplateResponse("daily.html", {
        "request": request, 
        "active_page": "daily",
        "current_user": current_user,
        "items": items,
        "total": total,
        "page": page
    })

@app.post("/daily/add")
async def daily_add(currency_code: int = Form(...), exchange_rate: float = Form(...), current_user: dict = Depends(login_required)):
    from datetime import datetime
    record_date = datetime.now().strftime('%Y-%m-%d')
    success, msg = add_daily(record_date, currency_code, exchange_rate)
    return RedirectResponse(url="/daily", status_code=303)

@app.post("/api/daily/update")
async def daily_update_api(id: int = Form(...), exchange_rate: float = Form(...), current_user: dict = Depends(login_required)):
    success, msg = update_daily(id, exchange_rate)
    return {"success": success, "message": msg}

@app.get("/emp", response_class=HTMLResponse)
async def emp_page(request: Request, page: int = 1, search: str = "", current_user: dict = Depends(login_required)):
    items, total = get_emp_list(page=page, search=search)
    return templates.TemplateResponse("emp.html", {
        "request": request, 
        "active_page": "emp",
        "current_user": current_user,
        "items": items,
        "total": total,
        "page": page,
        "search": search
    })

@app.post("/emp/add")
async def emp_add(
    emp_name: str = Form(...), department: str = Form(""), position: str = Form(""),
    contact: str = Form(""), account: str = Form(...), hire_date: str = Form(...),
    rule: str = Form("1"), remark: str = Form(""),
    current_user: dict = Depends(login_required)
):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/emp", status_code=303)
        
    data = {
        "emp_name": emp_name, "department": department, "position": position,
        "contact": contact, "account": account, "password": "12345",
        "hire_date": hire_date,
        "rule": rule, "remark": remark
    }
    success, msg = add_employee(data)
    return RedirectResponse(url="/emp", status_code=303)

@app.post("/emp/import")
async def emp_import(import_text: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/emp", status_code=303)
    success_count, errors = batch_import_text(import_text)
    return RedirectResponse(url=f"/emp?import_success={success_count}&errors={len(errors)}", status_code=303)

@app.post("/emp/import/csv")
async def emp_import_csv(csv_file: UploadFile = File(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/emp", status_code=303)
    content = await csv_file.read()
    try:
        text = content.decode('utf-8-sig').strip()
    except UnicodeDecodeError:
        text = content.decode('gbk', errors='replace').strip()
        
    if '\n' in text:
        text = text.split('\n', 1)[1] # skip header
    success_count, errors = batch_import_text(text)
    return RedirectResponse(url=f"/emp?import_success={success_count}&errors={len(errors)}", status_code=303)

@app.post("/api/emp/update")
async def emp_update_api(emp_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限"}
    # Only allow certain fields
    allowed_fields = ['emp_name', 'account', 'password', 'department', 'position', 'rule', 'contact', 'hire_date', 'remark']
    if field not in allowed_fields:
        return {"success": False, "message": "非法字段"}
    
    success, msg = update_employee(emp_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/emp/delete")
async def emp_delete_api(emp_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限"}
    success, msg = delete_employee(emp_id)
    return {"success": success, "message": msg}

from Sills.base import get_paginated_list

@app.get("/vendor", response_class=HTMLResponse)
async def vendor_page(request: Request, page: int = 1, search: str = "", current_user: dict = Depends(login_required)):
    search_kwargs = {"vendor_name": search} if search else None
    result = get_paginated_list("uni_vendor", page=page, search_kwargs=search_kwargs)
    return templates.TemplateResponse("vendor.html", {
        "request": request, "active_page": "vendor", "current_user": current_user,
        "items": result["items"], "total_pages": result["total_pages"], 
        "page": page, "search": search, "active_page": "vendor"
    })

@app.post("/vendor/add")
async def vendor_add(
    vendor_name: str = Form(...), address: str = Form(""), qq: str = Form(""),
    wechat: str = Form(""), email: str = Form(""), remark: str = Form(""),
    current_user: dict = Depends(login_required)
):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/vendor", status_code=303)
    data = {
        "vendor_name": vendor_name, "address": address, "qq": qq,
        "wechat": wechat, "email": email, "remark": remark
    }
    add_vendor(data)
    return RedirectResponse(url="/vendor", status_code=303)

@app.post("/vendor/import")
async def vendor_import(import_text: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/vendor", status_code=303)
    success_count, errors = batch_import_vendor_text(import_text)
    return RedirectResponse(url=f"/vendor?import_success={success_count}&errors={len(errors)}", status_code=303)

@app.post("/vendor/import/csv")
async def vendor_import_csv(csv_file: UploadFile = File(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/vendor", status_code=303)
    content = await csv_file.read()
    try:
        text = content.decode('utf-8-sig').strip()
    except UnicodeDecodeError:
        text = content.decode('gbk', errors='replace').strip()
        
    if '\n' in text:
        text = text.split('\n', 1)[1] # skip header
    success_count, errors = batch_import_vendor_text(text)
    return RedirectResponse(url=f"/vendor?import_success={success_count}&errors={len(errors)}", status_code=303)

@app.post("/api/vendor/update")
async def vendor_update_api(vendor_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限"}
    allowed_fields = ['vendor_name', 'address', 'qq', 'wechat', 'email', 'remark']
    if field not in allowed_fields:
        return {"success": False, "message": "非法字段"}
    success, msg = update_vendor(vendor_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/vendor/delete")
async def vendor_delete_api(vendor_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限"}
    success, msg = delete_vendor(vendor_id)
    return {"success": success, "message": msg}

@app.get("/cli", response_class=HTMLResponse)
async def cli_page(request: Request, page: int = 1, page_size: int = 20, search: str = "", current_user: dict = Depends(login_required)):
    # 限制每页最多100条
    page_size = min(max(1, page_size), 100)
    search_kwargs = {"cli_name": search} if search else None
    result = get_paginated_list("uni_cli", page=page, page_size=page_size, search_kwargs=search_kwargs)

    # Needs employees for dropdown
    employees, _ = get_emp_list(page=1, page_size=1000)

    return templates.TemplateResponse("cli.html", {
        "request": request, "active_page": "cli", "current_user": current_user,
        "items": result["items"], "total_pages": result["total_pages"], "total_count": result["total_count"],
        "page": page, "page_size": page_size, "search": search, "employees": employees
    })

@app.post("/cli/add")
async def cli_add(
    cli_name: str = Form(...), cli_full_name: str = Form(""), cli_name_en: str = Form(""),
    contact_name: str = Form(""), address: str = Form(""),
    region: str = Form("韩国"), credit_level: str = Form("A"),
    margin_rate: float = Form(10.0), emp_id: str = Form(...), website: str = Form(""),
    payment_terms: str = Form(""), email: str = Form(""), phone: str = Form(""),
    remark: str = Form(""), current_user: dict = Depends(login_required)
):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/cli", status_code=303)
    data = {
        "cli_name": cli_name, "cli_full_name": cli_full_name, "cli_name_en": cli_name_en,
        "contact_name": contact_name, "address": address,
        "region": region, "credit_level": credit_level,
        "margin_rate": margin_rate, "emp_id": emp_id, "website": website,
        "payment_terms": payment_terms, "email": email, "phone": phone, "remark": remark
    }
    add_cli(data)
    return RedirectResponse(url="/cli", status_code=303)

@app.post("/cli/import")
async def cli_import(import_text: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/cli", status_code=303)
    success_count, errors = batch_import_cli_text(import_text)
    return RedirectResponse(url=f"/cli?import_success={success_count}&errors={len(errors)}", status_code=303)

@app.post("/cli/import/csv")
async def cli_import_csv(csv_file: UploadFile = File(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/cli", status_code=303)
    content = await csv_file.read()
    try:
        text = content.decode('utf-8-sig').strip()
    except UnicodeDecodeError:
        text = content.decode('gbk', errors='replace').strip()
        
    if '\n' in text:
        text = text.split('\n', 1)[1] # skip header
    success_count, errors = batch_import_cli_text(text)
    return RedirectResponse(url=f"/cli?import_success={success_count}&errors={len(errors)}", status_code=303)

@app.post("/api/order/update")
async def order_update_api(order_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无修改权限"}

    allowed_fields = ['order_no', 'order_date', 'inquiry_mpn', 'inquiry_brand', 'price_rmb', 'price_kwr', 'price_usd', 'cost_price_rmb', 'paid_amount', 'return_status', 'remark', 'is_transferred']
    if field not in allowed_fields:
        return {"success": False, "message": f"非法字段: {field}"}

    if field in ['price_rmb', 'price_kwr', 'price_usd', 'cost_price_rmb', 'paid_amount']:
        try:
            val = float(value)

            # 当修改price_rmb时，自动计算price_kwr和price_usd
            if field == 'price_rmb':
                from Sills.base import get_exchange_rates
                krw_rate, usd_rate, _ = get_exchange_rates()

                # 汇率含义: 1 RMB = X 外币
                # price_kwr = price_rmb * krw_rate (韩元)
                # price_usd = price_rmb * usd_rate (美元)
                price_kwr = round(val * krw_rate, 1) if krw_rate else 0
                price_usd = round(val * usd_rate, 3) if usd_rate else 0

                # 同时更新三个价格字段
                success, msg = update_order(order_id, {
                    'price_rmb': val,
                    'price_kwr': price_kwr,
                    'price_usd': price_usd
                })

                # 获取成本价计算利润
                order_info = get_order_by_id(order_id)
                cost_price = float(order_info.get('cost_price_rmb') or 0) if order_info else 0
                profit = round(val - cost_price, 3)

                return {"success": success, "message": msg, "price_kwr": price_kwr, "price_usd": price_usd, "profit": profit}

            # 当修改cost_price_rmb时，计算利润
            if field == 'cost_price_rmb':
                success, msg = update_order(order_id, {field: val})

                # 获取销售价计算利润
                order_info = get_order_by_id(order_id)
                price_rmb = float(order_info.get('price_rmb') or 0) if order_info else 0
                profit = round(price_rmb - val, 3)

                return {"success": success, "message": msg, "profit": profit}

            success, msg = update_order(order_id, {field: val})
            return {"success": success, "message": msg}
        except:
            return {"success": False, "message": "必须是数字"}

    success, msg = update_order(order_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/order/update_status")
async def order_update_status_api(order_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    from Sills.db_order import update_order_status
    success, msg = update_order_status(order_id, field, value)
    return {"success": success, "message": msg}

@app.post("/api/cli/update")
async def cli_update_api(cli_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限"}
    allowed_fields = ['cli_name', 'cli_full_name', 'cli_name_en', 'contact_name', 'address', 'region', 'credit_level', 'margin_rate', 'emp_id', 'website', 'payment_terms', 'email', 'phone', 'remark']
    if field not in allowed_fields:
        return {"success": False, "message": "非法字段"}
    
    if field == 'margin_rate':
        try: 
            val = float(value)
            success, msg = update_cli(cli_id, {field: val})
            return {"success": success, "message": msg}
        except: 
            return {"success": False, "message": "利润率必须是数字"}
        
    success, msg = update_cli(cli_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/cli/delete")
async def cli_delete_api(cli_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    success, msg = delete_cli(cli_id)
    return {"success": success, "message": msg}

@app.post("/api/cli/batch_delete")
async def cli_batch_delete_api(request: Request, current_user: dict = Depends(login_required)):
    """批量删除客户"""
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}

    from Sills.db_cli import batch_delete_cli
    data = await request.json()
    cli_ids = data.get('ids', [])

    if not cli_ids:
        return {"success": False, "message": "未选择记录"}

    deleted_count, failed_count, msg = batch_delete_cli(cli_ids)
    return {
        "success": True,
        "deleted_count": deleted_count,
        "failed_count": failed_count,
        "message": msg
    }

@app.get("/api/cli/list")
async def cli_list_api(current_user: dict = Depends(get_current_user)):
    """获取客户列表API（用于邮件关联选择器）"""
    if not current_user:
        return {"success": False, "message": "未登录", "items": []}
    items, total = get_cli_list(page=1, page_size=1000)
    return {"success": True, "items": items, "total": total}


@app.get("/api/cli/export")
async def cli_export_api(current_user: dict = Depends(login_required)):
    """导出客户数据到Excel"""
    from Sills.db_cli import export_cli_to_excel
    from urllib.parse import quote
    import io

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"cli_export_{timestamp}.xlsx"

    output = io.BytesIO()
    success, result = export_cli_to_excel(output)

    if not success:
        return {"success": False, "message": result}

    output.seek(0)
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/order/list")
async def order_list_api(current_user: dict = Depends(get_current_user)):
    """获取订单列表API（用于邮件关联选择器）"""
    if not current_user:
        return {"success": False, "message": "未登录", "items": []}
    items, total = get_order_list(page=1, page_size=1000)
    return {"success": True, "items": items, "total": total}

# ---------------- Quote Module ----------------
@app.get("/api/price/query")
async def api_price_query(
    mpn: str,
    current_user: dict = Depends(login_required)
):
    """全球比价查询"""
    if not mpn:
        return {"success": False, "message": "型号不能为空"}
    
    try:
        engine = PriceEngine()
        # 使用 run_in_executor 避免阻塞异步循环
        loop = asyncio.get_event_loop()
        results = await loop.run_in_executor(None, engine.check_all, mpn)
        
        return {
            "success": True,
            "mpn": mpn,
            "results": results
        }
    except Exception as e:
        return {"success": False, "message": f"查询失败: {str(e)}"}


@app.get("/quote", response_class=HTMLResponse)
async def quote_page(request: Request, current_user: dict = Depends(login_required), page: int = 1, page_size: int = 20, search: str = "", start_date: str = "", end_date: str = "", cli_id: str = "", status: str = "", is_transferred: str = ""):
    # 从 session 获取筛选条件
    session = request.session
    # 检查 URL 中是否有筛选参数（包括空值）
    has_params = any(k in request.query_params for k in ['search', 'start_date', 'end_date', 'cli_id', 'status', 'is_transferred'])

    if not has_params:
        # 首次访问或无参数，从 session 读取
        search = session.get("quote_search", "")
        start_date = session.get("quote_start_date", "")
        end_date = session.get("quote_end_date", "")
        cli_id = session.get("quote_cli_id", "")
        status = session.get("quote_status", "")
        is_transferred = session.get("quote_is_transferred", "")
    else:
        # 有参数时（包括空值），保存到 session
        session["quote_search"] = search
        session["quote_start_date"] = start_date
        session["quote_end_date"] = end_date
        session["quote_cli_id"] = cli_id
        session["quote_status"] = status
        session["quote_is_transferred"] = is_transferred

    results, total = get_quote_list(page=page, page_size=page_size, search_kw=search, start_date=start_date, end_date=end_date, cli_id=cli_id, status=status, is_transferred=is_transferred)
    total_pages = (total + page_size - 1) // page_size
    cli_list, _ = get_cli_list(page=1, page_size=1000)
    cli_list = sorted(cli_list, key=lambda x: x.get('cli_name', ''))
    return templates.TemplateResponse("quote.html", {
        "request": request,
        "active_page": "quote",
        "current_user": current_user,
        "items": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "search": search,
        "start_date": start_date,
        "end_date": end_date,
        "cli_id": cli_id,
        "status": status,
        "is_transferred": is_transferred,
        "cli_list": cli_list
    })

@app.post("/quote/add")
async def quote_add(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/quote", status_code=303)
    form = await request.form()
    data = dict(form)
    ok, msg = add_quote(data)
    import urllib.parse
    msg_param = urllib.parse.quote(msg)
    success = 1 if ok else 0
    return RedirectResponse(url=f"/quote?msg={msg_param}&success={success}", status_code=303)

@app.post("/quote/import")
async def quote_import_text(batch_text: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/quote", status_code=303)
    success_count, errors = batch_import_quote_text(batch_text)
    err_msg = ""
    if errors:
        import urllib.parse
        err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/quote?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.get("/api/quote/template")
async def api_quote_template(current_user: dict = Depends(get_current_user)):
    """下载需求导入模板 (Excel格式)"""
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "需求导入模板"

    # 设置表头
    headers = ["日期", "客户名", "询价型号", "报价型号", "询价品牌", "询价数量", "目标价", "成本价", "批号", "交期", "状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 设置示例数据（日期格式为文本，避免Excel自动转换）
    ws.cell(row=2, column=1, value="2026-01-01")
    ws.cell(row=2, column=2, value="示例客户")
    ws.cell(row=2, column=3, value="STM32F103C8T6")
    ws.cell(row=2, column=4, value="STM32F103C8T6")
    ws.cell(row=2, column=5, value="ST")
    ws.cell(row=2, column=6, value=500)
    ws.cell(row=2, column=7, value=8.5)
    ws.cell(row=2, column=8, value=7.0)
    ws.cell(row=2, column=9, value="2912+")
    ws.cell(row=2, column=10, value="1~3days")
    ws.cell(row=2, column=11, value="询价中")
    ws.cell(row=2, column=12, value="示例数据")

    ws.cell(row=3, column=1, value="2026-01-01")
    ws.cell(row=3, column=2, value="示例客户")
    ws.cell(row=3, column=3, value="CC0603KRX7R9BB104")
    ws.cell(row=3, column=4, value="CC0603KRX7R9BB104")
    ws.cell(row=3, column=5, value="YAGEO")
    ws.cell(row=3, column=6, value=2500)
    ws.cell(row=3, column=7, value=0.02)
    ws.cell(row=3, column=8, value=0.015)
    ws.cell(row=3, column=9, value="2912+")
    ws.cell(row=3, column=10, value="1~3days")
    ws.cell(row=3, column=11, value="询价中")
    ws.cell(row=3, column=12, value="示例数据")

    # 设置日期列格式为文本，防止Excel自动转换
    for row in range(2, 4):
        ws.cell(row=row, column=1).number_format = '@'

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 15

    # 输出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quote_template.xlsx"}
    )

@app.post("/quote/import/csv")
async def quote_import_csv(csv_file: UploadFile = File(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/quote", status_code=303)

    import openpyxl
    import io as io_module

    content = await csv_file.read()
    filename = csv_file.filename or ""
    rows_data = []

    # 判断是否为Excel文件
    if filename.lower().endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(io_module.BytesIO(content))
            ws = wb.active
            # 将Excel行转换为列表
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                rows_data.append(row_values)
        except Exception as e:
            return RedirectResponse(url=f"/quote?msg=Excel解析失败: {str(e)}&success=0", status_code=303)
    else:
        # CSV文件解析
        try:
            text = content.decode('utf-8-sig').strip()
        except UnicodeDecodeError:
            text = content.decode('gbk', errors='replace').strip()
        import csv
        f = io_module.StringIO(text)
        reader = csv.reader(f)
        rows_data = list(reader)

    # 调用导入函数
    success_count, errors = batch_import_quote_from_rows(rows_data)
    err_msg = ""
    if errors:
        import urllib.parse
        err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/quote?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.post("/api/quote/update")
async def quote_update_api(quote_id: str = Form(...), field: str = Form(...), value: str = Form(default=""), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无修改权限"}
        
    allowed_fields = ['cli_id', 'inquiry_mpn', 'quoted_mpn', 'inquiry_brand', 'inquiry_qty', 'actual_qty', 'target_price_rmb', 'cost_price_rmb', 'date_code', 'delivery_date', 'status', 'is_transferred', 'remark']
    if field not in allowed_fields:
        return {"success": False, "message": f"非法字段: {field}"}

    if field in ['inquiry_qty', 'actual_qty', 'target_price_rmb', 'cost_price_rmb']:
        try:
            val = float(value) if 'price' in field else int(value)
            success, msg = update_quote(quote_id, {field: val})
            return {"success": success, "message": msg}
        except:
            return {"success": False, "message": "必须是数字"}
            
    success, msg = update_quote(quote_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/quote/delete")
async def quote_delete_api(quote_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    success, msg = delete_quote(quote_id)
    return {"success": success, "message": msg}

@app.post("/api/quote/batch_delete")
async def quote_batch_delete_api(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    data = await request.json()
    ids = data.get("ids", [])
    success, msg = batch_delete_quote(ids)
    return {"success": success, "message": msg}

@app.post("/api/quote/batch_copy")
async def quote_batch_copy_api(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限复制需求"}
    data = await request.json()
    ids = data.get("ids", [])
    success, msg = batch_copy_quote(ids)
    return {"success": success, "message": msg}

@app.post("/api/quote/batch_add")
async def quote_batch_add_api(request: Request):
    """批量添加询价记录 - 供 skill 调用，使用内部 API Key 认证"""
    # 检查内部 API Key
    internal_key = request.headers.get("X-Internal-API-Key")
    if internal_key != INTERNAL_API_KEY:
        return {"success": False, "message": "无权限访问"}

    try:
        data = await request.json()
        items = data.get("items", [])

        if not items:
            return {"success": False, "message": "缺少 items 参数"}

        success_count, errors, created_ids = batch_add_quotes(items)

        return {
            "success": True,
            "message": f"成功添加 {success_count} 条记录",
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors,
            "created_ids": created_ids
        }
    except Exception as e:
        return {"success": False, "message": f"批量添加失败: {str(e)}"}

@app.get("/api/quote/info")
async def get_quote_info_api(id: str, current_user: dict = Depends(login_required)):
    from Sills.base import get_db_connection
    with get_db_connection() as conn:
        row = conn.execute("SELECT q.*, c.cli_name FROM uni_quote q LEFT JOIN uni_cli c ON q.cli_id = c.cli_id WHERE q.quote_id = ?", (id,)).fetchone()
        if row:
            return {"success": True, "data": dict(row)}
        return {"success": False, "message": "未找到"}

@app.post("/api/quote/export_offer_csv")
async def quote_export_offer_csv(request: Request, current_user: dict = Depends(login_required)):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        return {"success": False, "message": "未选择任何记录进行导出"}

    from Sills.base import get_db_connection
    placeholders = ','.join(['?'] * len(ids))
    with get_db_connection() as conn:
        quotes = conn.execute(f"""
            SELECT q.*, c.cli_name
            FROM uni_quote q
            LEFT JOIN uni_cli c ON q.cli_id = c.cli_id
            WHERE q.quote_id IN ({placeholders})
        """, ids).fetchall()

    import io, csv
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    # 字段顺序与页面显示一致
    writer.writerow(['需求日期','需求编号','客户名称','询价型号','报价型号','询价品牌','需求数量','目标价','成本价','批号','交期','状态','已转','备注'])

    for q in quotes:
        q_dict = dict(q)
        writer.writerow([
            q_dict.get('quote_date', ''),
            q_dict.get('quote_id', ''),
            q_dict.get('cli_name', ''),
            q_dict.get('inquiry_mpn', ''),
            q_dict.get('quoted_mpn', ''),
            q_dict.get('inquiry_brand', ''),
            q_dict.get('inquiry_qty', 0),
            q_dict.get('target_price_rmb', 0.0),
            q_dict.get('cost_price_rmb', 0.0),
            q_dict.get('date_code', ''),
            q_dict.get('delivery_date', ''),
            q_dict.get('status', ''),
            q_dict.get('is_transferred', ''),
            q_dict.get('remark', '')
        ])

    return {"success": True, "csv_content": output.getvalue()}

# ---------------- Offer Module ----------------
@app.get("/offer", response_class=HTMLResponse)
async def offer_page(request: Request, current_user: dict = Depends(login_required), page: int = 1, page_size: int = 20, search: str = "", start_date: str = "", end_date: str = "", cli_id: str = "", is_transferred: str = ""):
    # 从 session 获取筛选条件
    session = request.session
    # 检查 URL 中是否有筛选参数（包括空值）
    has_params = any(k in request.query_params for k in ['search', 'start_date', 'end_date', 'cli_id', 'is_transferred'])

    if not has_params:
        # 首次访问或无参数，从 session 读取
        search = session.get("offer_search", "")
        start_date = session.get("offer_start_date", "")
        end_date = session.get("offer_end_date", "")
        cli_id = session.get("offer_cli_id", "")
        is_transferred = session.get("offer_is_transferred", "未转")
        page_size = session.get("offer_page_size", 20)
    else:
        # 有参数（包括空值），保存到 session
        session["offer_search"] = search
        session["offer_start_date"] = start_date
        session["offer_end_date"] = end_date
        session["offer_cli_id"] = cli_id
        session["offer_is_transferred"] = is_transferred
        session["offer_page_size"] = page_size

    # is_transferred 空字符串表示"全部"，直接传递给查询层处理
    # 首次访问时 session 默认为"未转"，用户选择"全部"后 session 保存空字符串
    query_is_transferred = is_transferred
    results, total = get_offer_list(page=page, page_size=page_size, search_kw=search, start_date=start_date, end_date=end_date, cli_id=cli_id, is_transferred=query_is_transferred)
    total_pages = (total + page_size - 1) // page_size
    from Sills.base import get_paginated_list
    vendor_data = get_paginated_list('uni_vendor', page=1, page_size=1000)
    vendor_list = vendor_data['items']
    cli_data = get_paginated_list('uni_cli', page=1, page_size=1000)
    cli_list = sorted(cli_data['items'], key=lambda x: x.get('cli_name', ''))
    return templates.TemplateResponse("offer.html", {
        "request": request,
        "active_page": "offer",
        "current_user": current_user,
        "items": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "search": search,
        "start_date": start_date,
        "end_date": end_date,
        "cli_id": cli_id,
        "is_transferred": is_transferred,
        "vendor_list": vendor_list,
        "cli_list": cli_list
    })

@app.post("/offer/add")
async def offer_add_route(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/offer", status_code=303)
    form = await request.form()
    data = dict(form)
    emp_id = current_user['emp_id']

    # 自动报价逻辑：如果报价为 0 且指定了需求，则联动客户利润率
    if (not data.get('offer_price_rmb') or float(data.get('offer_price_rmb') or 0) == 0) and data.get('quote_id'):
        from Sills.base import get_db_connection
        with get_db_connection() as conn:
            clip = conn.execute("""
                SELECT c.margin_rate, q.cost_price_rmb
                FROM uni_quote q
                JOIN uni_cli c ON q.cli_id = c.cli_id
                WHERE q.quote_id = ?
            """, (data['quote_id'],)).fetchone()
            if clip and clip['cost_price_rmb']:
                margin = float(clip['margin_rate'] or 10.0)
                cost = float(clip['cost_price_rmb'])
                data['offer_price_rmb'] = round(cost * (1 + margin / 100.0), 4)
                if not data.get('cost_price_rmb') or float(data.get('cost_price_rmb') or 0) == 0:
                    data['cost_price_rmb'] = cost

    ok, msg = add_offer(data, emp_id)
    import urllib.parse
    msg_param = urllib.parse.quote(msg)
    success = 1 if ok else 0
    return RedirectResponse(url=f"/offer?msg={msg_param}&success={success}", status_code=303)

@app.post("/offer/import")
async def offer_import_text(batch_text: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/offer", status_code=303)
    success_count, errors = batch_import_offer_text(batch_text, current_user['emp_id'])
    err_msg = ""
    if errors:
        import urllib.parse
        err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/offer?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.post("/offer/import/csv")
async def offer_import_csv(csv_file: UploadFile = File(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/offer", status_code=303)
    content = await csv_file.read()
    try:
        text = content.decode('utf-8-sig').strip()
    except UnicodeDecodeError:
        text = content.decode('gbk', errors='replace').strip()
        
    # Pass full text to sill
    success_count, errors = batch_import_offer_text(text, current_user['emp_id'])
    err_msg = ""
    if errors:
        import urllib.parse
        err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/offer?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.get("/api/exchange/rates")
async def get_exchange_rates_api(current_user: dict = Depends(login_required)):
    """获取最新汇率（KRW、USD 和 JPY）"""
    from Sills.base import get_exchange_rates
    krw, usd, jpy = get_exchange_rates()
    return {"success": True, "krw": krw, "usd": usd, "jpy": jpy}

@app.get("/api/server/env")
async def get_server_env_api():
    """获取服务器环境信息"""
    return {"success": True, "env": get_server_env()}

@app.get("/api/server/memory")
async def get_server_memory_api():
    """获取服务器内存使用情况（支持WSL）"""
    try:
        import psutil

        # 当前进程内存
        process = psutil.Process()
        process_memory_mb = process.memory_info().rss / 1024 / 1024

        # 系统内存
        virtual_mem = psutil.virtual_memory()
        total_mb = virtual_mem.total / 1024 / 1024
        available_mb = virtual_mem.available / 1024 / 1024
        used_mb = virtual_mem.used / 1024 / 1024
        percent = virtual_mem.percent

        # WSL特殊处理：尝试获取WSL内存限制
        wsl_memory_limit = None
        server_env = get_server_env()
        if server_env == "WSL":
            try:
                # 尝试从/proc/meminfo获取MemTotal作为WSL限制
                with open("/proc/meminfo", "r") as f:
                    for line in f:
                        if line.startswith("MemTotal:"):
                            wsl_memory_limit = float(line.split()[1]) / 1024  # KB to MB
                            break
            except:
                pass

        result = {
            "success": True,
            "env": server_env,
            "process": {
                "memory_mb": round(process_memory_mb, 2),
                "name": process.name()
            },
            "system": {
                "total_mb": round(total_mb, 2),
                "available_mb": round(available_mb, 2),
                "used_mb": round(used_mb, 2),
                "percent": percent
            }
        }

        if wsl_memory_limit:
            result["wsl"] = {
                "memory_limit_mb": round(wsl_memory_limit, 2)
            }

        return result
    except ImportError:
        return {"success": False, "message": "psutil未安装，请运行: pip install psutil"}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/offer/update")
async def offer_update_api(offer_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无修改权限"}
        
    allowed_fields = ['quote_id', 'inquiry_mpn', 'quoted_mpn', 'inquiry_brand', 'quoted_brand',
                      'inquiry_qty', 'actual_qty', 'quoted_qty', 'cost_price_rmb', 'offer_price_rmb',
                      'price_kwr', 'price_usd', 'vendor_id', 'cli_id', 'date_code', 'delivery_date', 'offer_statement', 'is_transferred', 'remark', 'status', 'target_price_rmb']
    if field not in allowed_fields:
        return {"success": False, "message": f"非法字段: {field}"}
        
    if field in ['inquiry_qty', 'actual_qty', 'quoted_qty', 'cost_price_rmb', 'offer_price_rmb', 'price_kwr', 'price_usd']:
        try:
            val = float(value) if 'price' in field else int(value)
            success, msg = update_offer(offer_id, {field: val})
            return {"success": success, "message": msg}
        except:
            return {"success": False, "message": "必须是数字"}
            
    success, msg = update_offer(offer_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/offer/delete")
async def offer_delete_api(offer_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    success, msg = delete_offer(offer_id)
    return {"success": success, "message": msg}

@app.post("/api/offer/batch_delete")
async def offer_batch_delete_api(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    data = await request.json()
    ids = data.get("ids", [])
    success, msg = batch_delete_offer(ids)
    return {"success": success, "message": msg}

@app.post("/api/offer/batch_price_increase")
async def offer_batch_price_increase_api(request: Request, current_user: dict = Depends(login_required)):
    """按比例加价：新报价RMB = 成本价 × (1 + 比例%)，同时更新KWR和USD"""
    if current_user['rule'] != '3' and current_user['rule'] != '0':
        return {"success": False, "message": "无权限执行此操作"}

    data = await request.json()
    ids = data.get("ids", [])
    ratio = data.get("ratio", 15)  # 默认15%

    if not ids:
        return {"success": False, "message": "未选择任何记录"}

    from Sills.base import get_db_connection

    updated_count = 0
    with get_db_connection() as conn:
        # 获取最新汇率
        try:
            krw_row = conn.execute("SELECT exchange_rate FROM uni_daily WHERE currency_code=2 ORDER BY record_date DESC LIMIT 1").fetchone()
            usd_row = conn.execute("SELECT exchange_rate FROM uni_daily WHERE currency_code=1 ORDER BY record_date DESC LIMIT 1").fetchone()
            krw_rate = float(krw_row[0]) if krw_row else 180.0
            usd_rate = float(usd_row[0]) if usd_row else 7.0
        except:
            krw_rate = 180.0
            usd_rate = 7.0

        # 批量更新：新报价RMB = 成本价 × (1 + ratio/100)
        for offer_id in ids:
            # 获取当前记录的成本价
            row = conn.execute("SELECT cost_price_rmb FROM uni_offer WHERE offer_id=?", (offer_id,)).fetchone()
            if row and row[0]:
                cost_price = float(row[0])
                new_offer_rmb = cost_price * (1 + ratio / 100)
                new_price_kwr = round(new_offer_rmb * krw_rate, 1)  # KWR保留1位小数
                new_price_usd = round(new_offer_rmb * usd_rate, 3)  # USD保留3位小数

                conn.execute("""
                    UPDATE uni_offer
                    SET offer_price_rmb=?, price_kwr=?, price_usd=?
                    WHERE offer_id=?
                """, (new_offer_rmb, new_price_kwr, new_price_usd, offer_id))
                updated_count += 1

        conn.commit()

    return {"success": True, "updated_count": updated_count, "message": f"成功更新 {updated_count} 条报价"}

@app.post("/api/offer/batch_send_email")
async def offer_batch_send_email_api(request: Request, current_user: dict = Depends(login_required)):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        return {"success": False, "message": "未选择任何记录"}

    # 1. Fetch Data
    from Sills.base import get_db_connection
    placeholders = ','.join(['?'] * len(ids))
    with get_db_connection() as conn:
        # Get KRW rate for calculation
        try:
            rate_row = conn.execute("SELECT exchange_rate FROM uni_daily WHERE currency_code=2 ORDER BY record_date DESC LIMIT 1").fetchone()
            krw_rate = float(rate_row[0]) if rate_row else 180.0
        except: krw_rate = 180.0

        query = f"""
            SELECT o.*, v.vendor_name, c.cli_name
            FROM uni_offer o
            LEFT JOIN uni_vendor v ON o.vendor_id = v.vendor_id
            LEFT JOIN uni_quote q ON o.quote_id = q.quote_id
            LEFT JOIN uni_cli c ON q.cli_id = c.cli_id
            WHERE o.offer_id IN ({placeholders})
        """
        rows = conn.execute(query, ids).fetchall()
    
    if not rows:
        return {"success": False, "message": "记录不存在"}

    # 2. Generate Excel and Body
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Offers"
    # Headers exactly as requested: Model, Brand, QTY(pcs), Price(KWR), DC, L/T, Remark
    headers = ['Model', 'Brand', 'QTY(pcs)', 'Price(KWR)', 'DC', 'L/T', 'Remark']
    ws.append(headers)
    
    email_body_sections = []
    
    for row_data in rows:
        r = dict(row_data)
        # Handle KRW price if missing
        pkwr = r.get('price_kwr')
        if not pkwr:
            try: pkwr = round(float(r['offer_price_rmb'] or 0) * krw_rate, 1)
            except: pkwr = 0.0
            
        # Format for Excel
        ws.append([
            r['quoted_mpn'] or r['inquiry_mpn'], r['quoted_brand'] or r['inquiry_brand'],
            r['quoted_qty'], pkwr, r['date_code'], r['delivery_date'], r['remark']
        ])
        
        # Format for Email Body (Template requested)
        email_body_sections.append(f"""================
Model：{r['quoted_mpn'] or r['inquiry_mpn']}
Brand：{r['quoted_brand'] or r['inquiry_brand']}
Amount(pcs)：{r['quoted_qty']}
Price(KRW)：{pkwr}
DC:{r['date_code']}
LeadTime：{r['delivery_date']}
Remark: {r['remark']}
================ """ )

    full_body_for_email = "您好，这是批量导出的报价信息，请查收附件：\n\n" + "\n\n".join(email_body_sections)
    full_body_for_clipboard = "\n\n".join(email_body_sections)
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 3. MIME
    message = MIMEMultipart()
    message['To'] = 'joy@unicornsemi.com'
    message['Subject'] = f"批量报价汇总 - {len(rows)}条记录"
    message.attach(MIMEText(full_body_for_email, 'plain'))
    
    part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    part.set_payload(excel_file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="batch_offers_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"')
    message.attach(part)
    
    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
    
    # 4. Gmail Skill (Maton Gateway)
    maton_key = "AIzaSyCgZbWmrE266eSmCynikDsoddpt_ERCbvs"
    try:
        url = 'https://gateway.maton.ai/google-mail/gmail/v1/users/me/messages/send'
        payload = json.dumps({'raw': raw_message}).encode('utf-8')
        
        req = urllib.request.Request(
            url,
            data=payload,
            method='POST'
        )
        req.add_header('Authorization', f'Bearer {maton_key}')
        req.add_header('Content-Type', 'application/json')
        
        opener = urllib.request.build_opener()
        try:
            with opener.open(req, timeout=30) as response:
                resp_body = response.read().decode('utf-8')
                # Prepare base64 for frontend download
                excel_b64 = base64.b64encode(excel_file.getvalue()).decode('utf-8')
                return {
                    "success": True, 
                    "message": f"成功发送 {len(rows)} 条报价到 joy@unicornsemi.com",
                    "clipboard": full_body_for_clipboard,
                    "excel_b64": excel_b64,
                    "filename": f"batch_offers_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
                }
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            return {"success": False, "message": f"邮件发送失败 (HTTP {e.code}): {err_body}"}
    except Exception as e:
        return {"success": False, "message": f"邮件发送失败: {str(e)}"}

@app.post("/api/offer/export_csv")
async def offer_export_csv(request: Request, current_user: dict = Depends(login_required)):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids:
        return {"success": False, "message": "未选择任何记录进行导出"}

    from Sills.base import get_db_connection, get_exchange_rates
    placeholders = ','.join(['?'] * len(ids))
    with get_db_connection() as conn:
        query = f"""
            SELECT o.*, v.vendor_name, e.emp_name, c.cli_name
            FROM uni_offer o
            LEFT JOIN uni_vendor v ON o.vendor_id = v.vendor_id
            LEFT JOIN uni_emp e ON o.emp_id = e.emp_id
            LEFT JOIN uni_quote q ON o.quote_id = q.quote_id
            LEFT JOIN uni_cli c ON q.cli_id = c.cli_id
            WHERE o.offer_id IN ({placeholders})
        """
        rows = conn.execute(query, ids).fetchall()

    # 获取汇率
    krw_rate, usd_rate, _ = get_exchange_rates()

    # CSV 头部 - 与页面展示一致
    headers = ['日期', '报价编号', '客户名称', '需求编号', '询价型号', '报价型号', '需求品牌', '报价品牌',
               '需求数量(pcs)', '报价数量(pcs)', '成本价', '报价(RMB)', '报价(KWR)', '报价(USD)',
               '利润', '总利润', '供应商', '批号(DC)', '交期(LT)', '负责人', '已转', '备注']

    csv_lines = [','.join(headers)]

    for row_data in rows:
        r = dict(row_data)

        # 计算价格
        offer_price_rmb = float(r.get('offer_price_rmb') or 0)
        cost_price_rmb = float(r.get('cost_price_rmb') or 0)
        quoted_qty = int(r.get('quoted_qty') or 0)

        # KWR 价格
        pkwr = r.get('price_kwr')
        if not pkwr or float(pkwr) == 0:
            if krw_rate > 10:
                pkwr = round(offer_price_rmb * krw_rate, 1)
            else:
                pkwr = round(offer_price_rmb / krw_rate, 1) if krw_rate else 0

        # USD 价格 (USD汇率表示 1 RMB = ? USD，直接乘)
        pusd = r.get('price_usd')
        if not pusd or float(pusd) == 0:
            pusd = round(offer_price_rmb * usd_rate, 2) if usd_rate else 0

        # 利润计算
        profit = round(offer_price_rmb - cost_price_rmb, 3)
        total_profit = int(round(profit * quoted_qty, 0))

        # 构建CSV行
        line = [
            r.get('offer_date', ''),
            r.get('offer_id', ''),
            r.get('cli_name', ''),
            r.get('quote_id', ''),
            r.get('inquiry_mpn', ''),
            r.get('quoted_mpn', ''),
            r.get('inquiry_brand', ''),
            r.get('quoted_brand', ''),
            r.get('inquiry_qty', 0),
            r.get('quoted_qty', 0),
            cost_price_rmb,
            offer_price_rmb,
            pkwr,
            pusd,
            profit,
            total_profit,
            r.get('vendor_name', ''),
            r.get('date_code', ''),
            r.get('delivery_date', ''),
            r.get('emp_name', ''),
            r.get('is_transferred', '未转'),
            (r.get('remark') or '').replace('\n', ' ').replace(',', '，')
        ]
        csv_lines.append(','.join([str(v) for v in line]))

    csv_content = '\n'.join(csv_lines)

    # 生成剪贴板内容（韩元报价模板格式）
    clipboard_krw_sections = []
    clipboard_usd_sections = []
    for row_data in rows:
        r = dict(row_data)
        offer_price_rmb = float(r.get('offer_price_rmb') or 0)
        quoted_qty = int(r.get('quoted_qty') or 0)

        # KWR 价格
        pkwr = r.get('price_kwr')
        if not pkwr or float(pkwr) == 0:
            if krw_rate > 10:
                pkwr = round(offer_price_rmb * krw_rate, 1)
            else:
                pkwr = round(offer_price_rmb / krw_rate, 1) if krw_rate else 0

        # USD 价格
        pusd = r.get('price_usd')
        if not pusd or float(pusd) == 0:
            pusd = round(offer_price_rmb * usd_rate, 2) if usd_rate else 0

        # 韩元报价格式
        clipboard_krw_sections.append(f"""================
Model：{r.get('quoted_mpn') or r.get('inquiry_mpn')}
Brand：{r.get('quoted_brand') or r.get('inquiry_brand')}
Amount(pcs)：{r.get('quoted_qty')}
Price(KRW)：{pkwr}
DC：{r.get('date_code')}
LeadTime：{r.get('delivery_date')}
Remark: {r.get('remark')}
================ """)

        # 美元报价格式
        clipboard_usd_sections.append(f"""================
Model：{r.get('quoted_mpn') or r.get('inquiry_mpn')}
Brand：{r.get('quoted_brand') or r.get('inquiry_brand')}
Amount(pcs)：{r.get('quoted_qty')}
Price(USD)：{pusd}
DC：{r.get('date_code')}
LeadTime：{r.get('delivery_date')}
Remark: {r.get('remark')}
================ """)

    clipboard_krw_content = "\n\n".join(clipboard_krw_sections)
    clipboard_usd_content = "\n\n".join(clipboard_usd_sections)

    return {
        "success": True,
        "csv_content": csv_content,
        "clipboard": clipboard_krw_content,
        "clipboard_krw": clipboard_krw_content,
        "clipboard_usd": clipboard_usd_content,
        "filename": f"报价卡片_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    }

@app.post("/api/offer/generate_koquote")
async def offer_generate_koquote(request: Request, current_user: dict = Depends(login_required)):
    """生成韩文Excel报价单"""
    from Sills.document_generator import generate_koquote

    data = await request.json()
    offer_ids = data.get("offer_ids", [])
    if not offer_ids:
        return {"success": False, "message": "未选择任何报价记录"}

    try:
        success, result = generate_koquote(offer_ids)

        if success:
            return {
                "success": True,
                "file_path": result.get("excel_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}
    except subprocess.TimeoutExpired:
        return {"success": False, "message": "生成超时，请稍后重试"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}

@app.get("/order", response_class=HTMLResponse)
async def order_page(request: Request, current_user: dict = Depends(login_required), page: int = 1, page_size: int = 20, search: str = "", cli_id: str = "", start_date: str = "", end_date: str = "", is_finished: str = "", is_transferred: str = ""):
    # 日期默认不选择，保持为空
    # is_finished 为空表示"全部状态"，不做强制设置
    results, total = get_order_list(page=page, page_size=page_size, search_kw=search, cli_id=cli_id, start_date=start_date, end_date=end_date, is_finished=is_finished, is_transferred=is_transferred)
    total_pages = (total + page_size - 1) // page_size
    from Sills.db_cli import get_cli_list
    from Sills.base import get_paginated_list
    cli_list, _ = get_cli_list(page=1, page_size=1000)
    cli_list = sorted(cli_list, key=lambda x: x.get('cli_name', ''))
    vendor_data = get_paginated_list('uni_vendor', page=1, page_size=1000)
    vendor_list = vendor_data['items']
    return templates.TemplateResponse("order.html", {
        "request": request, "active_page": "order", "current_user": current_user,
        "items": results, "total": total, "page": page, "page_size": page_size,
        "total_pages": total_pages, "search": search, "cli_id": cli_id,
        "start_date": start_date, "end_date": end_date, "cli_list": cli_list,
        "vendor_list": vendor_list,
        "is_finished": is_finished,
        "is_transferred": request.query_params.get("is_transferred", "")
    })

@app.post("/order/add")
async def order_add_route(
    cli_id: str = Form(...), offer_id: str = Form(None), 
    order_id: str = Form(None), order_date: str = Form(None),
    inquiry_mpn: str = Form(None), inquiry_brand: str = Form(None),
    is_finished: int = Form(0), is_paid: int = Form(0), 
    paid_amount: float = Form(0.0), remark: str = Form(""),
    current_user: dict = Depends(login_required)
):
    data = {
        "cli_id": cli_id, "offer_id": offer_id, "order_id": order_id, "order_date": order_date,
        "inquiry_mpn": inquiry_mpn, "inquiry_brand": inquiry_brand,
        "is_finished": is_finished, "is_paid": is_paid, 
        "paid_amount": paid_amount, "remark": remark
    }
    ok, msg = add_order(data)
    import urllib.parse
    return RedirectResponse(url=f"/order?msg={urllib.parse.quote(msg)}&success={1 if ok else 0}", status_code=303)

@app.post("/order/import")
async def order_import_text(batch_text: str = Form(None), csv_file: UploadFile = File(None), cli_id: str = Form(...), current_user: dict = Depends(login_required)):
    if batch_text:
        text = batch_text
    elif csv_file:
        content = await csv_file.read()
        try:
            text = content.decode('utf-8-sig').strip()
        except UnicodeDecodeError:
            text = content.decode('gbk', errors='replace').strip()
    else:
        return RedirectResponse(url="/order?msg=未提供导入内容&success=0", status_code=303)
        
    success_count, errors = batch_import_order(text, cli_id)
    import urllib.parse
    err_msg = ""
    if errors: err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/order?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.post("/api/order/update_status")
async def api_order_update_status(order_id: str = Form(...), field: str = Form(...), value: int = Form(...), current_user: dict = Depends(login_required)):
    ok, msg = update_order_status(order_id, field, value)
    return {"success": ok, "message": msg}

@app.post("/api/order/update")
async def api_order_update(order_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if field in ['paid_amount']:
        try: value = float(value)
        except: return {"success": False, "message": "必须是数字"}
    
    allowed_fields = ['order_no', 'order_date', 'cli_id', 'offer_id', 'inquiry_mpn', 'inquiry_brand', 'price_rmb', 'price_kwr', 'price_usd', 'cost_price_rmb', 'is_finished', 'is_paid', 'paid_amount', 'return_status', 'remark', 'is_transferred']
    if field not in allowed_fields:
        return {"success": False, "message": f"非法字段: {field}"}

    ok, msg = update_order(order_id, {field: value})
    return {"success": ok, "message": msg}

@app.post("/api/order/delete")
async def api_order_delete(order_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3': return {"success": False, "message": "无权限"}
    ok, msg = delete_order(order_id)
    return {"success": ok, "message": msg}

@app.post("/api/order/batch_delete")
async def api_order_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3': return {"success": False, "message": "仅管理员可删除"}
    data = await request.json()
    ids = data.get("ids", [])
    ok, msg = batch_delete_order(ids)
    return {"success": ok, "message": msg}

@app.post("/api/order/export_csv")
async def order_export_csv(request: Request, current_user: dict = Depends(login_required)):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids: return {"success": False, "message": "未选择记录"}
    placeholders = ','.join(['?'] * len(ids))
    with get_db_connection() as conn:
        orders = conn.execute(f"""
            SELECT ord.*, c.cli_name,
                   off.inquiry_qty, off.quoted_qty, off.date_code, off.delivery_date,
                   v.vendor_name
            FROM uni_order ord
            LEFT JOIN uni_cli c ON ord.cli_id = c.cli_id
            LEFT JOIN uni_offer off ON ord.offer_id = off.offer_id
            LEFT JOIN uni_vendor v ON off.vendor_id = v.vendor_id
            WHERE ord.order_id IN ({placeholders})
        """, ids).fetchall()

    krw_rate, usd_rate, _ = get_exchange_rates()

    import io, csv
    output = io.StringIO(); output.write('\ufeff')
    writer = csv.writer(output)
    # 与页面显示字段一致
    writer.writerow(['订单日期','订单编号','客户','报价编号','报价型号','品牌','报价(RMB)','报价(KWR)','报价(USD)','成本(RMB)','利润','需求数量(pcs)','报价数量(pcs)','供应商','批号(DC)','货期','退货状态','完结','付款','已付金额','已转','备注'])

    for r in orders:
        d = dict(r)
        price_rmb = float(d.get('price_rmb') or 0)
        cost_rmb = float(d.get('cost_price_rmb') or 0)
        qty = int(d.get('quoted_qty') or 0)
        profit = round(price_rmb - cost_rmb, 3)

        # KWR/USD 计算
        price_kwr = d.get('price_kwr')
        price_usd = d.get('price_usd')
        if not price_kwr:
            if krw_rate > 10: price_kwr = round(price_rmb * krw_rate, 1)
            else: price_kwr = round(price_rmb / krw_rate, 1) if krw_rate else 0
        if not price_usd:
            price_usd = round(price_rmb * usd_rate, 2) if usd_rate else 0

        writer.writerow([
            d.get('order_date', ''),
            d.get('order_no') or d.get('order_id', ''),
            d.get('cli_name', ''),
            d.get('offer_id') or '',
            d.get('inquiry_mpn') or '',
            d.get('inquiry_brand') or '',
            f"{price_rmb:.2f}",
            price_kwr,
            price_usd,
            f"{cost_rmb:.2f}",
            profit,
            d.get('inquiry_qty') or '',
            d.get('quoted_qty') or '',
            d.get('vendor_name') or '',
            d.get('date_code') or '',
            d.get('delivery_date') or '',
            d.get('return_status', '正常'),
            '已完结' if d.get('is_finished') == 1 else '未完结',
            '已付款' if d.get('is_paid') == 1 else '未付款',
            d.get('paid_amount', 0),
            d.get('is_transferred', '未转'),
            d.get('remark') or ''
        ])
    return {"success": True, "csv_content": output.getvalue()}

# ============ 客户订单管理 ============

@app.get("/order_manager", response_class=HTMLResponse)
async def order_manager_page(request: Request, current_user: dict = Depends(login_required), page: int = 1, page_size: int = 20, search: str = "", cli_id: str = "", start_date: str = "", end_date: str = "", is_paid: str = "", is_finished: str = ""):
    results, total = get_manager_list(page=page, page_size=page_size, search_kw=search, cli_id=cli_id, start_date=start_date, end_date=end_date, is_paid=is_paid, is_finished=is_finished)
    total_pages = (total + page_size - 1) // page_size
    from Sills.db_cli import get_cli_list
    cli_list, _ = get_cli_list(page=1, page_size=1000)
    cli_list = sorted(cli_list, key=lambda x: x.get('cli_name', ''))
    return templates.TemplateResponse("order_manager.html", {
        "request": request, "active_page": "order_manager", "current_user": current_user,
        "items": results, "total": total, "page": page, "page_size": page_size,
        "total_pages": total_pages, "search": search, "cli_id": cli_id,
        "start_date": start_date, "end_date": end_date, "cli_list": cli_list,
        "is_paid": is_paid, "is_finished": is_finished
    })

# 历史订单导入页面
@app.get("/order_manager/import", response_class=HTMLResponse)
async def order_manager_import_page(request: Request, current_user: dict = Depends(login_required)):
    return templates.TemplateResponse("order_manager_import.html", {
        "request": request, "active_page": "order_manager", "current_user": current_user
    })


@app.get("/api/order_manager/import_template")
async def api_order_manager_import_template(current_user: dict = Depends(get_current_user)):
    """下载历史订单导入模板"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "历史订单导入模板"

    # 表头（19个字段）
    headers = [
        "日期", "交易编码", "客户订单号", "客户名称", "询价型号", "报价型号",
        "询价品牌", "报价品牌", "询价数量", "报价数量", "目标价(RMB)", "成本价(RMB)",
        "报价(RMB)", "报价(KRW)", "报价(USD)", "报价(JPY)", "批号", "交期", "备注"
    ]

    # 必填列标识（客户订单号, 客户名称, 询价型号, 报价数量）
    required_cols = [2, 3, 4, 9]

    # 样式
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 必填列标红背景
        if col - 1 in required_cols:
            for row in range(2, 6):
                ws.cell(row=row, column=col).fill = required_fill

    # 添加示例数据行
    sample_data = [
        ["2026-01-01", "TX001", "CO20260101001", "示例客户A", "STM32F103C8T6", "STM32F103C8T6", "ST", "ST", 100, 100, 12.0, 8.0, 10.5, 2345.0, 1.5, 160.0, "DC202601", "2026-01-15", "示例备注"],
        ["2026-01-01", "TX001", "CO20260101001", "示例客户A", "LM358DR", "LM358DR", "TI", "TI", 200, 200, 3.0, 1.8, 2.5, 560.0, 0.35, 38.0, "", "", ""],
        ["2026-01-02", "", "CO20260101002", "示例客户B", "NE555D", "NE555D", "TI", "TI", 50, 50, 1.5, 0.8, 1.2, 270.0, 0.17, 18.0, "DC2350", "2026-01-20", ""],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    col_widths = [12, 15, 12, 18, 18, 10, 10, 10, 10, 12, 12, 12, 10, 12, 15, 10]
    for i, width in enumerate(col_widths, 1):
        if i <= 26:
            ws.column_dimensions[chr(64 + i)].width = width
        else:
            ws.column_dimensions['A' + chr(64 + i - 26)].width = width

    # 输出到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    from urllib.parse import quote
    filename = "历史订单导入模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# 历史订单导入模板下载API
@app.get("/api/order_manager/history_template")
async def download_history_template(current_user: dict = Depends(login_required)):
    """下载历史订单导入模板"""
    from io import BytesIO
    from openpyxl import Workbook
    from urllib.parse import quote

    wb = Workbook()
    ws = wb.active
    ws.title = "历史订单导入模板"

    # 新模板字段列表（19个字段）
    headers = [
        "日期", "交易编码", "客户订单号", "客户名称", "询价型号", "报价型号",
        "询价品牌", "报价品牌", "询价数量", "报价数量", "目标价(RMB)", "成本价(RMB)",
        "报价(RMB)", "报价(KRW)", "报价(USD)", "报价(JPY)", "批号", "交期", "备注"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 添加示例数据
    example_data = [
        "2024-01-15", "TX001", "ORD001", "客户A", "MPN-001", "MPN-001-A",
        "BrandA", "BrandB", 100, 100, 0.7, 0.5,
        0.8, 1000.0, 0.08, 120.0, "DC202401", "2024-02-15", "备注信息"
    ]
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "历史订单导入模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# 历史订单导入API
@app.post("/api/order_manager/import_history")
async def api_import_history(
    file: UploadFile = File(...),
    current_user: dict = Depends(login_required)
):
    import tempfile
    import os
    from Sills.db_history_import import import_history_orders

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"success": 0, "skip": 0, "fail": 0, "errors": ["请上传Excel文件 (.xlsx 或 .xls)"]}

    # 保存上传的文件到临时目录
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # 调用导入函数
        emp_id = current_user.get('emp_id', '')
        success, skip, fail, errors = import_history_orders(tmp_path, emp_id)
        return {"success": success, "skip": skip, "fail": fail, "errors": errors}
    except Exception as e:
        return {"success": 0, "skip": 0, "fail": 0, "errors": [f"导入失败: {str(e)}"]}
    finally:
        # 删除临时文件
        try:
            os.unlink(tmp_path)
        except:
            pass

@app.get("/order_manager/{manager_id}", response_class=HTMLResponse)
async def order_manager_detail_page(request: Request, manager_id: str, current_user: dict = Depends(login_required)):
    manager = get_manager_by_id(manager_id)
    if not manager:
        return RedirectResponse(url="/order_manager?msg=订单不存在&success=0", status_code=303)
    offers = get_manager_offers(manager_id)
    attachments = get_attachments(manager_id)
    available_offers = get_available_offers_for_manager(cli_id=manager['cli_id'], manager_id=manager_id)
    # 获取可转采购的报价（排除已转的）
    from Sills.db_order_manager import get_manager_orders_for_purchase
    unpurchased_orders = get_manager_orders_for_purchase(manager_id)
    from Sills.db_cli import get_cli_list
    from Sills.db_vendor import get_vendor_list
    cli_list, _ = get_cli_list(page=1, page_size=1000)
    cli_list = sorted(cli_list, key=lambda x: x.get('cli_name', ''))
    vendors, _ = get_vendor_list(page=1, page_size=1000)
    return templates.TemplateResponse("order_manager_detail.html", {
        "request": request, "active_page": "order_manager", "current_user": current_user,
        "manager": manager, "orders": offers, "attachments": attachments,
        "available_orders": available_offers, "cli_list": cli_list, "vendors": vendors,
        "unpurchased_orders": unpurchased_orders
    })

@app.post("/api/order_manager/add")
async def api_order_manager_add(
    customer_order_no: str = Form(None), cli_id: str = Form(...), order_date: str = Form(None),
    shipping_fee: float = Form(0), tracking_no: str = Form(""), query_link: str = Form(""),
    mail_id: str = Form(""), mail_notes: str = Form(""), remark: str = Form(""),
    current_user: dict = Depends(login_required)
):
    data = {
        "customer_order_no": customer_order_no, "cli_id": cli_id, "order_date": order_date,
        "shipping_fee": shipping_fee, "tracking_no": tracking_no, "query_link": query_link,
        "mail_id": mail_id, "mail_notes": mail_notes, "remark": remark
    }
    ok, msg = add_manager(data)
    import urllib.parse
    if ok and isinstance(msg, dict):
        return RedirectResponse(url=f"/order_manager/{msg['manager_id']}?msg=创建成功&success=1", status_code=303)
    return RedirectResponse(url=f"/order_manager?msg={urllib.parse.quote(str(msg))}&success={1 if ok else 0}", status_code=303)

@app.post("/api/order_manager/update")
async def api_order_manager_update(manager_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if field in ['shipping_fee', 'paid_amount']:
        try: value = float(value)
        except: return {"success": False, "message": "必须是数字"}
    if field in ['is_paid', 'is_finished']:
        value = int(value)
    ok, msg = update_manager(manager_id, {field: value})
    return {"success": ok, "message": msg}

@app.post("/api/order_manager/delete")
async def api_order_manager_delete(manager_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3': return {"success": False, "message": "无权限"}
    ok, msg = delete_manager(manager_id)
    return {"success": ok, "message": msg}

@app.post("/api/order_manager/batch_delete")
async def api_order_manager_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3': return {"success": False, "message": "仅管理员可删除"}
    data = await request.json()
    ids = data.get("ids", [])
    ok, msg = batch_delete_managers(ids)
    return {"success": ok, "message": msg}

@app.post("/api/order_manager/batch_link_transactions")
async def api_order_manager_batch_link_transactions(request: Request, current_user: dict = Depends(login_required)):
    """批量补关联银行流水（根据交易编码）"""
    if current_user['rule'] not in ('2', '3'):
        return {"success": False, "message": "无权限执行此操作"}

    data = await request.json()
    manager_ids = data.get("manager_ids", [])

    if not manager_ids:
        return {"success": False, "message": "请先选择要补关联的订单"}

    from Sills.db_order_manager import batch_link_transactions
    result = batch_link_transactions(manager_ids, current_user.get('emp_id'))

    # 构建返回消息
    msg = f"成功关联 {result['linked_count']} 条"
    if result['skipped_count'] > 0:
        msg += f"，跳过 {result['skipped_count']} 条（无交易编码或已关联）"
    if result['failed_count'] > 0:
        msg += f"，失败 {result['failed_count']} 条"

    return {
        "success": result['linked_count'] > 0 or result['skipped_count'] > 0,
        "message": msg,
        "details": result['details'],
        "linked_count": result['linked_count'],
        "skipped_count": result['skipped_count'],
        "failed_count": result['failed_count']
    }

@app.get("/api/order_manager/template")
async def api_order_manager_template(current_user: dict = Depends(get_current_user)):
    """下载客户订单导入模板"""
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户订单导入模板"

    # 设置表头
    headers = ["日期", "客户名", "订单号", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 设置示例数据（日期格式为文本，避免Excel自动转换）
    ws.cell(row=2, column=1, value="2026-01-01")
    ws.cell(row=2, column=2, value="示例客户")
    ws.cell(row=2, column=3, value="CO20260101001")
    ws.cell(row=2, column=4, value="示例备注")

    ws.cell(row=3, column=1, value="2026-01-01")
    ws.cell(row=3, column=2, value="示例客户")
    ws.cell(row=3, column=3, value="CO20260101002")
    ws.cell(row=3, column=4, value="示例备注")

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_order_template.xlsx"}
    )

@app.post("/api/order_manager/batch_import")
async def api_order_manager_batch_import(batch_text: str = Form(None), csv_file: UploadFile = File(None), cli_id: str = Form(None), current_user: dict = Depends(login_required)):
    import openpyxl
    import io

    rows_data = []

    if batch_text:
        # 手动输入的文本，按CSV解析
        text = batch_text.strip()
        import csv
        f = io.StringIO(text)
        reader = csv.reader(f)
        rows_data = list(reader)
    elif csv_file:
        content = await csv_file.read()
        filename = csv_file.filename or ""

        # 判断是否为Excel文件
        if filename.lower().endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                # 将Excel行转换为列表
                for row in ws.iter_rows(values_only=True):
                    # 将None转换为空字符串，其他值转换为字符串
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    rows_data.append(row_values)
            except Exception as e:
                return RedirectResponse(url=f"/order_manager?msg=Excel解析失败: {str(e)}&success=0", status_code=303)
        else:
            # CSV文件解析
            try:
                text = content.decode('utf-8-sig').strip()
            except UnicodeDecodeError:
                text = content.decode('gbk', errors='replace').strip()
            import csv
            f = io.StringIO(text)
            reader = csv.reader(f)
            rows_data = list(reader)
    else:
        return RedirectResponse(url="/order_manager?msg=未提供导入内容&success=0", status_code=303)

    # 调用导入函数
    success_count, errors = batch_import_manager_from_rows(rows_data, cli_id)
    import urllib.parse
    err_msg = ""
    if errors: err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/order_manager?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.post("/api/order_manager/add_order")
async def api_order_manager_add_order(manager_id: str = Form(...), offer_id: str = Form(...), current_user: dict = Depends(login_required)):
    ok, msg = add_offer_to_manager(manager_id, offer_id)
    return {"success": ok, "message": msg}

@app.post("/api/order_manager/remove_order")
async def api_order_manager_remove_order(manager_id: str = Form(...), offer_id: str = Form(...), current_user: dict = Depends(login_required)):
    ok, msg = remove_offer_from_manager(manager_id, offer_id)
    return {"success": ok, "message": msg}

@app.get("/api/order_manager/list")
async def api_order_manager_list_api(page: int = 1, page_size: int = 20, search: str = "", cli_id: str = "", current_user: dict = Depends(login_required)):
    results, total = get_manager_list(page=page, page_size=page_size, search_kw=search, cli_id=cli_id)
    return {"items": results, "total": total, "page": page, "page_size": page_size}

@app.get("/api/order_manager/{manager_id}/orders")
async def api_order_manager_get_orders(manager_id: str, current_user: dict = Depends(login_required)):
    offers = get_manager_offers(manager_id)
    return {"orders": offers}

@app.post("/api/order_manager/{manager_id}/attachment")
async def api_order_manager_upload_attachment(manager_id: str, file: UploadFile = File(...), file_type: str = Form("其他"), current_user: dict = Depends(login_required)):
    import os
    from werkzeug.utils import secure_filename

    manager = get_manager_by_id(manager_id)
    if not manager:
        return {"success": False, "message": "客户订单不存在"}

    # 创建存储目录
    base_path = os.environ.get('ATTACHMENT_PATH', 'E:/1_Business/Attachments')
    customer_order_no = manager['customer_order_no']
    dir_path = os.path.join(base_path, customer_order_no)
    os.makedirs(dir_path, exist_ok=True)

    # 保存文件
    filename = secure_filename(file.filename)
    file_path = os.path.join(dir_path, filename)
    content = await file.read()
    with open(file_path, 'wb') as f:
        f.write(content)

    # 记录到数据库
    ok, msg = add_attachment(manager_id, file_path, file_type, filename)
    return {"success": ok, "message": msg, "file_path": file_path}

@app.get("/api/order_manager/{manager_id}/attachments")
async def api_order_manager_get_attachments(manager_id: str, current_user: dict = Depends(login_required)):
    attachments = get_attachments(manager_id)
    return {"attachments": attachments}

@app.delete("/api/order_manager/attachment/{attachment_id}")
async def api_order_manager_delete_attachment(attachment_id: int, current_user: dict = Depends(login_required)):
    ok, msg = delete_attachment(attachment_id)
    return {"success": ok, "message": msg}

@app.get("/api/order_manager/by_cli/{cli_id}")
async def api_order_manager_by_cli(cli_id: str, current_user: dict = Depends(login_required)):
    """获取指定客户的客户订单列表（用于报价转订单）"""
    from Sills.db_order_manager import get_manager_list_by_cli
    managers = get_manager_list_by_cli(cli_id)
    return managers

@app.post("/api/order_manager/batch_convert_offers")
async def api_order_manager_batch_convert_offers(request: Request, current_user: dict = Depends(login_required)):
    """批量将报价转入客户订单"""
    from Sills.db_order_manager import batch_convert_offers_to_manager
    data = await request.json()
    offer_ids = data.get('offer_ids', [])
    manager_id = data.get('manager_id')

    if not manager_id:
        return {"success": False, "message": "请选择目标客户订单"}

    ok, msg = batch_convert_offers_to_manager(offer_ids, manager_id)
    return {"success": ok, "message": msg}

@app.post("/api/order_manager/batch_to_purchase")
async def api_order_manager_batch_to_purchase(request: Request, current_user: dict = Depends(login_required)):
    """将客户订单中选中的报价订单转采购（先创建销售订单，再创建采购单）"""
    from Sills.db_order_manager import get_all_manager_orders_for_purchase
    from Sills.db_buy import add_buy
    from Sills.db_order import add_order
    from Sills.base import get_db_connection

    data = await request.json()
    manager_ids = data.get('manager_ids', [])
    selected_offer_ids = data.get('offer_ids', [])  # 选中的报价ID列表

    if not manager_ids:
        return {"success": False, "message": "请选择客户订单"}

    offers = get_all_manager_orders_for_purchase(manager_ids)
    if not offers:
        return {"success": False, "message": "客户订单中没有可转采购的报价订单"}

    # 如果有选中的报价ID，只处理选中的
    if selected_offer_ids:
        offers = [o for o in offers if o['offer_id'] in selected_offer_ids]

    if not offers:
        return {"success": False, "message": "请选择要转采购的报价订单"}

    success_count = 0
    errors = []

    for offer in offers:
        try:
            offer_id = offer['offer_id']

            # 1. 检查是否已有销售订单
            with get_db_connection() as conn:
                existing_order = conn.execute(
                    "SELECT order_id FROM uni_order WHERE offer_id = ?",
                    (offer_id,)
                ).fetchone()

                if existing_order:
                    order_id = existing_order['order_id']
                else:
                    # 2. 创建销售订单
                    order_date = datetime.now().strftime("%Y-%m-%d")

                    # 生成递增的5位数销售订单编号
                    last_order = conn.execute("SELECT order_id FROM uni_order WHERE order_id LIKE 'd%' ORDER BY order_id DESC LIMIT 1").fetchone()
                    if last_order:
                        try:
                            last_num = int(last_order['order_id'][1:])
                            new_num = last_num + 1
                        except:
                            new_num = 1
                    else:
                        new_num = 1
                    order_id = f"d{new_num:05d}"

                    conn.execute("""
                        INSERT INTO uni_order (
                            order_id, order_no, order_date, cli_id, offer_id,
                            inquiry_mpn, inquiry_brand, price_rmb, price_kwr, price_usd,
                            cost_price_rmb, is_finished, is_paid, paid_amount, return_status,
                            remark, is_transferred
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        order_id, order_id, order_date, offer['cli_id'], offer_id,
                        offer.get('inquiry_mpn'), '', offer.get('price_rmb', 0), 0, offer.get('price_usd', 0),
                        offer.get('cost_price_rmb', 0), 0, 0, 0.0, '正常',
                        '', '未转'
                    ))
                    conn.commit()

            # 3. 创建采购单
            buy_data = {
                'order_id': order_id,
                'buy_mpn': offer.get('inquiry_mpn'),
                'buy_brand': offer.get('buy_brand', ''),
                'vendor_id': offer.get('vendor_id'),
                'buy_price_rmb': offer.get('cost_price_rmb', 0),
                'buy_qty': offer.get('quoted_qty', 1),
                'sales_price_rmb': offer.get('price_rmb', 0),
            }
            ok, msg = add_buy(buy_data)
            if ok:
                success_count += 1
            else:
                errors.append(f"{offer_id}: {msg}")
        except Exception as e:
            errors.append(f"{offer.get('offer_id', 'unknown')}: {str(e)}")

    if success_count == 0 and errors:
        return {"success": False, "message": errors[0]}

    return {"success": True, "message": f"成功转采购 {success_count} 条" + (f" (失败 {len(errors)} 条)" if errors else "")}


@app.post("/api/order_manager/generate_pi_ci_kr")
async def api_order_manager_generate_pi_ci_kr(request: Request, current_user: dict = Depends(login_required)):
    """批量生成PI-CI-KR文件"""
    from Sills.db_order_manager import get_manager_by_id, get_manager_offers
    from Sills.document_generator import generate_pi_from_offers
    from Sills.ci_generator import generate_ci_kr_from_offers

    data = await request.json()
    manager_ids = data.get('manager_ids', [])

    if not manager_ids:
        return {"success": False, "message": "请选择客户订单"}

    # 创建输出目录
    from datetime import datetime
    import os
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_dir = os.path.join(r"E:\1_Business\1_unicorn\3_合规\审计资料2026\3_data\客户订单", f"PI-CI-KR-{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    success_count = 0
    errors = []
    generated_files = []

    for manager_id in manager_ids:
        try:
            # 获取客户订单信息
            manager = get_manager_by_id(manager_id)
            if not manager:
                errors.append(f"{manager_id}: 客户订单不存在")
                continue

            customer_order_no = manager.get('customer_order_no', 'UNKNOWN')
            cli_name = manager.get('cli_name', 'Unknown')

            # 获取关联的报价订单
            offers = get_manager_offers(manager_id)
            if not offers:
                errors.append(f"{customer_order_no}: 没有关联的报价订单")
                continue

            offer_ids = [o['offer_id'] for o in offers]

            # 生成 PI（传入customer_order_no作为invoice_no）
            pi_success, pi_result = generate_pi_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            if pi_success:
                # 重命名文件使用客户订单号
                old_path = pi_result.get('excel_path', '')
                if old_path and os.path.exists(old_path):
                    new_name = f"Proforma Invoice_{cli_name}_{customer_order_no}.xlsx"
                    new_path = os.path.join(output_dir, new_name)
                    os.rename(old_path, new_path)
                    generated_files.append(new_name)

                    # 重命名 PDF
                    old_pdf = pi_result.get('pdf_path', '')
                    if old_pdf and os.path.exists(old_pdf):
                        new_pdf_name = f"Proforma Invoice_{cli_name}_{customer_order_no}.pdf"
                        new_pdf_path = os.path.join(output_dir, new_pdf_name)
                        os.rename(old_pdf, new_pdf_path)
                        generated_files.append(new_pdf_name)
            else:
                errors.append(f"{customer_order_no} PI生成失败: {pi_result}")

            # 生成 CI（传入customer_order_no作为invoice_no）
            ci_success, ci_result = generate_ci_kr_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            if ci_success:
                # 重命名文件使用客户订单号
                old_path = ci_result.get('excel_path', '')
                if old_path and os.path.exists(old_path):
                    new_name = f"COMMERCIAL INVOICE_{cli_name}_{customer_order_no}.xlsx"
                    new_path = os.path.join(output_dir, new_name)
                    os.rename(old_path, new_path)
                    generated_files.append(new_name)

                    # 重命名 PDF
                    old_pdf = ci_result.get('pdf_path', '')
                    if old_pdf and os.path.exists(old_pdf):
                        new_pdf_name = f"COMMERCIAL INVOICE_{cli_name}_{customer_order_no}.pdf"
                        new_pdf_path = os.path.join(output_dir, new_pdf_name)
                        os.rename(old_pdf, new_pdf_path)
                        generated_files.append(new_pdf_name)
            else:
                errors.append(f"{customer_order_no} CI生成失败: {ci_result}")

            success_count += 1

        except Exception as e:
            errors.append(f"{manager_id}: {str(e)}")

    return {
        "success": success_count > 0,
        "message": f"成功生成 {success_count} 个客户订单的PI-CI文件" + (f" (失败 {len(errors)} 条)" if errors else ""),
        "output_dir": output_dir,
        "generated_files": generated_files,
        "errors": errors
    }


@app.post("/api/order_manager/generate_pi_ci_us")
async def api_order_manager_generate_pi_ci_us(request: Request, current_user: dict = Depends(login_required)):
    """批量生成PI-CI-US文件"""
    from Sills.db_order_manager import get_manager_by_id, get_manager_offers
    from Sills.document_generator import generate_pi_us_from_offers, generate_ci_us_from_offers

    data = await request.json()
    manager_ids = data.get('manager_ids', [])

    if not manager_ids:
        return {"success": False, "message": "请选择客户订单"}

    # 创建输出目录
    from datetime import datetime
    import os
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_dir = os.path.join(r"E:\1_Business\1_unicorn\3_合规\审计资料2026\3_data\客户订单", f"PI-CI-US-{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    success_count = 0
    errors = []
    generated_files = []

    for manager_id in manager_ids:
        try:
            # 获取客户订单信息
            manager = get_manager_by_id(manager_id)
            if not manager:
                errors.append(f"{manager_id}: 客户订单不存在")
                continue

            customer_order_no = manager.get('customer_order_no', 'UNKNOWN')
            cli_name = manager.get('cli_name', 'Unknown')

            # 获取关联的报价订单
            offers = get_manager_offers(manager_id)
            if not offers:
                errors.append(f"{customer_order_no}: 没有关联的报价订单")
                continue

            offer_ids = [o['offer_id'] for o in offers]

            # 生成 PI-US（传入customer_order_no作为invoice_no）
            pi_success, pi_result = generate_pi_us_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            if pi_success:
                # 重命名文件使用客户订单号
                old_path = pi_result.get('excel_path', '')
                if old_path and os.path.exists(old_path):
                    new_name = f"Proforma Invoice_{cli_name}_{customer_order_no}.xlsx"
                    new_path = os.path.join(output_dir, new_name)
                    os.rename(old_path, new_path)
                    generated_files.append(new_name)

                    # 重命名 PDF
                    old_pdf = pi_result.get('pdf_path', '')
                    if old_pdf and os.path.exists(old_pdf):
                        new_pdf_name = f"Proforma Invoice_{cli_name}_{customer_order_no}.pdf"
                        new_pdf_path = os.path.join(output_dir, new_pdf_name)
                        os.rename(old_pdf, new_pdf_path)
                        generated_files.append(new_pdf_name)
            else:
                errors.append(f"{customer_order_no} PI生成失败: {pi_result}")

            # 生成 CI-US（传入customer_order_no作为invoice_no）
            ci_success, ci_result = generate_ci_us_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            if ci_success:
                # 重命名文件使用客户订单号
                old_path = ci_result.get('excel_path', '')
                if old_path and os.path.exists(old_path):
                    new_name = f"COMMERCIAL INVOICE_{cli_name}_{customer_order_no}.xlsx"
                    new_path = os.path.join(output_dir, new_name)
                    os.rename(old_path, new_path)
                    generated_files.append(new_name)

                    # 重命名 PDF
                    old_pdf = ci_result.get('pdf_path', '')
                    if old_pdf and os.path.exists(old_pdf):
                        new_pdf_name = f"COMMERCIAL INVOICE_{cli_name}_{customer_order_no}.pdf"
                        new_pdf_path = os.path.join(output_dir, new_pdf_name)
                        os.rename(old_pdf, new_pdf_path)
                        generated_files.append(new_pdf_name)
            else:
                errors.append(f"{customer_order_no} CI生成失败: {ci_result}")

            success_count += 1

        except Exception as e:
            errors.append(f"{manager_id}: {str(e)}")

    return {
        "success": success_count > 0,
        "message": f"成功生成 {success_count} 个客户订单的PI-CI文件" + (f" (失败 {len(errors)} 条)" if errors else ""),
        "output_dir": output_dir,
        "generated_files": generated_files,
        "errors": errors
    }


@app.post("/api/order_manager/generate_pi")
async def api_order_manager_generate_pi(request: Request, current_user: dict = Depends(login_required)):
    """批量生成PI文件（智能判断币种）

    判断逻辑：
    1. 首先检查客户订单汇总字段：total_price_kwr, total_price_usd, total_price_jpy
    2. 如果汇总字段都为零，检查报价单价格字段：price_kwr, price_usd, price_jpy
    3. 取首个非零值决定币种，优先级：KRW > USD > JPY
    """
    from Sills.db_order_manager import get_manager_by_id, get_manager_offers
    from Sills.document_generator import generate_pi_from_offers, generate_pi_us_from_offers, generate_pi_jp_from_offers

    data = await request.json()
    manager_ids = data.get('manager_ids', [])

    if not manager_ids:
        return {"success": False, "message": "请选择客户订单"}

    # 创建输出目录
    from datetime import datetime
    import os
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_dir = os.path.join(r"E:\1_Business\1_unicorn\3_合规\审计资料2026\3_data\客户订单", f"PI-{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    success_count = 0
    errors = []
    generated_files = []

    for manager_id in manager_ids:
        try:
            manager = get_manager_by_id(manager_id)
            if not manager:
                errors.append(f"{manager_id}: 客户订单不存在")
                continue

            customer_order_no = manager.get('customer_order_no', 'UNKNOWN')
            cli_name = manager.get('cli_name', 'Unknown')

            offers = get_manager_offers(manager_id)
            if not offers:
                errors.append(f"{customer_order_no}: 没有关联的报价订单")
                continue

            offer_ids = [o['offer_id'] for o in offers]

            # 判断币种：优先使用客户订单汇总字段
            currency_type = None
            total_kwr = float(manager.get('total_price_kwr') or 0)
            total_usd = float(manager.get('total_price_usd') or 0)
            total_jpy = float(manager.get('total_price_jpy') or 0)

            if total_kwr > 0:
                currency_type = "KRW"
            elif total_usd > 0:
                currency_type = "USD"
            elif total_jpy > 0:
                currency_type = "JPY"

            # 如果汇总字段都为零，检查报价单价格字段
            if not currency_type:
                for offer in offers:
                    price_kwr = float(offer.get('price_kwr') or 0)
                    price_usd = float(offer.get('price_usd') or 0)
                    price_jpy = float(offer.get('price_jpy') or 0)

                    if price_kwr > 0:
                        currency_type = "KRW"
                        break
                    elif price_usd > 0:
                        currency_type = "USD"
                        break
                    elif price_jpy > 0:
                        currency_type = "JPY"
                        break

            # 默认使用 KRW
            if not currency_type:
                currency_type = "KRW"

            # 根据币种生成 PI（传入customer_order_no作为invoice_no）
            if currency_type == "KRW":
                pi_success, pi_result = generate_pi_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            elif currency_type == "USD":
                pi_success, pi_result = generate_pi_us_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            elif currency_type == "JPY":
                pi_success, pi_result = generate_pi_jp_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)

            if pi_success:
                old_path = pi_result.get('excel_path', '')
                if old_path and os.path.exists(old_path):
                    # 文件名包含币种标识
                    new_name = f"Proforma Invoice_{cli_name}_{customer_order_no}_{currency_type}.xlsx"
                    new_path = os.path.join(output_dir, new_name)
                    os.rename(old_path, new_path)
                    generated_files.append(new_name)
                success_count += 1
            else:
                errors.append(f"{customer_order_no} PI生成失败: {pi_result}")

        except Exception as e:
            errors.append(f"{manager_id}: {str(e)}")

    return {
        "success": success_count > 0,
        "message": f"成功生成 {success_count} 个PI文件" + (f" (失败 {len(errors)} 条)" if errors else ""),
        "output_dir": output_dir,
        "generated_files": generated_files,
        "errors": errors
    }


@app.post("/api/order_manager/generate_ci")
async def api_order_manager_generate_ci(request: Request, current_user: dict = Depends(login_required)):
    """批量生成CI文件（智能判断币种）

    判断逻辑：
    1. 首先检查客户订单汇总字段：total_price_kwr, total_price_usd, total_price_jpy
    2. 如果汇总字段都为零，检查报价单价格字段：price_kwr, price_usd, price_jpy
    3. 取首个非零值决定币种，优先级：KRW > USD > JPY
    """
    from Sills.db_order_manager import get_manager_by_id, get_manager_offers
    from Sills.ci_generator import generate_ci_kr_from_offers, generate_ci_jp_from_offers
    from Sills.document_generator import generate_ci_us_from_offers

    data = await request.json()
    manager_ids = data.get('manager_ids', [])

    if not manager_ids:
        return {"success": False, "message": "请选择客户订单"}

    # 创建输出目录
    from datetime import datetime
    import os
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_dir = os.path.join(r"E:\1_Business\1_unicorn\3_合规\审计资料2026\3_data\客户订单", f"CI-{timestamp}")
    os.makedirs(output_dir, exist_ok=True)

    success_count = 0
    errors = []
    generated_files = []

    for manager_id in manager_ids:
        try:
            manager = get_manager_by_id(manager_id)
            if not manager:
                errors.append(f"{manager_id}: 客户订单不存在")
                continue

            customer_order_no = manager.get('customer_order_no', 'UNKNOWN')
            cli_name = manager.get('cli_name', 'Unknown')

            offers = get_manager_offers(manager_id)
            if not offers:
                errors.append(f"{customer_order_no}: 没有关联的报价订单")
                continue

            offer_ids = [o['offer_id'] for o in offers]

            # 判断币种：优先使用客户订单汇总字段
            currency_type = None
            total_kwr = float(manager.get('total_price_kwr') or 0)
            total_usd = float(manager.get('total_price_usd') or 0)
            total_jpy = float(manager.get('total_price_jpy') or 0)

            if total_kwr > 0:
                currency_type = "KRW"
            elif total_usd > 0:
                currency_type = "USD"
            elif total_jpy > 0:
                currency_type = "JPY"

            # 如果汇总字段都为零，检查报价单价格字段
            if not currency_type:
                for offer in offers:
                    price_kwr = float(offer.get('price_kwr') or 0)
                    price_usd = float(offer.get('price_usd') or 0)
                    price_jpy = float(offer.get('price_jpy') or 0)

                    if price_kwr > 0:
                        currency_type = "KRW"
                        break
                    elif price_usd > 0:
                        currency_type = "USD"
                        break
                    elif price_jpy > 0:
                        currency_type = "JPY"
                        break

            # 默认使用 KRW
            if not currency_type:
                currency_type = "KRW"

            # 根据币种生成 CI（传入customer_order_no作为invoice_no）
            if currency_type == "KRW":
                ci_success, ci_result = generate_ci_kr_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            elif currency_type == "USD":
                ci_success, ci_result = generate_ci_us_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)
            elif currency_type == "JPY":
                ci_success, ci_result = generate_ci_jp_from_offers(offer_ids, output_base=output_dir, invoice_no=customer_order_no)

            if ci_success:
                old_path = ci_result.get('excel_path', '')
                if old_path and os.path.exists(old_path):
                    # 文件名包含币种标识
                    new_name = f"COMMERCIAL INVOICE_{cli_name}_{customer_order_no}_{currency_type}.xlsx"
                    new_path = os.path.join(output_dir, new_name)
                    os.rename(old_path, new_path)
                    generated_files.append(new_name)
                success_count += 1
            else:
                errors.append(f"{customer_order_no} CI生成失败: {ci_result}")

        except Exception as e:
            errors.append(f"{manager_id}: {str(e)}")

    return {
        "success": success_count > 0,
        "message": f"成功生成 {success_count} 个CI文件" + (f" (失败 {len(errors)} 条)" if errors else ""),
        "output_dir": output_dir,
        "generated_files": generated_files,
        "errors": errors
    }


# ============ 采购管理 ============

@app.get("/buy", response_class=HTMLResponse)
async def buy_page(request: Request, current_user: dict = Depends(login_required), page: int = 1, page_size: int = 20, search: str = "", order_id: str = "", start_date: str = "", end_date: str = "", cli_id: str = "", is_shipped: str = ""):
    # 日期默认不选择，保持为空
    # is_shipped 为空表示"全部状态"，不做强制设置
    results, total = get_buy_list(page=page, page_size=page_size, search_kw=search, order_id=order_id, start_date=start_date, end_date=end_date, cli_id=cli_id, is_shipped=is_shipped)
    total_pages = (total + page_size - 1) // page_size
    with get_db_connection() as conn:
        vendors = conn.execute("SELECT vendor_id, vendor_name, address FROM uni_vendor").fetchall()
        orders = conn.execute("SELECT order_id, order_no FROM uni_order").fetchall()
        clis = conn.execute("SELECT cli_id, cli_name FROM uni_cli ORDER BY cli_name").fetchall()
        vendor_addresses = {str(v['vendor_id']): (v['address'] or "") for v in vendors}
    return templates.TemplateResponse("buy.html", {
        "request": request, "active_page": "buy", "current_user": current_user,
        "items": results, "total": total, "page": page, "page_size": page_size,
        "total_pages": total_pages, "search": search, "order_id": order_id,
        "start_date": start_date, "end_date": end_date, "cli_id": cli_id,
        "vendor_list": vendors, "order_list": orders, "cli_list": clis,
        "is_shipped": is_shipped, "vendor_addresses": vendor_addresses
    })

@app.post("/buy/import")
async def buy_import_text(batch_text: str = Form(None), csv_file: UploadFile = File(None), current_user: dict = Depends(login_required)):
    if batch_text:
        text = batch_text
    elif csv_file:
        content = await csv_file.read()
        try:
            text = content.decode('utf-8-sig').strip()
        except UnicodeDecodeError:
            text = content.decode('gbk', errors='replace').strip()
    else:
        return RedirectResponse(url="/buy?import_success=0&errors=1&msg=未提供导入内容", status_code=303)
        
    success_count, errors = batch_import_buy(text)
    import urllib.parse
    err_msg = ""
    if errors: err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/buy?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)

@app.post("/buy/add")
async def buy_add_route(
    order_id: str = Form(...), vendor_id: str = Form(...),
    buy_mpn: str = Form(...), buy_brand: str = Form(""),
    buy_price_rmb: float = Form(...), buy_qty: int = Form(...),
    sales_price_rmb: float = Form(0.0), remark: str = Form(""),
    current_user: dict = Depends(login_required)
):
    data = {
        "order_id": order_id, "vendor_id": vendor_id,
        "buy_mpn": buy_mpn, "buy_brand": buy_brand,
        "buy_price_rmb": buy_price_rmb, "buy_qty": buy_qty,
        "sales_price_rmb": sales_price_rmb, "remark": remark
    }
    ok, msg = add_buy(data)
    import urllib.parse
    msg_param = urllib.parse.quote(msg)
    success = 1 if ok else 0
    return RedirectResponse(url=f"/buy?msg={msg_param}&success={success}", status_code=303)

@app.post("/api/buy/update_node")
async def api_buy_update_node(buy_id: str = Form(...), field: str = Form(...), value: int = Form(...), current_user: dict = Depends(login_required)):
    from Sills.db_buy import update_buy_node
    ok, msg = update_buy_node(buy_id, field, value)
    return {"success": ok, "message": msg}

@app.post("/api/buy/update")
async def api_buy_update(buy_id: str = Form(...), field: str = Form(...), value: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限"}
    allowed_fields = ['order_id', 'vendor_id', 'buy_mpn', 'buy_brand', 'buy_price_rmb', 'buy_qty', 'sales_price_rmb', 'remark', 'price_kwr', 'price_usd']
    if field not in allowed_fields:
        return {"success": False, "message": f"非法字段: {field}"}
    from Sills.db_buy import update_buy
    success, msg = update_buy(buy_id, {field: value})
    return {"success": success, "message": msg}

@app.post("/api/buy/delete")
async def api_buy_delete(buy_id: str = Form(...), current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3': return {"success": False, "message": "仅管理员可删除"}
    from Sills.db_buy import delete_buy
    ok, msg = delete_buy(buy_id)
    return {"success": ok, "message": msg}

@app.post("/api/buy/batch_delete")
async def api_buy_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3': return {"success": False, "message": "仅管理员可删除"}
    data = await request.json()
    ids = data.get("ids", [])
    ok, msg = batch_delete_buy(ids)
    return {"success": ok, "message": msg}

@app.post("/api/buy/export_csv")
async def buy_export_csv(request: Request, current_user: dict = Depends(login_required)):
    data = await request.json()
    ids = data.get("ids", [])
    if not ids: return {"success": False, "message": "未选择记录"}
    placeholders = ','.join(['?'] * len(ids))
    with get_db_connection() as conn:
        buys = conn.execute(f"""
            SELECT b.*, v.vendor_name, v.address as vendor_address, ord.order_no,
                   c.cli_id, c.cli_name, off.date_code, off.delivery_date
            FROM uni_buy b
            LEFT JOIN uni_vendor v ON b.vendor_id = v.vendor_id
            LEFT JOIN uni_order ord ON b.order_id = ord.order_id
            LEFT JOIN uni_cli c ON ord.cli_id = c.cli_id
            LEFT JOIN uni_offer off ON ord.offer_id = off.offer_id
            WHERE b.buy_id IN ({placeholders})
        """, ids).fetchall()

    krw_rate, usd_rate, _ = get_exchange_rates()

    import io, csv
    output = io.StringIO(); output.write('\ufeff')
    writer = csv.writer(output)
    # 与页面显示字段一致
    writer.writerow(['日期','采购编号','对应销售单','客户ID','客户名称','供应商','供应商地址','型号','品牌','批号(DC)','货期','采购单价(RMB)','销售报价(RMB)','采购单价(KWR)','采购单价(USD)','数量','总额(RMB)','货源确认','下单确认','入库确认','发货确认','备注'])

    for r in buys:
        d = dict(r)
        buy_price = float(d.get('buy_price_rmb') or 0)

        # KWR/USD 计算
        price_kwr = d.get('price_kwr')
        price_usd = d.get('price_usd')
        if not price_kwr or float(price_kwr or 0) == 0:
            if krw_rate > 10: price_kwr = round(buy_price * krw_rate, 1)
            else: price_kwr = round(buy_price / krw_rate, 1) if krw_rate else 0
        if not price_usd or float(price_usd or 0) == 0:
            price_usd = round(buy_price * usd_rate, 2) if usd_rate else 0

        writer.writerow([
            d.get('buy_date', ''),
            d.get('buy_id', ''),
            d.get('order_no') or '',
            d.get('cli_id') or '',
            d.get('cli_name') or '',
            d.get('vendor_name') or '',
            d.get('vendor_address') or '',
            d.get('buy_mpn') or '',
            d.get('buy_brand') or '',
            d.get('date_code') or '',
            d.get('delivery_date') or '',
            f"{buy_price:.2f}",
            f"{float(d.get('sales_price_rmb') or 0):.2f}",
            price_kwr,
            price_usd,
            d.get('buy_qty') or 0,
            d.get('total_amount') or 0,
            '是' if d.get('is_source_confirmed') == 1 else '否',
            '是' if d.get('is_ordered') == 1 else '否',
            '是' if d.get('is_instock') == 1 else '否',
            '是' if d.get('is_shipped') == 1 else '否',
            d.get('remark') or ''
        ])
    return {"success": True, "csv_content": output.getvalue()}
# --- New Workflow API endpoints ---

@app.post("/api/quote/batch_to_offer")
async def api_quote_batch_to_offer(data: dict, current_user: dict = Depends(login_required)):
    ids = data.get('ids', [])
    if not ids: return {"success": False, "message": "未选中记录"}
    try:
        ok, msg = batch_convert_from_quote(ids, current_user['emp_id'])
        return {"success": ok, "message": msg}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/offer/batch_to_order")
async def api_offer_batch_to_order(data: dict, current_user: dict = Depends(login_required)):
    ids = data.get('ids', [])
    cli_id = data.get('cli_id')
    if not ids: return {"success": False, "message": "未选中记录"}
    try:
        ok, msg = batch_convert_from_offer(ids, cli_id)
        return {"success": ok, "message": msg}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/order/batch_to_buy")
async def api_order_batch_to_buy(data: dict, current_user: dict = Depends(login_required)):
    ids = data.get('ids', [])
    if not ids: return {"success": False, "message": "未选中记录"}
    try:
        ok, msg = batch_convert_from_order(ids)
        return {"success": ok, "message": msg}
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/api/order/generate_pi")
async def api_order_generate_pi(request: Request, current_user: dict = Depends(login_required)):
    """生成PI文件"""
    from Sills.document_generator import generate_pi

    data = await request.json()
    order_ids = data.get("order_ids", [])
    if not order_ids:
        return {"success": False, "message": "未选择任何订单"}

    try:
        success, result = generate_pi(order_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "pdf_path": result.get("pdf_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.post("/api/order/generate_pi_us")
async def api_order_generate_pi_us(request: Request, current_user: dict = Depends(login_required)):
    """生成PI-US文件（美元版）"""
    from Sills.document_generator import generate_pi_us

    data = await request.json()
    order_ids = data.get("order_ids", [])
    if not order_ids:
        return {"success": False, "message": "未选择任何订单"}

    try:
        success, result = generate_pi_us(order_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.post("/api/order/generate_ci_kr")
async def api_order_generate_ci_kr(request: Request, current_user: dict = Depends(login_required)):
    """生成CI-KR文件"""
    from Sills.ci_generator import generate_ci_kr

    data = await request.json()
    order_ids = data.get("order_ids", [])
    if not order_ids:
        return {"success": False, "message": "未选择任何订单"}

    try:
        success, result = generate_ci_kr(order_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "pdf_path": result.get("pdf_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.post("/api/order/generate_ci_us")
async def api_order_generate_ci_us(request: Request, current_user: dict = Depends(login_required)):
    """生成CI-US文件（美元版）"""
    from Sills.document_generator import generate_ci_us

    data = await request.json()
    order_ids = data.get("order_ids", [])
    if not order_ids:
        return {"success": False, "message": "未选择任何订单"}

    try:
        success, result = generate_ci_us(order_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "pdf_path": result.get("pdf_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


# ============================================================
# 基于报价的 PI/CI 生成 API
# ============================================================

@app.post("/api/offer/generate_pi")
async def api_offer_generate_pi(request: Request, current_user: dict = Depends(login_required)):
    """基于报价生成PI-KR文件"""
    from Sills.document_generator import generate_pi_from_offers

    data = await request.json()
    offer_ids = data.get("offer_ids", [])
    if not offer_ids:
        return {"success": False, "message": "未选择任何报价"}

    try:
        success, result = generate_pi_from_offers(offer_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.post("/api/offer/generate_pi_us")
async def api_offer_generate_pi_us(request: Request, current_user: dict = Depends(login_required)):
    """基于报价生成PI-US文件（美元版）"""
    from Sills.document_generator import generate_pi_us_from_offers

    data = await request.json()
    offer_ids = data.get("offer_ids", [])
    if not offer_ids:
        return {"success": False, "message": "未选择任何报价"}

    try:
        success, result = generate_pi_us_from_offers(offer_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.post("/api/offer/generate_ci_kr")
async def api_offer_generate_ci_kr(request: Request, current_user: dict = Depends(login_required)):
    """基于报价生成CI-KR文件"""
    from Sills.ci_generator import generate_ci_kr_from_offers

    data = await request.json()
    offer_ids = data.get("offer_ids", [])
    if not offer_ids:
        return {"success": False, "message": "未选择任何报价"}

    try:
        success, result = generate_ci_kr_from_offers(offer_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "pdf_path": result.get("pdf_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.post("/api/offer/generate_ci_us")
async def api_offer_generate_ci_us(request: Request, current_user: dict = Depends(login_required)):
    """基于报价生成CI-US文件（美元版）"""
    from Sills.document_generator import generate_ci_us_from_offers

    data = await request.json()
    offer_ids = data.get("offer_ids", [])
    if not offer_ids:
        return {"success": False, "message": "未选择任何报价"}

    try:
        success, result = generate_ci_us_from_offers(offer_ids)

        if success:
            return {
                "success": True,
                "excel_path": result.get("excel_path", ""),
                "pdf_path": result.get("pdf_path", ""),
                "count": result.get("count", 0),
                "cli_name": result.get("cli_name", ""),
                "invoice_no": result.get("invoice_no", "")
            }
        else:
            return {"success": False, "message": f"生成失败: {result}"}
    except Exception as e:
        return {"success": False, "message": f"生成异常: {str(e)}"}


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return RedirectResponse(url="/", status_code=303)

    # Get backup path info - 根据系统选择路径
    is_windows = platform.system() == "Windows"
    backup_root = r"E:\WorkPlace\1_AIemployee\备份目录" if is_windows else "/home/kim/workspace/DbBackup"

    return templates.TemplateResponse("settings.html", {
        "request": request,
        "active_page": "settings",
        "current_user": current_user,
        "backup_path": backup_root
    })

@app.post("/api/backup")
async def api_backup(current_user: dict = Depends(login_required)):
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可执行备份"}

    try:
        count, backup_dir = do_backup()
        return {"success": True, "message": f"备份成功！已备份 {count} 个数据库文件和 static 目录到 {backup_dir}"}
    except Exception as e:
        return {"success": False, "message": f"备份失败: {str(e)}"}

@app.get("/api/backup/list")
async def api_backup_list(current_user: dict = Depends(login_required)):
    """获取备份目录列表"""
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可执行"}

    try:
        backup_root = get_backup_root()
        if not os.path.exists(backup_root):
            return {"success": True, "backups": []}

        backups = []
        for item in os.listdir(backup_root):
            try:
                item_path = os.path.join(backup_root, item)
                if os.path.isdir(item_path) and item.startswith("backup_"):
                    # 获取目录修改时间
                    mtime = os.path.getmtime(item_path)
                    mtime_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
                    # 获取目录大小
                    total_size = 0
                    for f in os.listdir(item_path):
                        fpath = os.path.join(item_path, f)
                        if os.path.isfile(fpath):
                            total_size += os.path.getsize(fpath)
                    backups.append({
                        "name": item,
                        "path": item_path,
                        "mtime": mtime_str,
                        "size": f"{total_size / 1024:.1f} KB"
                    })
            except Exception as e:
                print(f"处理备份目录 {item} 时出错: {str(e)}")
                continue

        # 按修改时间降序排序
        backups.sort(key=lambda x: x['mtime'], reverse=True)
        return {"success": True, "backups": backups}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "message": f"获取备份列表失败: {str(e)}"}

@app.post("/api/backup/restore")
async def api_backup_restore(backup_path: str = Form(...), current_user: dict = Depends(login_required)):
    """从备份目录恢复数据库 - 支持 SQLite 和 PostgreSQL"""
    import subprocess

    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可执行恢复"}

    try:
        if not os.path.exists(backup_path):
            return {"success": False, "message": f"备份目录不存在: {backup_path}"}

        project_root = os.path.dirname(os.path.abspath(__file__))
        restored_count = 0

        # 关闭所有数据库连接
        from Sills.base import clear_cache, close_all_connections
        close_all_connections()
        clear_cache()

        # 根据数据库类型选择恢复方式
        if is_postgresql():
            # PostgreSQL 恢复：使用 psql 执行 SQL 文件
            sql_file = os.path.join(backup_path, "uni_platform.sql")
            if not os.path.exists(sql_file):
                return {"success": False, "message": "备份目录中没有找到 uni_platform.sql 文件"}

            pg_config = get_pg_config()

            # 构建 psql 命令
            cmd = [
                "psql",
                "-h", pg_config['host'],
                "-p", str(pg_config['port']),
                "-U", pg_config['user'],
                "-d", pg_config['database'],
                "-f", sql_file
            ]

            # 设置环境变量传递密码
            env = os.environ.copy()
            env["PGPASSWORD"] = pg_config['password']

            try:
                result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=600)
                if result.returncode == 0:
                    restored_count = 1
                    print(f"[恢复] PostgreSQL 恢复成功")
                else:
                    # psql 可能返回错误但部分成功，检查是否有输出
                    if "ERROR" in result.stderr:
                        return {"success": False, "message": f"psql 执行出错: {result.stderr[:500]}"}
                    restored_count = 1  # 警告但视为成功
            except FileNotFoundError:
                return {"success": False, "message": "psql 命令未找到，请确保 PostgreSQL 客户端已安装并添加到 PATH"}
        else:
            # SQLite 恢复：复制文件
            for db_file in os.listdir(backup_path):
                if db_file.endswith(".db"):
                    src = os.path.join(backup_path, db_file)
                    dst = os.path.join(project_root, db_file)
                    # 删除目标文件（包括 WAL 和 SHM 文件）
                    if os.path.exists(dst):
                        os.remove(dst)
                    wal_file = dst + "-wal"
                    shm_file = dst + "-shm"
                    if os.path.exists(wal_file):
                        os.remove(wal_file)
                    if os.path.exists(shm_file):
                        os.remove(shm_file)
                    # 复制备份文件
                    shutil.copy2(src, dst)
                    restored_count += 1

        if restored_count == 0:
            return {"success": False, "message": "备份目录中没有找到数据库文件"}

        # 再次清除缓存
        clear_cache()

        db_type = "PostgreSQL" if is_postgresql() else "SQLite"
        return {"success": True, "message": f"恢复成功！已恢复 {db_type} 数据库（请刷新页面）"}
    except Exception as e:
        return {"success": False, "message": f"恢复失败: {str(e)}"}


@app.post("/api/backup/delete")
async def api_backup_delete(backup_path: str = Form(...), current_user: dict = Depends(login_required)):
    """删除备份目录"""
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可执行删除"}

    try:
        # 安全检查：确保路径在备份目录内
        backup_root = get_backup_root()
        if not os.path.exists(backup_path):
            return {"success": False, "message": "备份目录不存在"}

        # 确保是备份目录
        if not backup_path.startswith(backup_root):
            return {"success": False, "message": "非法路径"}

        if not os.path.basename(backup_path).startswith("backup_"):
            return {"success": False, "message": "不是有效的备份目录"}

        # 删除目录
        shutil.rmtree(backup_path)
        return {"success": True, "message": f"已删除备份: {os.path.basename(backup_path)}"}
    except Exception as e:
        return {"success": False, "message": f"删除失败: {str(e)}"}


# ==================== SmartMail 邮件模块路由 ====================

@app.get("/mail", response_class=HTMLResponse)
async def mail_page(request: Request, current_user: dict = Depends(login_required)):
    """邮件中心页面"""
    from Sills.db_mail import get_mail_config
    mail_config = get_mail_config()
    current_email = mail_config.get('username', '') if mail_config else ''
    return templates.TemplateResponse("mail.html", {
        "request": request,
        "active_page": "mail",
        "current_user": current_user,
        "current_email": current_email
    })


@app.get("/email_task", response_class=HTMLResponse)
async def email_task_page(request: Request, current_user: dict = Depends(login_required)):
    """开发信任务页面"""
    return templates.TemplateResponse("email_task.html", {
        "request": request,
        "active_page": "email_task",
        "current_user": current_user
    })


@app.get("/api/mail/list")
async def api_mail_list(
    folder: str = "inbox",
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取邮件列表（用户隔离）"""
    # 限制每页数量在1-1000之间
    page_size = max(1, min(1000, page_size))
    is_sent = 1 if folder == "sent" else 0
    # 获取当前邮件账户ID
    config = get_mail_config()
    account_id = config.get('account_id') if config else None
    result = get_mail_list(page=page, limit=page_size, is_sent=is_sent, search=search, account_id=account_id)
    return result


@app.post("/api/mail/send")
async def api_mail_send(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """发送邮件"""
    from Sills.db_mail import get_signature

    data = await request.json()
    to = data.get('to', '')
    subject = data.get('subject', '')
    body = data.get('body', '')
    html_body = data.get('html_body', '')
    cc = data.get('cc', '')

    if not to or not subject:
        return {"success": False, "message": "收件人和主题不能为空"}

    # 追加签名（签名现在是HTML格式）
    signature = get_signature()
    if signature:
        # 纯文本body：从HTML中提取纯文本（简单处理）
        import re
        plain_signature = re.sub(r'<[^>]+>', '', signature).replace('&nbsp;', ' ')
        body = body + "\n\n" + plain_signature if body else plain_signature
        # HTML body：直接追加HTML签名（空字符串也会进入else分支）
        if html_body and html_body.strip():
            html_body = html_body + "<br><br>" + signature
        else:
            html_body = signature

    result = send_email_now(to=to, subject=subject, body=body, html_body=html_body, cc=cc)

    if result["success"]:
        return {"success": True, "message": "邮件发送成功", "message_id": result.get("message_id")}
    else:
        # 发送失败，自动保存草稿
        from Sills.db_mail import save_draft
        account = get_mail_config()
        account_id = account.get('id') if account else None

        draft_data = {
            'subject': subject,
            'to_addr': to,
            'cc_addr': cc,
            'content': body,
            'html_content': html_body,
            'account_id': account_id
        }
        draft_id = save_draft(draft_data)

        return {
            "success": False,
            "message": f"发送失败: {result.get('error', '未知错误')}",
            "draft_saved": True,
            "draft_id": draft_id
        }


@app.post("/api/mail/send-with-attachments")
async def api_mail_send_with_attachments(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """发送带附件的邮件"""
    from fastapi import UploadFile, File, Form
    from Sills.db_mail import get_signature
    import tempfile
    import os

    form = await request.form()
    to = form.get('to', '')
    subject = form.get('subject', '')
    body = form.get('body', '')
    html_body = form.get('html_body', '')
    cc = form.get('cc', '')

    if not to or not subject:
        return {"success": False, "message": "收件人和主题不能为空"}

    # 追加签名（签名现在是HTML格式）
    signature = get_signature()
    if signature:
        # 纯文本body：从HTML中提取纯文本（简单处理）
        import re
        plain_signature = re.sub(r'<[^>]+>', '', signature).replace('&nbsp;', ' ')
        body = body + "\n\n" + plain_signature if body else plain_signature
        # HTML body：直接追加HTML签名（空字符串也会进入else分支）
        if html_body and html_body.strip():
            html_body = html_body + "<br><br>" + signature
        else:
            html_body = signature

    # 获取所有附件（使用getlist获取多个同名字段）
    attachments = []
    attachment_files = form.getlist('attachments')
    for value in attachment_files:
        if hasattr(value, 'filename') and value.filename:
            # 保存到临时文件
            content = await value.read()
            if content:
                temp_dir = tempfile.gettempdir()
                temp_path = os.path.join(temp_dir, value.filename)
                with open(temp_path, 'wb') as f:
                    f.write(content)
                attachments.append({
                    'path': temp_path,
                    'filename': value.filename,
                    'content_type': value.content_type or 'application/octet-stream'
                })

    from Sills.mail_service import send_email_with_attachments
    result = send_email_with_attachments(
        to=to, subject=subject, body=body,
        html_body=html_body, cc=cc, attachments=attachments
    )

    # 清理临时文件
    for att in attachments:
        try:
            os.remove(att['path'])
        except:
            pass

    if result["success"]:
        return {"success": True, "message": "邮件发送成功", "message_id": result.get("message_id")}
    else:
        # 发送失败，自动保存草稿
        from Sills.db_mail import save_draft
        account = get_mail_config()
        account_id = account.get('id') if account else None

        draft_data = {
            'subject': subject,
            'to_addr': to,
            'cc_addr': cc,
            'content': body,
            'html_content': html_body,
            'account_id': account_id
        }
        draft_id = save_draft(draft_data)

        return {
            "success": False,
            "message": f"发送失败: {result.get('error', '未知错误')}",
            "draft_saved": True,
            "draft_id": draft_id
        }


@app.post("/api/mail/sync")
async def api_mail_sync(current_user: dict = Depends(login_required)):
    """同步邮件（后台异步）"""
    if is_sync_locked():
        return {"success": False, "message": "同步任务正在进行中，请稍后"}

    result = sync_inbox_async()
    return {"success": True, "message": "同步任务已启动"}


@app.post("/api/mail/refresh")
async def api_mail_refresh(current_user: dict = Depends(login_required)):
    """刷新邮件：只同步上次同步之后的新邮件"""
    from Sills.mail_service import refresh_emails_async
    if is_sync_locked():
        return {"success": False, "message": "同步任务正在进行中，请稍后"}

    refresh_emails_async()
    return {"success": True, "message": "刷新任务已启动"}


@app.post("/api/mail/sync-new")
async def api_mail_sync_new(current_user: dict = Depends(login_required)):
    """增量同步：只获取新邮件"""
    from Sills.mail_service import sync_new_emails_async
    if is_sync_locked():
        return {"success": False, "message": "同步任务正在进行中，请稍后"}

    result = sync_new_emails_async()
    return {"success": True, "message": "增量同步任务已启动"}


@app.get("/api/mail/sync/status")
async def api_mail_sync_status(current_user: dict = Depends(login_required)):
    """获取同步状态和进度"""
    progress = get_sync_progress()
    return {
        "success": True,
        **progress
    }


@app.post("/api/mail/sync/cancel")
async def api_mail_sync_cancel(current_user: dict = Depends(login_required)):
    """取消同步"""
    from Sills.mail_service import request_cancel_sync, is_sync_cancelled
    request_cancel_sync()
    return {"success": True, "message": "已发送取消请求"}


@app.get("/api/mail/config")
async def api_mail_config_get(current_user: dict = Depends(login_required)):
    """获取邮件配置"""
    config = get_mail_config()
    if config:
        # 隐藏密码
        config["password"] = "******" if config.get("password") else ""
    return {
        "success": True,
        "config": config
    }


@app.get("/api/mail/settings")
async def api_mail_settings_get(current_user: dict = Depends(login_required)):
    """获取邮件设置（合并多个设置项）"""
    from Sills.db_mail import get_signature, get_sync_days, get_sync_deleted_setting, get_undo_send_seconds
    sync_interval = get_sync_interval()
    undo_seconds = get_undo_send_seconds()
    sync_range = get_sync_days()
    sync_deleted = get_sync_deleted_setting()
    signature = get_signature()

    return {
        "success": True,
        "settings": {
            "syncInterval": sync_interval,
            "undoSendSeconds": undo_seconds,
            "syncRange": sync_range,
            "syncDeleted": sync_deleted,
            "signature": signature,
            "geminiKey": ""
        }
    }


@app.post("/api/mail/settings")
async def api_mail_settings_post(request: Request, current_user: dict = Depends(login_required)):
    """更新邮件设置"""
    data = await request.json()

    if 'syncInterval' in data:
        set_sync_interval(data['syncInterval'])
    if 'undoSendSeconds' in data:
        set_undo_send_seconds(data['undoSendSeconds'])
    if 'syncRange' in data:
        set_sync_range(data['syncRange'])
    if 'syncDeleted' in data:
        set_sync_deleted(data['syncDeleted'])
    if 'signature' in data:
        from Sills.db_mail import set_signature
        set_signature(data['signature'])
    if 'geminiKey' in data:
        # Gemini API key 设置需要单独处理
        pass

    return {"success": True, "message": "设置已保存"}


@app.post("/api/mail/config/batch")
async def api_mail_config_batch(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置分批同步配置"""
    from Sills.db_mail import update_mail_account, get_mail_config
    try:
        data = await request.json()
        batch_size = data.get('batch_size', 100)
        pause_seconds = data.get('pause_seconds', 1.0)

        # 获取当前账户
        config = get_mail_config()
        if not config:
            return {"success": False, "message": "未找到邮件账户"}

        account_id = config.get('account_id')

        # 更新配置
        update_mail_account(account_id, {
            'sync_batch_size': batch_size,
            'sync_pause_seconds': pause_seconds
        })

        return {"success": True, "message": f"分批配置已保存"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.post("/api/mail/config")
async def api_mail_config_update(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """更新邮件配置（含连接验证）"""
    from pydantic import BaseModel
    from Sills.mail_service import IMAPClient

    class MailConfigRequest(BaseModel):
        imap_server: str = ""
        imap_port: int = 993
        smtp_server: str = ""
        smtp_port: int = 587
        username: str = ""
        password: str = ""
        use_tls: int = 1
        sync_interval: int = 5

    try:
        data = await request.json()
        config_data = MailConfigRequest(**data)
    except Exception as e:
        return {"success": False, "message": f"请求数据格式错误: {str(e)}"}

    # 验证必填字段
    if not config_data.imap_server or not config_data.smtp_server or not config_data.username:
        return {"success": False, "message": "服务器地址和用户名不能为空"}

    # 验证IMAP连接（如果提供了密码）
    if config_data.password and config_data.password != "******":
        test_config = {
            'imap_server': config_data.imap_server,
            'imap_port': config_data.imap_port,
            'username': config_data.username,
            'password': config_data.password,
            'use_tls': config_data.use_tls
        }
        try:
            imap_client = IMAPClient(test_config)
            imap_client.connect()
            imap_client.disconnect()
        except Exception as e:
            return {"success": False, "message": f"IMAP连接验证失败: {str(e)}"}

    config = {
        "imap_server": config_data.imap_server,
        "imap_port": config_data.imap_port,
        "smtp_server": config_data.smtp_server,
        "smtp_port": config_data.smtp_port,
        "username": config_data.username,
        "use_tls": config_data.use_tls,
        "sync_interval": config_data.sync_interval
    }

    # 只有提供了新密码才更新
    if config_data.password and config_data.password != "******":
        config["password"] = config_data.password

    try:
        result = update_mail_config(config)
        if result:
            return {"success": True, "message": "设置保存成功"}
        else:
            return {"success": False, "message": "保存失败"}
    except ValueError as e:
        # 加密密钥未配置
        return {"success": False, "message": f"系统配置错误: {str(e)}"}
    except Exception as e:
        return {"success": False, "message": f"保存失败: {str(e)}"}


@app.get("/api/mail/accounts")
async def api_mail_accounts_list(current_user: dict = Depends(login_required)):
    """获取所有邮件账户列表"""
    accounts = get_all_mail_accounts()
    return {
        "success": True,
        "accounts": accounts
    }


@app.get("/api/mail/account/current")
async def api_mail_account_current(current_user: dict = Depends(login_required)):
    """获取当前邮件账户配置"""
    config = get_mail_config()
    if config:
        # 隐藏密码
        config['password'] = '******' if config.get('password') else ''
    return {
        "success": True,
        "config": config
    }


@app.get("/api/mail/account/{account_id}")
async def api_mail_account_get(account_id: int, current_user: dict = Depends(login_required)):
    """获取指定邮件账户配置"""
    config = get_mail_account_by_id(account_id)
    if not config:
        return {"success": False, "message": "账户不存在"}

    # 隐藏密码
    config['password'] = '******' if config.get('password') else ''
    return {
        "success": True,
        "config": config
    }


@app.get("/api/mail/sync-interval")
async def api_mail_sync_interval_get(current_user: dict = Depends(login_required)):
    """获取同步间隔设置"""
    interval = get_sync_interval()
    return {
        "success": True,
        "interval": interval
    }


@app.post("/api/mail/sync-interval")
async def api_mail_sync_interval_set(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置同步间隔"""
    try:
        data = await request.json()
        interval = data.get('interval', 30)
        if not isinstance(interval, int) or interval < 1:
            return {"success": False, "message": "同步间隔必须为正整数"}

        set_sync_interval(interval)
        return {"success": True, "message": f"同步间隔已设置为 {interval} 分钟"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.get("/api/mail/sync-days")
async def api_mail_sync_days_get(current_user: dict = Depends(login_required)):
    """获取同步时间范围设置"""
    from Sills.db_mail import get_sync_days
    days = get_sync_days()
    return {
        "success": True,
        "days": days
    }


@app.post("/api/mail/sync-days")
async def api_mail_sync_days_set(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置同步时间范围"""
    from Sills.db_mail import set_sync_days, clear_sync_date_range
    try:
        data = await request.json()
        days = data.get('days', 90)
        if not isinstance(days, int) or days < 1 or days > 365:
            return {"success": False, "message": "同步时间范围必须在1-365天之间"}

        set_sync_days(days)
        # 清除自定义日期范围
        clear_sync_date_range()
        return {"success": True, "message": f"同步时间范围已设置为 {days} 天"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.get("/api/mail/sync-deleted")
async def api_mail_sync_deleted_get(current_user: dict = Depends(login_required)):
    """获取"同步已删除邮件"开关设置"""
    from Sills.db_mail import get_sync_deleted_setting
    enabled = get_sync_deleted_setting()
    return {
        "success": True,
        "enabled": enabled
    }


@app.post("/api/mail/sync-deleted")
async def api_mail_sync_deleted_set(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置"同步已删除邮件"开关"""
    from Sills.db_mail import set_sync_deleted_setting
    try:
        data = await request.json()
        enabled = data.get('enabled', True)
        if not isinstance(enabled, bool):
            return {"success": False, "message": "enabled必须为布尔值"}

        set_sync_deleted_setting(enabled)
        return {"success": True, "message": f"设置已更新"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.get("/api/mail/sync-range")
async def api_mail_sync_range_get(current_user: dict = Depends(login_required)):
    """获取同步范围设置（快捷或自定义）"""
    from Sills.db_mail import get_sync_days, get_sync_date_range
    start_date, end_date = get_sync_date_range()
    if start_date and end_date:
        return {
            "success": True,
            "mode": "custom",
            "start_date": start_date,
            "end_date": end_date
        }
    else:
        return {
            "success": True,
            "mode": "quick",
            "days": get_sync_days()
        }


@app.post("/api/mail/sync-date-range")
async def api_mail_sync_date_range_set(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置自定义同步日期范围"""
    from Sills.db_mail import set_sync_date_range
    try:
        data = await request.json()
        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')

        if not start_date or not end_date:
            return {"success": False, "message": "请选择起始日期和结束日期"}

        # 验证日期格式
        from datetime import datetime
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
            datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return {"success": False, "message": "日期格式无效，请使用 YYYY-MM-DD 格式"}

        set_sync_date_range(start_date, end_date)
        return {"success": True, "message": f"同步范围已设置为 {start_date} 至 {end_date}"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.get("/api/mail/undo-send-seconds")
async def api_mail_undo_send_seconds_get(current_user: dict = Depends(login_required)):
    """获取发送撤销时间设置"""
    from Sills.db_mail import get_undo_send_seconds
    seconds = get_undo_send_seconds()
    return {
        "success": True,
        "seconds": seconds
    }


@app.post("/api/mail/undo-send-seconds")
async def api_mail_undo_send_seconds_set(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置发送撤销时间"""
    from Sills.db_mail import set_undo_send_seconds
    try:
        data = await request.json()
        seconds = data.get('seconds', 5)
        if not isinstance(seconds, int) or seconds < 0 or seconds > 30:
            return {"success": False, "message": "撤销时间必须在0-30秒之间"}

        set_undo_send_seconds(seconds)
        return {"success": True, "message": f"撤销时间已设置为 {seconds} 秒"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.get("/api/mail/signature")
async def api_mail_signature_get(current_user: dict = Depends(login_required)):
    """获取邮件签名"""
    from Sills.db_mail import get_signature
    signature = get_signature()
    return {
        "success": True,
        "signature": signature
    }


@app.post("/api/mail/signature")
async def api_mail_signature_set(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """设置邮件签名"""
    from Sills.db_mail import set_signature
    try:
        data = await request.json()
        signature = data.get('signature', '')
        set_signature(signature)
        return {"success": True, "message": "签名设置成功"}
    except Exception as e:
        return {"success": False, "message": f"设置失败: {str(e)}"}


@app.post("/api/mail/account/add")
async def api_mail_account_add(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """添加新邮件账户（含连接验证）"""
    from Sills.mail_service import IMAPClient

    try:
        data = await request.json()

        # 验证必填字段
        if not data.get('imap_server') or not data.get('smtp_server') or not data.get('username'):
            return {"success": False, "message": "服务器地址和用户名不能为空"}

        # 验证IMAP连接（如果提供了密码）
        if data.get('password') and data['password'] != "******":
            test_config = {
                'imap_server': data.get('imap_server'),
                'imap_port': data.get('imap_port', 993),
                'username': data.get('username'),
                'password': data['password'],
                'use_tls': data.get('use_tls', 1)
            }
            try:
                imap_client = IMAPClient(test_config)
                imap_client.connect()
                imap_client.disconnect()
            except Exception as e:
                return {"success": False, "message": f"IMAP连接验证失败: {str(e)}"}

        account_id = add_mail_account(data)
        return {
            "success": True,
            "message": "账户添加成功",
            "account_id": account_id
        }

    except ValueError as e:
        return {"success": False, "message": f"系统配置错误: {str(e)}"}
    except Exception as e:
        return {"success": False, "message": f"添加失败: {str(e)}"}


@app.post("/api/mail/account/update")
async def api_mail_account_update(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """更新邮件账户（含连接验证）"""
    from Sills.mail_service import IMAPClient

    try:
        data = await request.json()
        account_id = data.get('id')

        if not account_id:
            return {"success": False, "message": "账户ID不能为空"}

        # 验证必填字段
        if not data.get('imap_server') or not data.get('smtp_server') or not data.get('username'):
            return {"success": False, "message": "服务器地址和用户名不能为空"}

        # 验证IMAP连接（如果提供了新密码）
        if data.get('password') and data['password'] != "******":
            test_config = {
                'imap_server': data.get('imap_server'),
                'imap_port': data.get('imap_port', 993),
                'username': data.get('username'),
                'password': data['password'],
                'use_tls': data.get('use_tls', 1)
            }
            try:
                imap_client = IMAPClient(test_config)
                imap_client.connect()
                imap_client.disconnect()
            except Exception as e:
                return {"success": False, "message": f"IMAP连接验证失败: {str(e)}"}

        result = update_mail_account(account_id, data)
        if result:
            return {"success": True, "message": "账户更新成功"}
        else:
            return {"success": False, "message": "更新失败"}

    except ValueError as e:
        return {"success": False, "message": f"系统配置错误: {str(e)}"}
    except Exception as e:
        return {"success": False, "message": f"更新失败: {str(e)}"}


@app.post("/api/mail/account/switch")
async def api_mail_account_switch(
    request: Request,
    current_user: dict = Depends(login_required)
):
    """切换当前邮件账户"""
    try:
        data = await request.json()
        account_id = data.get('account_id')

        if not account_id:
            return {"success": False, "message": "账户ID不能为空"}

        result = switch_current_account(account_id)
        if result:
            return {"success": True, "message": "已切换到新账户"}
        else:
            return {"success": False, "message": "切换失败，账户不存在"}

    except Exception as e:
        return {"success": False, "message": f"切换失败: {str(e)}"}


@app.delete("/api/mail/account/{account_id}")
async def api_mail_account_delete(
    account_id: int,
    current_user: dict = Depends(login_required)
):
    """删除邮件账户"""
    try:
        result = delete_mail_account(account_id)
        if result.get('success'):
            return {"success": True, "message": result.get('message', '删除成功')}
        else:
            return {"success": False, "message": result.get('message', '删除失败')}
    except Exception as e:
        return {"success": False, "message": f"删除失败: {str(e)}"}


# ==================== 邮件文件夹 API (必须在 /api/mail/{mail_id} 之前) ====================

@app.get("/api/mail/folders")
async def api_get_folders(current_user: dict = Depends(login_required)):
    """获取文件夹列表"""
    account = get_mail_config()
    account_id = account.get('id') if account else None
    folders = get_folders(account_id)
    # 为每个文件夹添加邮件数量
    for folder in folders:
        folder['mail_count'] = get_mail_count_by_folder(folder['id'], account_id)
    return {"success": True, "folders": folders}


@app.post("/api/mail/folder/add")
async def api_add_folder(request: Request, current_user: dict = Depends(login_required)):
    """添加文件夹"""
    data = await request.json()
    folder_name = data.get('folder_name', '').strip()
    if not folder_name:
        return {"success": False, "message": "文件夹名称不能为空"}

    account = get_mail_config()
    account_id = account.get('id') if account else None

    folder_id = add_folder({
        'folder_name': folder_name,
        'folder_icon': data.get('folder_icon', 'folder'),
        'sort_order': data.get('sort_order', 0),
        'account_id': account_id
    })
    return {"success": True, "folder_id": folder_id}


@app.post("/api/mail/folder/update")
async def api_update_folder(request: Request, current_user: dict = Depends(login_required)):
    """更新文件夹"""
    data = await request.json()
    folder_id = data.get('folder_id')
    if not folder_id:
        return {"success": False, "message": "缺少文件夹ID"}

    success = update_folder(folder_id, {
        'folder_name': data.get('folder_name'),
        'folder_icon': data.get('folder_icon'),
        'sort_order': data.get('sort_order', 0)
    })
    return {"success": success}


@app.post("/api/mail/folder/delete")
async def api_delete_folder(request: Request, current_user: dict = Depends(login_required)):
    """删除文件夹"""
    data = await request.json()
    folder_id = data.get('folder_id')
    if not folder_id:
        return {"success": False, "message": "缺少文件夹ID"}

    success = delete_folder(folder_id)
    return {"success": success}


# ==================== 邮件过滤规则 API ====================

@app.get("/api/mail/filter-rules")
async def api_get_filter_rules(folder_id: int = None, current_user: dict = Depends(login_required)):
    """获取过滤规则列表"""
    rules = get_filter_rules(folder_id)
    return {"success": True, "rules": rules}


@app.post("/api/mail/filter-rule/add")
async def api_add_filter_rule(request: Request, current_user: dict = Depends(login_required)):
    """添加过滤规则"""
    data = await request.json()
    folder_id = data.get('folder_id')
    keyword = data.get('keyword', '').strip()

    if not folder_id or not keyword:
        return {"success": False, "message": "文件夹和关键词不能为空"}

    rule_id = add_filter_rule({
        'folder_id': folder_id,
        'keyword': keyword,
        'priority': data.get('priority', 0),
        'is_enabled': data.get('is_enabled', 1)
    })
    return {"success": True, "rule_id": rule_id}


@app.post("/api/mail/filter-rule/update")
async def api_update_filter_rule(request: Request, current_user: dict = Depends(login_required)):
    """更新过滤规则"""
    data = await request.json()
    rule_id = data.get('rule_id')
    if not rule_id:
        return {"success": False, "message": "缺少规则ID"}

    success = update_filter_rule(rule_id, {
        'keyword': data.get('keyword'),
        'priority': data.get('priority', 0),
        'is_enabled': data.get('is_enabled', 1)
    })
    return {"success": success}


@app.post("/api/mail/filter-rule/delete")
async def api_delete_filter_rule(request: Request, current_user: dict = Depends(login_required)):
    """删除过滤规则"""
    data = await request.json()
    rule_id = data.get('rule_id')
    if not rule_id:
        return {"success": False, "message": "缺少规则ID"}

    success = delete_filter_rule(rule_id)
    return {"success": success}


# ==================== 自动分类 API ====================

@app.post("/api/mail/auto-classify")
async def api_auto_classify(current_user: dict = Depends(login_required)):
    """自动分类邮件（文件夹规则 + 邮件类型分类）"""
    account = get_mail_config()
    account_id = account.get('id') if account else None

    # 文件夹规则分类
    folder_result = auto_classify_emails(account_id)

    # 邮件类型分类（已读/未读回执、退信）
    type_result = classify_mails(account_id)

    return {
        "success": True,
        "classified_count": folder_result['classified_count'],
        "rule_count": folder_result['rule_count'],
        "type_classification": {
            "read_receipts": type_result['read_receipts'],
            "unread_receipts": type_result['unread_receipts'],
            "bounced": type_result['bounced'],
            "recipients_extracted": type_result['recipients_extracted']
        }
    }


@app.get("/api/mail/folder/{folder_id}")
async def api_get_mails_by_folder(
    folder_id: int,
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取指定文件夹的邮件列表"""
    page_size = max(1, min(1000, page_size))
    account = get_mail_config()
    account_id = account.get('id') if account else None

    result = get_mails_by_folder(folder_id, page, page_size, search, account_id)
    return result


@app.get("/api/mail/trash")
async def api_mail_trash_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取回收站邮件列表"""
    from Sills.db_mail import get_trash_list
    page_size = max(1, min(1000, page_size))
    account = get_mail_config()
    account_id = account.get('id') if account else None
    result = get_trash_list(page=page, limit=page_size, search=search, account_id=account_id)
    return result


@app.post("/api/mail/trash/empty")
async def api_mail_empty_trash(current_user: dict = Depends(login_required)):
    """清空回收站"""
    from Sills.db_mail import empty_trash
    deleted = empty_trash()
    return {"success": True, "deleted": deleted}


# ============ 草稿箱 API ============

@app.get("/api/mail/draft")
async def api_mail_draft_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取草稿列表"""
    from Sills.db_mail import get_draft_list
    page_size = max(1, min(1000, page_size))
    account = get_mail_config()
    account_id = account.get('id') if account else None
    result = get_draft_list(page=page, limit=page_size, search=search, account_id=account_id)
    return result


@app.post("/api/mail/draft")
async def api_mail_save_draft(request: Request, current_user: dict = Depends(login_required)):
    """保存草稿"""
    from Sills.db_mail import save_draft
    data = await request.json()

    account = get_mail_config()
    account_id = account.get('id') if account else None

    draft_data = {
        'subject': data.get('subject', ''),
        'to_addr': data.get('to', ''),
        'cc_addr': data.get('cc', ''),
        'content': data.get('body', ''),
        'html_content': data.get('html_body', ''),
        'account_id': account_id
    }

    draft_id = save_draft(draft_data)
    return {"success": True, "draft_id": draft_id}


@app.get("/api/mail/draft/{draft_id}")
async def api_mail_get_draft(draft_id: int, current_user: dict = Depends(login_required)):
    """获取草稿详情"""
    from Sills.db_mail import get_draft_by_id
    draft = get_draft_by_id(draft_id)
    if not draft:
        raise HTTPException(status_code=404, detail="草稿不存在")
    return {"success": True, "draft": draft}


@app.put("/api/mail/draft/{draft_id}")
async def api_mail_update_draft(
    draft_id: int,
    request: Request,
    current_user: dict = Depends(login_required)
):
    """更新草稿"""
    from Sills.db_mail import update_draft
    data = await request.json()

    draft_data = {
        'subject': data.get('subject', ''),
        'to_addr': data.get('to', ''),
        'cc_addr': data.get('cc', ''),
        'content': data.get('body', ''),
        'html_content': data.get('html_body', '')
    }

    success = update_draft(draft_id, draft_data)
    return {"success": success}


@app.delete("/api/mail/draft/{draft_id}")
async def api_mail_delete_draft(draft_id: int, current_user: dict = Depends(login_required)):
    """删除草稿"""
    from Sills.db_mail import delete_draft
    success = delete_draft(draft_id)
    return {"success": success}


# ============ 黑名单邮件箱 API ============

@app.get("/api/mail/blacklisted")
async def api_mail_blacklisted_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取黑名单邮件列表"""
    from Sills.db_mail import get_blacklisted_list
    page_size = max(1, min(1000, page_size))
    account = get_mail_config()
    account_id = account.get('id') if account else None
    result = get_blacklisted_list(page=page, limit=page_size, search=search, account_id=account_id)
    return result


@app.get("/api/mail/spam")
async def api_mail_spam_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取垃圾邮件列表"""
    from Sills.db_mail import get_spam_list
    page_size = max(1, min(1000, page_size))
    account = get_mail_config()
    account_id = account.get('id') if account else None
    result = get_spam_list(page=page, limit=page_size, search=search, account_id=account_id)
    return result


@app.post("/api/mail/{mail_id}/move")
async def api_mail_move_to_folder(
    mail_id: int,
    request: Request,
    current_user: dict = Depends(login_required)
):
    """移动邮件到指定文件夹"""
    from Sills.db_mail import move_email_to_folder
    data = await request.json()
    folder_id = data.get('folder_id')  # None表示移回收件箱
    success = move_email_to_folder(mail_id, folder_id)
    return {"success": success}


@app.post("/api/mail/batch-move")
async def api_mail_batch_move(request: Request, current_user: dict = Depends(login_required)):
    """批量移动邮件到指定文件夹"""
    from Sills.db_mail import move_emails_to_folder
    data = await request.json()
    mail_ids = data.get('ids', [])
    folder_id = data.get('folder_id')  # None表示移回收件箱
    if not mail_ids:
        return {"success": False, "message": "No mail ids provided"}
    moved = move_emails_to_folder(mail_ids, folder_id)
    return {"success": True, "moved": moved}


@app.post("/api/mail/{mail_id}/blacklist")
async def api_mail_mark_blacklisted(mail_id: int, current_user: dict = Depends(login_required)):
    """将邮件移入黑名单邮件箱"""
    from Sills.db_mail import mark_email_as_blacklisted
    success = mark_email_as_blacklisted(mail_id)
    return {"success": success}


@app.post("/api/mail/{mail_id}/unblacklist")
async def api_mail_unmark_blacklisted(mail_id: int, current_user: dict = Depends(login_required)):
    """将邮件移出黑名单邮件箱"""
    from Sills.db_mail import unmark_email_as_blacklisted
    success = unmark_email_as_blacklisted(mail_id)
    return {"success": success}


@app.post("/api/mail/blacklist/classify")
async def api_mail_auto_classify_blacklist(current_user: dict = Depends(login_required)):
    """自动分类黑名单邮件"""
    from Sills.db_mail import auto_classify_blacklist
    account = get_mail_config()
    account_id = account.get('id') if account else None
    count = auto_classify_blacklist(account_id)
    return {"success": True, "classified_count": count}


# ==================== 黑名单地址管理 API ====================

@app.get("/api/mail/blacklist")
async def api_get_blacklist(current_user: dict = Depends(login_required)):
    """获取黑名单列表"""
    from Sills.db_mail import get_blacklist_list
    account = get_mail_config()
    account_id = account.get('id') if account else None
    blacklist = get_blacklist_list(account_id)
    return {"success": True, "blacklist": blacklist}


@app.post("/api/mail/blacklist")
async def api_add_blacklist(request: Request, current_user: dict = Depends(login_required)):
    """添加黑名单"""
    from Sills.db_mail import add_to_blacklist
    data = await request.json()
    email_addr = data.get('email_addr', '').strip()
    reason = data.get('reason', '')

    if not email_addr:
        return {"success": False, "message": "邮箱地址不能为空"}

    account = get_mail_config()
    account_id = account.get('id') if account else None

    success = add_to_blacklist(email_addr, reason, account_id)
    if success:
        return {"success": True, "message": "已添加到黑名单"}
    else:
        return {"success": False, "message": "添加失败或已存在"}


@app.delete("/api/mail/blacklist/{blacklist_id}")
async def api_remove_blacklist(blacklist_id: int, current_user: dict = Depends(login_required)):
    """移除黑名单"""
    from Sills.db_mail import remove_from_blacklist
    success = remove_from_blacklist(blacklist_id)
    return {"success": success}


@app.get("/api/mail/{mail_id}")
async def api_mail_detail(mail_id: int, current_user: dict = Depends(login_required)):
    """获取邮件详情"""
    email = get_mail_by_id(mail_id)
    if not email:
        raise HTTPException(status_code=404, detail="邮件不存在")

    # 获取关联信息
    relations = get_mail_relations(mail_id)

    return {
        "success": True,
        "email": email,
        "relations": relations
    }


@app.post("/api/mail/{mail_id}/relate")
async def api_mail_relate(
    mail_id: int,
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    current_user: dict = Depends(login_required)
):
    """关联邮件到客户或订单"""
    if entity_type not in ["client", "order"]:
        return {"success": False, "message": "无效的实体类型"}

    result = create_mail_relation(mail_id, entity_type, entity_id)
    return result


@app.delete("/api/mail/{mail_id}/relate/{relation_id}")
async def api_mail_unrelate(
    mail_id: int,
    relation_id: int,
    current_user: dict = Depends(login_required)
):
    """移除邮件关联"""
    result = remove_mail_relation(relation_id)
    return result


@app.delete("/api/mail/{mail_id}")
async def api_mail_delete(mail_id: int, current_user: dict = Depends(login_required)):
    """删除邮件（移入回收站）"""
    result = delete_email(mail_id)
    return {"success": result}


@app.post("/api/mail/{mail_id}/restore")
async def api_mail_restore(mail_id: int, current_user: dict = Depends(login_required)):
    """恢复邮件（从回收站）"""
    from Sills.db_mail import restore_email
    result = restore_email(mail_id)
    return {"success": result}


@app.delete("/api/mail/{mail_id}/permanent")
async def api_mail_permanent_delete(mail_id: int, current_user: dict = Depends(login_required)):
    """永久删除邮件"""
    from Sills.db_mail import permanently_delete_email
    result = permanently_delete_email(mail_id)
    return {"success": result}


@app.post("/api/mail/batch-delete")
async def api_mail_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    """批量删除邮件（移入回收站）"""
    from Sills.db_mail import batch_delete_emails
    data = await request.json()
    mail_ids = data.get('ids', [])
    if not mail_ids:
        return {"success": False, "message": "未选择邮件"}
    deleted = batch_delete_emails(mail_ids)
    return {"success": True, "deleted": deleted}


@app.post("/api/mail/batch-restore")
async def api_mail_batch_restore(request: Request, current_user: dict = Depends(login_required)):
    """批量恢复邮件（从回收站）"""
    from Sills.db_mail import batch_restore_emails
    data = await request.json()
    mail_ids = data.get('ids', [])
    if not mail_ids:
        return {"success": False, "message": "未选择邮件"}
    restored = batch_restore_emails(mail_ids)
    return {"success": True, "restored": restored}


@app.post("/api/mail/batch-permanent-delete")
async def api_mail_batch_permanent_delete(request: Request, current_user: dict = Depends(login_required)):
    """批量永久删除邮件"""
    from Sills.db_mail import batch_permanently_delete_emails
    data = await request.json()
    mail_ids = data.get('ids', [])
    if not mail_ids:
        return {"success": False, "message": "未选择邮件"}
    deleted = batch_permanently_delete_emails(mail_ids)
    return {"success": True, "deleted": deleted}


@app.post("/api/mail/cleanup-duplicates")
async def api_mail_cleanup_duplicates(current_user: dict = Depends(login_required)):
    """清理重复邮件"""
    from Sills.db_mail import cleanup_duplicate_emails
    result = cleanup_duplicate_emails()
    return {"success": True, **result}


@app.post("/api/mail/account/{account_id}/clear")
async def api_mail_clear_account(account_id: int, current_user: dict = Depends(login_required)):
    """清空指定账户的所有本地邮件"""
    from Sills.db_mail import clear_account_emails
    result = clear_account_emails(account_id)
    return {"success": True, **result}


@app.post("/api/mail/{mail_id}/read")
async def api_mail_mark_read(mail_id: int, current_user: dict = Depends(login_required)):
    """标记邮件为已读"""
    from Sills.db_mail import mark_email_read
    mark_email_read(mail_id)
    return {"success": True}


@app.get("/api/mail/{mail_id}/analyze")
async def api_mail_analyze(mail_id: int, current_user: dict = Depends(login_required)):
    """AI 分析邮件"""
    email = get_mail_by_id(mail_id)
    if not email:
        raise HTTPException(status_code=404, detail="邮件不存在")

    content = email.get("content", "") or email.get("html_content", "")
    subject = email.get("subject", "")

    result = intent_recognizer.analyze(content, subject)
    return {
        "success": True,
        "analysis": result
    }


@app.get("/api/mail/{mail_id}/suggest-reply")
async def api_mail_suggest_reply(mail_id: int, current_user: dict = Depends(login_required)):
    """AI 建议回复"""
    email = get_mail_by_id(mail_id)
    if not email:
        raise HTTPException(status_code=404, detail="邮件不存在")

    content = email.get("content", "") or email.get("html_content", "")
    sender = email.get("from_addr", "")

    # 解析发件人名称
    sender_name = ""
    if sender:
        from email.utils import parseaddr
        sender_name, _ = parseaddr(sender)

    reply = smart_replier.generate_reply(content, {"client_name": sender_name})
    return {
        "success": True,
        "suggested_reply": reply
    }

# ==================== SmartMail 路由结束 ====================

# ==================== 联系人管理模块（营销） ====================

@app.get("/contact", response_class=HTMLResponse)
async def contact_page(request: Request, current_user: dict = Depends(login_required)):
    """联系人管理页面"""
    from Sills.db_contact import get_contact_countries, get_marketing_stats
    countries = get_contact_countries()
    stats = get_marketing_stats()
    return templates.TemplateResponse("contact.html", {
        "request": request,
        "active_page": "contact",
        "current_user": current_user,
        "countries": countries,
        "stats": stats
    })


@app.get("/api/contact/list")
async def api_contact_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    cli_id: str = None,
    country: str = None,
    is_bounced: int = None,
    is_read: int = None,
    has_sent: int = None,
    current_user: dict = Depends(login_required)
):
    """获取联系人列表"""
    from Sills.db_contact import get_contact_list
    filters = {}
    if cli_id:
        filters['cli_id'] = cli_id
    if country:
        filters['country'] = country
    if is_bounced is not None:
        filters['is_bounced'] = is_bounced
    if is_read is not None:
        filters['is_read'] = is_read
    if has_sent is not None:
        filters['has_sent'] = has_sent

    items, total = get_contact_list(
        page=page,
        page_size=page_size,
        search_kw=search or "",
        filters=filters if filters else None
    )
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@app.get("/api/contact/template")
async def api_contact_template(current_user: dict = Depends(login_required)):
    """下载联系人导入模板"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "联系人导入模板"
    ws.append(['域名*', '邮箱*', '姓名', '职位', '备注'])
    ws.append(['example.com', 'zhangsan@example.com', '张三', '经理', '备注信息'])
    ws.append(['test.com', 'lisi@test.com', '李四', '总监', ''])

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contact_template.xlsx"}
    )


@app.get("/api/contact/countries")
async def api_contact_countries(current_user: dict = Depends(login_required)):
    """获取所有国家列表"""
    from Sills.db_contact import get_contact_countries
    return {"countries": get_contact_countries()}


@app.get("/api/contact/stats")
async def api_contact_stats(current_user: dict = Depends(login_required)):
    """获取营销统计数据"""
    from Sills.db_contact import get_marketing_stats
    return get_marketing_stats()


@app.get("/api/contact/{contact_id}")
async def api_contact_get(contact_id: str, current_user: dict = Depends(login_required)):
    """获取联系人详情"""
    from Sills.db_contact import get_contact_by_id
    contact = get_contact_by_id(contact_id)
    if not contact:
        raise HTTPException(status_code=404, detail="联系人不存在")
    return contact


@app.post("/api/contact/add")
async def api_contact_add(request: Request, current_user: dict = Depends(login_required)):
    """添加联系人"""
    from Sills.db_contact import add_contact
    data = await request.json()
    success, message = add_contact(data)
    return {"success": success, "message": message}


@app.post("/api/contact/update")
async def api_contact_update(request: Request, current_user: dict = Depends(login_required)):
    """更新联系人"""
    from Sills.db_contact import update_contact
    data = await request.json()
    contact_id = data.get('contact_id')
    if not contact_id:
        return {"success": False, "message": "缺少联系人ID"}

    # 移除不应更新的字段
    update_data = {k: v for k, v in data.items() if k != 'contact_id'}
    success, message = update_contact(contact_id, update_data)
    return {"success": success, "message": message}


@app.post("/api/contact/delete")
async def api_contact_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除联系人"""
    from Sills.db_contact import delete_contact
    data = await request.json()
    contact_id = data.get('contact_id')
    if not contact_id:
        return {"success": False, "message": "缺少联系人ID"}
    success, message = delete_contact(contact_id)
    return {"success": success, "message": message}


@app.post("/api/contact/batch_delete")
async def api_contact_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    """批量删除联系人"""
    from Sills.db_contact import batch_delete_contacts
    data = await request.json()
    contact_ids = data.get('contact_ids', [])
    deleted, failed, message = batch_delete_contacts(contact_ids)
    return {"success": True, "deleted": deleted, "failed": failed, "message": message}


@app.post("/api/contact/clear_all")
async def api_contact_clear_all(current_user: dict = Depends(login_required)):
    """清空所有联系人"""
    from Sills.base import get_db_connection
    with get_db_connection() as conn:
        deleted = conn.execute("SELECT COUNT(*) FROM uni_contact").fetchone()[0]
        conn.execute("DELETE FROM uni_contact")
        conn.commit()
    return {"success": True, "deleted": deleted}


@app.post("/api/contact/import")
async def api_contact_import(request: Request, current_user: dict = Depends(login_required)):
    """批量导入联系人（支持文本和JSON数组）"""
    from Sills.db_contact import batch_import_contacts
    import re
    data = await request.json()
    auto_create_cli = data.get('auto_create_cli', False)

    # 支持两种格式：contacts数组或data文本
    contacts = data.get('contacts', [])
    if not contacts and data.get('data'):
        text = data.get('data', '')

        # 第一步：清理引号和回车符
        text = text.replace('"', '').replace("'", "").replace('\r', '')

        # 第二步：将连续多个空格/制表符替换为单个制表符
        text = re.sub(r'[ \t]{2,}', '\t', text)

        # 第三步：智能合并跨行数据
        # Excel复制时，带引号的单元格会导致换行，需要合并
        lines = text.strip().split('\n')
        merged_lines = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue

            # 检查当前行是否缺少邮箱（只有域名，没有@）
            parts = line.split('\t') if '\t' in line else line.split(',')
            has_email = any('@' in p for p in parts)

            if not has_email and i + 1 < len(lines):
                # 当前行没有邮箱，检查下一行是否是邮箱
                next_line = lines[i + 1].strip()
                if '@' in next_line and ('@' in next_line.split('\t')[0] if '\t' in next_line else '@' in next_line.split(',')[0]):
                    # 下一行是邮箱，合并
                    line = line + '\t' + next_line
                    i += 2
                    merged_lines.append(line)
                    continue

            merged_lines.append(line)
            i += 1

        # 第四步：解析合并后的行
        for line in merged_lines:
            line = line.strip()
            if not line:
                continue
            parts = line.split('\t') if '\t' in line else line.split(',')
            # 清理每个字段内的换行符
            parts = [p.strip().replace('\n', '').replace('\r', '') for p in parts]

            # 需要至少2个字段，且第二个字段是有效邮箱
            if len(parts) >= 2:
                domain = parts[0] if len(parts) > 0 else ''
                email = parts[1] if len(parts) > 1 else ''

                # 验证邮箱格式
                if email and '@' in email:
                    contacts.append({
                        'domain': domain,
                        'email': email,
                        'contact_name': parts[2] if len(parts) > 2 else '',
                        'position': parts[3] if len(parts) > 3 else '',
                        'remark': parts[4] if len(parts) > 4 else ''
                    })

    if not contacts:
        return {"success": False, "message": "未提供数据"}

    success_count, errors, new_clients = batch_import_contacts(contacts, auto_create_cli)
    return {
        "success": True,
        "imported": success_count,
        "skipped": len(contacts) - success_count,
        "errors": errors[:10] if errors else [],
        "new_clients": new_clients
    }


@app.post("/api/contact/import/file")
async def api_contact_import_file(request: Request, current_user: dict = Depends(login_required)):
    """通过Excel文件批量导入联系人"""
    import io
    from fastapi import UploadFile, File
    from openpyxl import load_workbook
    from Sills.db_contact import batch_import_contacts

    print("===== 联系人文件导入开始 =====")
    print(f"用户: {current_user}")

    form = await request.form()
    print(f"FormData keys: {list(form.keys())}")
    file = form.get('file')
    if not file:
        print("错误: 未上传文件")
        return {"success": False, "message": "未上传文件"}

    print(f"文件名: {file.filename if hasattr(file, 'filename') else 'unknown'}")

    try:
        contents = await file.read()
        print(f"文件大小: {len(contents)} bytes")
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        print(f"Sheet名称: {ws.title}, 行数: {ws.max_row}, 列数: {ws.max_column}")

        # 打印表头
        header_row = [cell.value for cell in ws[1]]
        print(f"表头: {header_row}")

        contacts = []
        row_num = 2
        for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
            if not row or len(row) < 2 or not row[1]:  # 邮箱为空跳过
                row_num += 1
                continue
            # 安全获取每个字段，避免索引越界
            contact = {
                'domain': str(row[0]).strip() if len(row) > 0 and row[0] else '',
                'email': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                'contact_name': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                'position': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                'phone': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                'remark': str(row[5]).strip() if len(row) > 5 and row[5] else ''
            }
            contacts.append(contact)
            row_num += 1

        print(f"解析出的联系人数量: {len(contacts)}")
        if not contacts:
            print("错误: 文件中没有有效数据")
            return {"success": False, "message": "文件中没有有效数据"}

        print("调用 batch_import_contacts...")
        success_count, errors, new_clients = batch_import_contacts(contacts, False)
        print(f"导入结果: 成功={success_count}, 错误={errors}, 新客户={new_clients}")
        print("===== 联系人文件导入结束 =====")
        return {
            "success": True,
            "imported": success_count,
            "skipped": len(contacts) - success_count,
            "errors": errors[:10] if errors else [],
            "new_clients": new_clients
        }
    except Exception as e:
        import traceback
        print("===== 联系人导入异常 =====")
        traceback.print_exc()
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {str(e)}")
        print("===== 异常结束 =====")
        return {"success": False, "message": f"解析文件失败: {str(e)}"}


# ==================== 待开发客户(Prospect)模块 ====================

@app.get("/api/prospect/list")
async def api_prospect_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    country: str = None,
    status: str = None,
    is_public: int = None,
    current_user: dict = Depends(login_required)
):
    """获取Prospect列表"""
    from Sills.db_prospect import get_prospect_list, get_prospect_stats
    filters = {}
    if country:
        filters['country'] = country
    if status:
        filters['status'] = status
    if is_public is not None:
        filters['is_public'] = is_public

    prospects, total = get_prospect_list(page, page_size, search or "", filters)
    stats = get_prospect_stats()
    return {"success": True, "prospects": prospects, "total": total, "stats": stats}


# ========== 数据同步 API ==========
@app.post("/api/sync/start")
async def api_sync_start(current_user: dict = Depends(login_required)):
    """开始数据同步（异步执行）"""
    from Sills.db_sync import run_sync_async
    success, message = run_sync_async()
    return {"success": success, "message": message}


@app.get("/api/sync/status")
async def api_sync_status(current_user: dict = Depends(login_required)):
    """获取同步状态和进度"""
    from Sills.db_sync import get_sync_status
    return get_sync_status()


@app.post("/api/sync/stop")
async def api_sync_stop(current_user: dict = Depends(login_required)):
    """停止同步"""
    from Sills.db_sync import stop_sync
    stop_sync()
    return {"success": True, "message": "同步已停止"}


@app.get("/api/contact/template")
async def api_contact_template(current_user: dict = Depends(login_required)):
    """下载联系人导入模板"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "联系人导入模板"
    ws.append(['域名', '邮箱*', '姓名', '职位', '电话', '备注'])
    ws.append(['samsung.com', 'john@samsung.com', 'John Doe', 'Manager', '+82-10-1234-5678', '重要客户'])
    ws.append(['lg.com', 'jane@lg.com', 'Jane Smith', 'Director', '+1-555-123-4567', '潜在客户'])

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contact_template.xlsx"}
    )


@app.get("/api/prospect/template")
async def api_prospect_template(current_user: dict = Depends(login_required)):
    """下载Prospect导入模板"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io
    from urllib.parse import quote

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospect导入模板"
    # 新模板字段：客户名称、公司网站、域名、国家、主要业务、业务明细、价值分级、备注
    ws.append(['客户名称*', '公司网站', '域名*', '国家', '主要业务', '业务明细', '价值分级(1-3)', '备注'])
    ws.append(['示例公司', 'https://example.com', 'example.com', '中国', '电子元器件', '半导体分销', '2', '备注信息'])

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "prospect_template.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/prospect/stats")
async def api_prospect_stats(current_user: dict = Depends(login_required)):
    """获取Prospect统计数据"""
    from Sills.db_prospect import get_prospect_stats
    return {"success": True, **get_prospect_stats()}


@app.get("/api/prospect/countries")
async def api_prospect_countries(current_user: dict = Depends(login_required)):
    """获取Prospect国家列表"""
    from Sills.db_prospect import get_prospect_countries
    return {"success": True, "countries": get_prospect_countries()}


@app.get("/api/prospect/{prospect_id}")
async def api_prospect_get(prospect_id: str, current_user: dict = Depends(login_required)):
    """获取单个Prospect详情"""
    from Sills.db_prospect import get_prospect_by_id
    prospect = get_prospect_by_id(prospect_id)
    if not prospect:
        return {"success": False, "message": "Prospect不存在"}
    return {"success": True, "prospect": prospect}


@app.post("/api/prospect/update")
async def api_prospect_update(request: Request, current_user: dict = Depends(login_required)):
    """更新Prospect"""
    from Sills.db_prospect import update_prospect
    data = await request.json()
    prospect_id = data.get('prospect_id')
    if not prospect_id:
        return {"success": False, "message": "缺少prospect_id"}
    ok, msg = update_prospect(prospect_id, data)
    return {"success": ok, "message": msg}


@app.post("/api/prospect/delete")
async def api_prospect_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除Prospect"""
    from Sills.db_prospect import delete_prospect
    data = await request.json()
    prospect_id = data.get('prospect_id')
    if not prospect_id:
        return {"success": False, "message": "缺少prospect_id"}
    ok, msg = delete_prospect(prospect_id)
    return {"success": ok, "message": msg}


@app.post("/api/prospect/batch_delete")
async def api_prospect_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    """批量删除Prospect"""
    from Sills.db_prospect import batch_delete_prospects
    print("===== Prospect批量删除开始 =====")
    data = await request.json()
    print(f"请求数据: {data}")
    prospect_ids = data.get('prospect_ids', [])
    print(f"要删除的ID列表: {prospect_ids}")
    if not prospect_ids:
        print("错误: 未选择任何记录")
        return {"success": False, "message": "未选择任何记录"}

    print("调用 batch_delete_prospects...")
    ok, msg = batch_delete_prospects(prospect_ids)
    print(f"删除结果: ok={ok}, msg={msg}")

    if msg is None or msg == '':
        msg = "删除操作返回空消息"
        print(f"消息为空，设置为: {msg}")

    print(f"返回: success={ok}, message={msg}")
    print("===== Prospect批量删除结束 =====")
    return {"success": ok, "message": msg}


@app.post("/api/prospect/clear_all")
async def api_prospect_clear_all(current_user: dict = Depends(login_required)):
    """清空所有Prospect"""
    from Sills.base import get_db_connection
    with get_db_connection() as conn:
        deleted = conn.execute("SELECT COUNT(*) FROM uni_prospect").fetchone()[0]
        conn.execute("DELETE FROM uni_prospect")
        conn.commit()
    return {"success": True, "deleted": deleted}


@app.post("/api/prospect/add")
async def api_prospect_add(request: Request, current_user: dict = Depends(login_required)):
    """添加Prospect"""
    from Sills.db_prospect import add_prospect
    data = await request.json()
    ok, msg = add_prospect(data)
    return {"success": ok, "message": msg}


@app.post("/api/prospect/import")
async def api_prospect_import(request: Request, current_user: dict = Depends(login_required)):
    """批量导入Prospect"""
    from Sills.db_prospect import import_prospects
    from openpyxl import load_workbook
    import io

    print("===== Prospect文件导入开始 =====")
    print(f"用户: {current_user}")

    form = await request.form()
    print(f"FormData keys: {list(form.keys())}")
    file = form.get('file')
    if not file:
        print("错误: 未上传文件")
        return {"success": False, "message": "未上传文件"}

    print(f"文件名: {file.filename if hasattr(file, 'filename') else 'unknown'}")

    try:
        content = await file.read()
        print(f"文件大小: {len(content)} bytes")
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        print(f"Sheet名称: {ws.title}, 行数: {ws.max_row}, 列数: {ws.max_column}")

        # 打印表头
        header_row = [cell.value for cell in ws[1]]
        print(f"表头: {header_row}")

        # 解析Excel数据
        data_list = []
        row_num = 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"第{row_num}行数据: {row}")
            if not row or not row[0]:
                print(f"  -> 跳过: 数据为空或客户名称为空")
                row_num += 1
                continue
            item = {
                'prospect_name': str(row[0] or '').strip(),
                'company_website': str(row[1] or '').strip(),
                'domain': str(row[2] or '').strip(),
                'country': str(row[3] or '').strip(),
                'business_type': str(row[4] or '').strip(),
                'business_detail': str(row[5] or '').strip(),
                'value_level': row[6] if row[6] else 0,
                'remark': str(row[7] or '').strip()
            }
            print(f"  -> 解析结果: {item}")
            data_list.append(item)
            row_num += 1

        print(f"解析出的数据数量: {len(data_list)}")
        if not data_list:
            print("错误: 文件中没有有效数据")
            return {"success": False, "message": "文件中没有有效数据"}

        print("调用 import_prospects...")
        imported, skipped, errors = import_prospects(data_list)
        print(f"导入结果: 成功={imported}, 跳过={skipped}, 错误={errors}")
        print("===== Prospect文件导入结束 =====")
        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:10] if errors else []
        }
    except Exception as e:
        import traceback
        print("===== Prospect导入异常 =====")
        traceback.print_exc()
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {str(e)}")
        print("===== 异常结束 =====")
        return {"success": False, "message": f"解析文件失败: {str(e)}"}


@app.post("/api/prospect/import/text")
async def api_prospect_import_text(request: Request, current_user: dict = Depends(login_required)):
    """批量导入Prospect (文本格式)"""
    from Sills.db_prospect import import_prospects

    data = await request.json()
    data_list = data.get('data_list', [])

    if not data_list:
        return {"success": False, "message": "未提供数据"}

    imported, skipped, errors = import_prospects(data_list)
    return {
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10] if errors else []
    }


@app.post("/api/prospect/convert")
async def api_prospect_convert(request: Request, current_user: dict = Depends(login_required)):
    """将Prospect转化为CLI客户"""
    from Sills.db_prospect import convert_prospect_to_cli
    data = await request.json()
    prospect_id = data.get('prospect_id')
    if not prospect_id:
        return {"success": False, "message": "缺少prospect_id"}

    ok, msg = convert_prospect_to_cli(prospect_id)
    return {"success": ok, "message": msg}


@app.post("/api/prospect/refresh_counts")
async def api_prospect_refresh_counts(current_user: dict = Depends(login_required)):
    """刷新所有Prospect的关联联系人数量"""
    from Sills.db_prospect import refresh_all_contact_counts
    updated = refresh_all_contact_counts()
    return {"success": True, "message": f"已更新 {updated} 条记录的联系人数量"}


# ==================== 开发信管理模块 (Email Task Manager) ====================
# 替换原有Marketing模块,支持任务管理、联系人组、发件人账号等功能

from Sills.db_contact_group import (
    get_group_list, get_group_by_id, add_group, update_group, delete_group,
    get_group_contacts, get_all_groups_contacts
)
from Sills.db_email_account import (
    get_account_list, get_account_by_id, get_account_by_email,
    add_account, update_account, delete_account, reset_daily_count,
    can_send_today, get_smtp_server_for_email
)
from Sills.db_email_task import (
    get_task_list, get_task_by_id, get_active_task, has_running_task,
    create_task, start_task, update_task_progress, cancel_task,
    complete_task, get_task_progress, get_task_contacts,
    delete_task, delete_tasks_batch
)
from Sills.db_email_template import (
    get_template_list, get_template_by_id, create_template,
    update_template, delete_template, delete_templates_batch
)
from Sills.db_email_log import (
    get_task_logs, get_failed_logs, get_task_stats
)
from Sills.email_sender import (
    start_email_worker, send_test_email
)


# ==================== 联系人组管理 ====================

@app.get("/api/group/list")
async def api_group_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取联系人组列表"""
    groups, total = get_group_list(page, page_size, search or "")
    return {"success": True, "groups": groups, "total": total, "page": page, "page_size": page_size}


@app.post("/api/group/add")
async def api_group_add(request: Request, current_user: dict = Depends(login_required)):
    """添加联系人组"""
    data = await request.json()
    group_name = data.get('group_name', '')
    filter_criteria = data.get('filter_criteria', {})

    success, message = add_group(group_name, filter_criteria)
    return {"success": success, "message": message}


@app.post("/api/group/delete")
async def api_group_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除联系人组"""
    data = await request.json()
    group_id = data.get('group_id', '')

    success, message = delete_group(group_id)
    return {"success": success, "message": message}


@app.post("/api/group/update")
async def api_group_update(request: Request, current_user: dict = Depends(login_required)):
    """更新联系人组"""
    data = await request.json()
    group_id = data.get('group_id', '')
    group_name = data.get('group_name', '')
    filter_criteria = data.get('filter_criteria', {})

    success, message = update_group(group_id, group_name, filter_criteria)
    return {"success": success, "message": message}


@app.get("/api/group/{group_id}/contacts")
async def api_group_contacts(
    group_id: str,
    page: int = 1,
    page_size: int = 100,
    current_user: dict = Depends(login_required)
):
    """预览组内联系人"""
    contacts, total = get_group_contacts(group_id, page, page_size)
    return {"success": True, "contacts": contacts, "total": total}


@app.get("/api/group/{group_id}/export")
async def api_group_export(group_id: str, current_user: dict = Depends(login_required)):
    """导出联系人组为Excel"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io
    from Sills.db_contact_group import get_group_by_id, get_group_contacts_all_types

    group = get_group_by_id(group_id)
    if not group:
        return {"success": False, "message": "组不存在"}

    # 使用统一的 get_group_contacts_all_types 函数，支持静态组和动态组
    contacts, _ = get_group_contacts_all_types(group_id, page_size=10000)  # 获取全部

    wb = Workbook()
    ws = wb.active
    ws.title = group.get('group_name', '联系人')
    ws.append(['联系人ID', '客户ID', '邮箱', '域名', '姓名', '国家', '职位',
               '电话', '公司', '是否退信', '是否已读', '发送次数', '退信次数',
               '已读次数', '最后发送时间', '备注', '待开发客户名'])

    for c in contacts:
        ws.append([
            c.get('contact_id', ''),
            c.get('cli_id', ''),
            c.get('email', ''),
            c.get('domain', ''),
            c.get('contact_name', ''),
            c.get('country', ''),
            c.get('position', ''),
            c.get('phone', ''),
            c.get('company', ''),
            '是' if c.get('is_bounced') else '否',
            '是' if c.get('is_read') else '否',
            c.get('send_count', 0),
            c.get('bounce_count', 0),
            c.get('read_count', 0),
            c.get('last_sent_at', ''),
            c.get('remark', ''),
            c.get('prospect_name', '')
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{group.get('group_name', 'group')}_contacts.xlsx"
    # 使用 URL 编码处理中文文件名
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ==================== 发件人账号管理 ====================

@app.get("/api/account/list")
async def api_account_list(
    page: int = 1,
    page_size: int = 20,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取发件人账号列表"""
    accounts, total = get_account_list(page, page_size, search or "")
    return {"success": True, "accounts": accounts, "total": total}


@app.post("/api/account/add")
async def api_account_add(request: Request, current_user: dict = Depends(login_required)):
    """添加发件人账号"""
    data = await request.json()
    email = data.get('email', '')
    password = data.get('password', '')
    smtp_server = data.get('smtp_server', None)
    daily_limit = data.get('daily_limit', 1800)

    success, message = add_account(email, password, smtp_server, daily_limit)
    return {"success": success, "message": message}


@app.post("/api/account/update")
async def api_account_update(request: Request, current_user: dict = Depends(login_required)):
    """更新发件人账号"""
    data = await request.json()
    account_id = data.get('account_id', '')
    password = data.get('password', None)
    smtp_server = data.get('smtp_server', None)
    daily_limit = data.get('daily_limit', None)

    success, message = update_account(account_id, password, smtp_server, daily_limit)
    return {"success": success, "message": message}


@app.post("/api/account/delete")
async def api_account_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除发件人账号"""
    data = await request.json()
    account_id = data.get('account_id', '')

    success, message = delete_account(account_id)
    return {"success": success, "message": message}


@app.post("/api/account/test")
async def api_account_test(request: Request, current_user: dict = Depends(login_required)):
    """发送测试邮件"""
    data = await request.json()
    account_id = data.get('account_id', '')
    to_email = data.get('to_email', '')

    success, message = send_test_email(account_id, to_email)
    return {"success": success, "message": message}


# ==================== 邮件任务管理 ====================

@app.get("/api/task/list")
async def api_task_list(
    page: int = 1,
    page_size: int = 20,
    status: str = None,
    search: str = None,
    current_user: dict = Depends(login_required)
):
    """获取邮件任务列表"""
    tasks, total = get_task_list(page, page_size, status or "", search or "")
    return {"success": True, "tasks": tasks, "total": total, "page": page, "page_size": page_size}


@app.get("/api/task/active")
async def api_task_active(current_user: dict = Depends(login_required)):
    """获取当前活跃任务"""
    task = get_active_task()
    if task:
        # 解密密码供前端显示(可选)
        progress = get_task_progress(task['task_id'])
        return {"success": True, "task": task, "progress": progress}
    return {"success": True, "task": None}


@app.post("/api/task/create")
async def api_task_create(request: Request, current_user: dict = Depends(login_required)):
    """创建邮件任务"""
    data = await request.json()

    task_name = data.get('task_name', '')
    account_ids = data.get('account_ids', [])    # 多账号ID列表（支持轮换）
    group_ids = data.get('group_ids', [])
    subject = data.get('subject', '')
    body = data.get('body', '')
    placeholders = data.get('placeholders', None)
    schedule_start = data.get('schedule_start', None)
    schedule_end = data.get('schedule_end', None)
    send_interval = data.get('send_interval', 2)  # 发送间隔（秒）
    skip_enabled = data.get('skip_enabled', 1)    # 是否启用跳过规则（默认开启）
    skip_days = data.get('skip_days', 7)          # 跳过天数（默认7天）
    daily_limit_per_account = data.get('daily_limit_per_account', 1800)  # 单账号日发送上限

    success, result = create_task(
        task_name, account_ids, group_ids, subject, body,
        placeholders, schedule_start, schedule_end, send_interval,
        skip_enabled, skip_days, daily_limit_per_account
    )

    if success:
        # 创建成功，不自动启动，等待用户手动执行
        return {"success": True, "task_id": result}
    else:
        return {"success": False, "message": result}


@app.post("/api/task/start")
async def api_task_start(request: Request, current_user: dict = Depends(login_required)):
    """手动启动任务执行"""
    data = await request.json()
    task_id = data.get('task_id', '')

    if not task_id:
        return {"success": False, "message": "任务ID不能为空"}

    # 检查任务状态
    task = get_task_by_id(task_id)
    if not task:
        return {"success": False, "message": "任务不存在"}

    if task.get('status') == 'running':
        return {"success": False, "message": "任务已在执行中"}

    if task.get('status') == 'completed':
        return {"success": False, "message": "任务已完成，无法重新执行"}

    # 启动任务
    start_task(task_id)
    start_email_worker(task_id)
    return {"success": True, "message": "任务已启动"}


@app.post("/api/task/cancel")
async def api_task_cancel(request: Request, current_user: dict = Depends(login_required)):
    """取消任务（暂停执行）"""
    data = await request.json()
    task_id = data.get('task_id', '')

    success, message = cancel_task(task_id)
    return {"success": success, "message": message}


@app.post("/api/task/delete")
async def api_task_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除单个任务"""
    data = await request.json()
    task_id = data.get('task_id', '')

    if not task_id:
        return {"success": False, "message": "任务ID不能为空"}

    success, message = delete_task(task_id)
    return {"success": success, "message": message}


@app.post("/api/task/update-account")
async def api_task_update_account(request: Request, current_user: dict = Depends(login_required)):
    """更新任务发件人账号（仅非执行状态可用）"""
    from Sills.db_email_task import update_task_account
    data = await request.json()
    task_id = data.get('task_id', '')
    new_account_id = data.get('account_id', '')

    if not task_id or not new_account_id:
        return {"success": False, "message": "任务ID和账号ID不能为空"}

    success, message = update_task_account(task_id, new_account_id)
    return {"success": success, "message": message}


@app.post("/api/task/batch-delete")
async def api_task_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    """批量删除任务"""
    data = await request.json()
    task_ids = data.get('task_ids', [])

    if not task_ids or len(task_ids) == 0:
        return {"success": False, "message": "请选择要删除的任务"}

    success_count, failed_list = delete_tasks_batch(task_ids)
    return {
        "success": True,
        "message": f"成功删除 {success_count} 个任务",
        "success_count": success_count,
        "failed_list": failed_list
    }


@app.get("/api/task/{task_id}/progress")
async def api_task_progress(task_id: str, current_user: dict = Depends(login_required)):
    """获取任务实时进度"""
    progress = get_task_progress(task_id)
    if progress:
        stats = get_task_stats(task_id)
        return {"success": True, "progress": progress, "stats": stats}
    return {"success": False, "message": "任务不存在"}


@app.get("/api/task/{task_id}/logs")
async def api_task_logs(
    task_id: str,
    page: int = 1,
    page_size: int = 50,
    current_user: dict = Depends(login_required)
):
    """获取任务发送日志"""
    logs, total = get_task_logs(task_id, page, page_size)
    return {"success": True, "logs": logs, "total": total, "page": page, "page_size": page_size}


@app.get("/api/task/{task_id}/failed")
async def api_task_failed(task_id: str, current_user: dict = Depends(login_required)):
    """获取任务失败日志"""
    failed = get_failed_logs(task_id)
    return {"success": True, "failed": failed}


@app.post("/api/task/retry")
async def api_task_retry(request: Request, current_user: dict = Depends(login_required)):
    """重试任务中发送失败的邮件"""
    from Sills.db_email_task import retry_failed_task, get_task_by_id
    from Sills.email_sender import start_email_worker

    data = await request.json()
    task_id = data.get('task_id', '')

    success, message = retry_failed_task(task_id)
    if success:
        # 启动后台Worker（重试模式）
        start_email_worker(task_id, retry_mode=True)
        return {"success": True, "message": message}
    return {"success": False, "message": message}


@app.get("/api/task/{task_id}/stats")
async def api_task_stats(task_id: str, current_user: dict = Depends(login_required)):
    """获取任务完整统计信息"""
    from Sills.db_email_task import get_task_full_stats
    stats = get_task_full_stats(task_id)
    if stats:
        return {"success": True, "stats": stats}
    return {"success": False, "message": "任务不存在"}


@app.get("/api/task/{task_id}/export")
async def api_task_export(task_id: str, current_user: dict = Depends(login_required)):
    """导出任务联系人发送状态到Excel（直接下载）"""
    from Sills.db_email_log import export_task_contacts_to_excel
    from fastapi.responses import FileResponse
    import os
    import tempfile

    # 使用临时目录生成文件
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"task_{task_id}_contacts_{timestamp}.xlsx"

    # 保存到临时目录，浏览器会自动下载到用户指定目录
    temp_dir = tempfile.gettempdir()
    output_path = os.path.join(temp_dir, filename)

    success, result = export_task_contacts_to_excel(task_id, output_path)
    if success:
        # 直接返回文件供下载
        return FileResponse(
            path=output_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return {"success": False, "message": result}


# ==================== 邮件模板管理 ====================

@app.get("/api/template/list")
async def api_template_list(current_user: dict = Depends(login_required)):
    """获取邮件模板列表（当前用户创建的模板）"""
    emp_id = current_user.get('emp_id', '')
    templates = get_template_list(emp_id)
    return {"success": True, "templates": templates}


@app.get("/api/template/{template_id}")
async def api_template_get(template_id: str, current_user: dict = Depends(login_required)):
    """获取单个模板详情"""
    template = get_template_by_id(template_id)
    if template:
        return {"success": True, "template": template}
    return {"success": False, "message": "模板不存在"}


@app.post("/api/template/add")
async def api_template_add(request: Request, current_user: dict = Depends(login_required)):
    """创建邮件模板"""
    data = await request.json()
    template_name = data.get('template_name', '')
    subject = data.get('subject', '')
    body = data.get('body', '')
    emp_id = current_user.get('emp_id', '')

    success, message = create_template(template_name, subject, body, emp_id)
    return {"success": success, "message": message}


@app.post("/api/template/update")
async def api_template_update(request: Request, current_user: dict = Depends(login_required)):
    """更新邮件模板"""
    data = await request.json()
    template_id = data.get('template_id', '')
    template_name = data.get('template_name', '')
    subject = data.get('subject', '')
    body = data.get('body', '')

    success, message = update_template(template_id, template_name, subject, body)
    return {"success": success, "message": message}


@app.post("/api/template/delete")
async def api_template_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除邮件模板"""
    data = await request.json()
    template_id = data.get('template_id', '')

    success, message = delete_template(template_id)
    return {"success": success, "message": message}


@app.post("/api/template/batch-delete")
async def api_template_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    """批量删除邮件模板"""
    data = await request.json()
    template_ids = data.get('template_ids', [])

    if not template_ids or len(template_ids) == 0:
        return {"success": False, "message": "请选择要删除的模板"}

    success_count, failed_list = delete_templates_batch(template_ids)
    return {
        "success": True,
        "message": f"成功删除 {success_count} 个模板",
        "success_count": success_count,
        "failed_list": failed_list
    }


@app.post("/api/group/add-static")
async def api_group_add_static(request: Request, current_user: dict = Depends(login_required)):
    """添加静态邮件组（手动邮件列表）"""
    from Sills.db_contact_group import add_static_group

    data = await request.json()
    group_name = data.get('group_name', '')
    email_list = data.get('email_list', [])  # [{"email": "x@x.com", "company": "公司名"}, ...]
    description = data.get('description', '')

    success, message = add_static_group(group_name, email_list, description)
    return {"success": success, "message": message}


@app.post("/api/group/create-from-contacts")
async def api_group_create_from_contacts(request: Request, current_user: dict = Depends(login_required)):
    """从现有联系人组创建静态邮件组（用于测试）"""
    from Sills.db_contact_group import add_static_group, get_group_contacts_all_types

    data = await request.json()
    group_name = data.get('group_name', '')
    source_group_ids = data.get('source_group_ids', [])  # 源组ID列表
    description = data.get('description', '')

    if not source_group_ids:
        return {"success": False, "message": "请选择至少一个源联系人组"}

    # 从源组获取联系人
    contacts = []
    for group_id in source_group_ids:
        group_contacts, _ = get_group_contacts_all_types(group_id)
        for c in group_contacts:
            contacts.append({
                'email': c.get('email', ''),
                'company': c.get('company', ''),
                'contact_name': c.get('contact_name', '')
            })

    if not contacts:
        return {"success": False, "message": "源组中没有联系人"}

    success, message = add_static_group(group_name, contacts, description)
    return {"success": success, "message": message}


@app.get("/api/group/{group_id}/contacts-all")
async def api_group_contacts_all(
    group_id: str,
    page: int = 1,
    page_size: int = 100,
    current_user: dict = Depends(login_required)
):
    """获取联系人组内联系人（支持动态+手动邮件合并）"""
    from Sills.db_contact_group import get_group_contacts_all_types
    contacts, total = get_group_contacts_all_types(group_id, page, page_size)
    return {"success": True, "contacts": contacts, "total": total, "page": page, "page_size": page_size}


@app.get("/api/group/{group_id}/manual-emails")
async def api_group_manual_emails(group_id: str, current_user: dict = Depends(login_required)):
    """获取联系人组的手动邮件列表"""
    from Sills.db_contact_group import get_group_manual_emails
    emails = get_group_manual_emails(group_id)
    return {"success": True, "emails": emails}


@app.post("/api/group/{group_id}/add-emails")
async def api_group_add_emails(
    group_id: str,
    request: Request,
    current_user: dict = Depends(login_required)
):
    """向联系人组添加手动邮件（数量不限）"""
    from Sills.db_contact_group import add_manual_emails_to_group

    data = await request.json()
    emails = data.get('emails', [])  # [{"email": "x@x.com", "company": "公司名"}, ...]

    if not emails:
        return {"success": False, "message": "请提供要添加的邮件列表"}

    success, message = add_manual_emails_to_group(group_id, emails)
    return {"success": success, "message": message}


@app.post("/api/group/{group_id}/remove-email")
async def api_group_remove_email(
    group_id: str,
    request: Request,
    current_user: dict = Depends(login_required)
):
    """从联系人组移除手动邮件"""
    from Sills.db_contact_group import remove_manual_email_from_group

    data = await request.json()
    email = data.get('email', '')

    if not email:
        return {"success": False, "message": "请提供要移除的邮箱地址"}

    success, message = remove_manual_email_from_group(group_id, email)
    return {"success": success, "message": message}


@app.post("/api/group/update-with-emails")
async def api_group_update_with_emails(request: Request, current_user: dict = Depends(login_required)):
    """更新联系人组（支持筛选条件 + 手动邮件）"""
    data = await request.json()
    group_id = data.get('group_id', '')
    group_name = data.get('group_name', '')
    filter_criteria = data.get('filter_criteria', {})
    manual_emails = data.get('manual_emails', None)  # [{"email": "x@x.com", "company": "公司名"}, ...]

    success, message = update_group(group_id, group_name, filter_criteria, manual_emails)
    return {"success": success, "message": message}

# ==================== 开发信管理模块结束 ====================

if __name__ == "__main__":
    # 根据环境选择端口: Windows=8001, WSL=8000
    env = get_server_env()
    port = 8001 if env == "Windows" else 8000
    uvicorn.run("main:app", host="127.0.0.1", port=port, reload=True)


# ==================== Gemini AI API ====================

@app.get("/api/gemini/config")
async def api_get_gemini_config(current_user: dict = Depends(login_required)):
    """获取 Gemini 配置状态"""
    from Sills.gemini_service import is_gemini_configured, get_gemini_api_key
    return {
        "success": True,
        "configured": is_gemini_configured(),
        "api_key": get_gemini_api_key()
    }


@app.post("/api/gemini/config")
async def api_set_gemini_config(request: Request, current_user: dict = Depends(login_required)):
    """设置 Gemini API Key"""
    from Sills.gemini_service import set_gemini_api_key_permanent
    data = await request.json()
    api_key = data.get('api_key', '').strip()

    if not api_key:
        return {"success": False, "message": "API Key 不能为空"}

    result = set_gemini_api_key_permanent(api_key)

    if result["windows"] or result["wsl"]:
        return {
            "success": True,
            "message": "Gemini API Key 设置成功",
            "details": result
        }
    else:
        return {
            "success": False,
            "message": result.get("message", "设置失败")
        }


@app.post("/api/gemini/suggest-reply")
async def api_gemini_suggest_reply(request: Request, current_user: dict = Depends(login_required)):
    """
    Gemini AI 建议回复
    需要提供：邮件内容 + 用户回复意图
    """
    from Sills.gemini_service import suggest_email_reply
    data = await request.json()

    email_content = data.get('email_content', '')
    user_instruction = data.get('user_instruction', '')
    sender_name = data.get('sender_name', '')
    email_subject = data.get('email_subject', '')

    if not email_content:
        return {"success": False, "message": "邮件内容不能为空"}
    if not user_instruction:
        return {"success": False, "message": "请输入您想要回复的内容或方向"}

    result = suggest_email_reply(
        email_content=email_content,
        user_instruction=user_instruction,
        sender_name=sender_name,
        email_subject=email_subject
    )

    return result


# ==================== 数据中心模块路由 ====================

@app.get("/datacenter", response_class=HTMLResponse)
async def datacenter_page(request: Request, current_user: dict = Depends(login_required)):
    """数据中心页面 - 仅管理员可访问"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return templates.TemplateResponse("datacenter.html", {
        "request": request,
        "active_page": "datacenter",
        "current_user": current_user
    })


@app.get("/api/datacenter/tables")
async def api_datacenter_tables(current_user: dict = Depends(login_required)):
    """获取所有数据库表"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import get_all_tables
    return get_all_tables()


@app.get("/api/datacenter/tables/{table_name}/structure")
async def api_datacenter_table_structure(table_name: str, current_user: dict = Depends(login_required)):
    """获取表结构"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import get_table_structure
    return get_table_structure(table_name)


@app.get("/api/datacenter/tables/{table_name}/data")
async def api_datacenter_table_data(table_name: str, current_user: dict = Depends(login_required)):
    """获取表数据"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import get_table_data
    data, columns = get_table_data(table_name)
    return {"data": data, "columns": columns}


@app.post("/api/datacenter/execute")
async def api_datacenter_execute(request: Request, current_user: dict = Depends(login_required)):
    """执行 SQL"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import execute_sql
    data = await request.json()
    sql = data.get("sql", "")
    if not sql:
        return {"success": False, "message": "SQL 不能为空", "data": []}

    success, result, message = execute_sql(sql)
    return {"success": success, "data": result if success else [], "message": message, "error": "" if success else str(result)}


@app.post("/api/datacenter/queries")
async def api_datacenter_save_query(request: Request, current_user: dict = Depends(login_required)):
    """保存查询"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import save_query
    data = await request.json()
    name = data.get("name", "").strip()
    sql = data.get("sql", "").strip()
    if not name or not sql:
        return {"success": False, "message": "名称和 SQL 不能为空"}

    success, message = save_query(name, sql, current_user.get("account"))
    return {"success": success, "message": message}


@app.get("/api/datacenter/queries")
async def api_datacenter_get_queries(current_user: dict = Depends(login_required)):
    """获取保存的查询"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import get_saved_queries
    return get_saved_queries()


@app.delete("/api/datacenter/queries/{query_id}")
async def api_datacenter_delete_query(query_id: int, current_user: dict = Depends(login_required)):
    """删除保存的查询"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import delete_saved_query
    success, message = delete_saved_query(query_id)
    return {"success": success, "message": message}


@app.post("/api/datacenter/export")
async def api_datacenter_export(request: Request, current_user: dict = Depends(login_required)):
    """导出 Excel"""
    if current_user.get("rule") != "3":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    from Sills.db_datacenter import export_to_excel
    from fastapi.responses import Response

    data = await request.json()
    result_data = data.get("data", [])
    columns = data.get("columns", [])
    selected_fields = data.get("selected_fields", None)

    # 限制最大导出行数
    if len(result_data) > 10000:
        return {"success": False, "message": "导出数据不能超过 10000 行"}

    excel_bytes = export_to_excel(result_data, columns, selected_fields)

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=data_export.xlsx"}
    )

    return result


# ============ 银行流水管理（财务管理模块） ============

@app.get("/bank", response_class=HTMLResponse)
async def bank_page(request: Request, current_user: dict = Depends(login_required)):
    """银行流水管理页面"""
    return templates.TemplateResponse("bank.html", {
        "request": request,
        "current_user": current_user
    })


@app.get("/bank/import", response_class=HTMLResponse)
async def bank_import_page(request: Request, current_user: dict = Depends(login_required)):
    """银行流水导入页面"""
    return templates.TemplateResponse("bank_import.html", {
        "request": request,
        "current_user": current_user
    })


@app.get("/api/bank/list")
async def api_bank_list(
    page: int = 1,
    page_size: int = 20,
    start_date: str = "",
    end_date: str = "",
    transaction_type: str = "",
    is_matched: str = "",
    payer_name: str = "",
    min_amount: str = "",
    max_amount: str = "",
    import_batch: str = "",
    current_user: dict = Depends(login_required)
):
    """获取银行流水列表"""
    results, total = get_transaction_list(
        page=page,
        page_size=page_size,
        start_date=start_date,
        end_date=end_date,
        transaction_type=transaction_type,
        is_matched=is_matched,
        payer_name=payer_name,
        min_amount=min_amount,
        max_amount=max_amount,
        import_batch=import_batch
    )
    return {
        "items": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


@app.get("/api/bank/template")
async def api_bank_template(current_user: dict = Depends(login_required)):
    """下载银行流水导入模板"""
    import pandas as pd
    from fastapi.responses import StreamingResponse

    # 创建模板DataFrame
    columns = ['交易时间', '交易编号', '记账流水号', '交易类型', '交易详情', '币种',
               '交易金额', '总余额', '付款人名称', '付款银行', '付款账号',
               '收款人名称', '收款银行', '收款账号', '收款人备注名', '附言', '重复']

    # 示例数据（帮助用户理解格式）
    sample_data = [
        ['2026-04-20 17:33:20', '81990102355000026042005520183', '8990100034000026042072297477',
         '手续费', '银行账户管理费', 'CNY', 41.09, 0.00, '', '', '', '', '', '', '', '', ''],
        ['2026-04-20 17:33:20', '81990102355000026042005520184', '8990100034000026042072297475',
         '支出', '转账-货款支付', 'CNY', 10229.95, 41.09, '深圳某科技有限公司', '', '',
         '收款方公司', '招商银行', '6225880123456789', '', '备注信息', ''],
        ['2026-04-20 17:33:19', '10990102137000026042007242615', '8990100034000026042072297471',
         '收入', '收入-货款收款', 'CNY', 10271.04, 10271.04, '', '', '',
         '付款方公司名', '工商银行', '1234567890', '', '', ''],
    ]

    df = pd.DataFrame(sample_data, columns=columns)

    # 导出为Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='银行流水模板')
    output.seek(0)

    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bank_transaction_template.xlsx"}
    )


@app.get("/api/bank/batches")
async def api_bank_batches(current_user: dict = Depends(login_required)):
    """获取导入批次列表"""
    batches = get_batch_list()
    return {"items": batches}


@app.get("/api/bank/{transaction_id}")
async def api_bank_detail(transaction_id: str, current_user: dict = Depends(login_required)):
    """获取流水详情"""
    tx = get_transaction_by_id(transaction_id)
    if not tx:
        raise HTTPException(status_code=404, detail="流水不存在")
    return tx


@app.post("/api/bank/import")
async def api_bank_import(request: Request, current_user: dict = Depends(login_required)):
    """Excel批量导入银行流水"""
    data = await request.json()
    rows = data.get("rows", [])
    source_file = data.get("source_file", "")

    if not rows:
        return {"success": False, "message": "无数据"}

    success_count, errors, batch_id = batch_import_transactions(rows, source_file)

    return {
        "success": success_count > 0,
        "message": f"成功导入 {success_count} 条流水" + (f"，失败 {len(errors)} 条" if errors else ""),
        "batch_id": batch_id,
        "errors": errors
    }


@app.post("/api/bank/upload")
async def api_bank_upload(
    file: UploadFile = File(...),
    current_user: dict = Depends(login_required)
):
    """上传Excel文件并解析"""
    import pandas as pd

    try:
        # 读取Excel文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # 字段映射（Excel列名 -> 数据库字段名）
        column_mapping = {
            '交易时间': 'transaction_time',
            '交易编号': 'transaction_no',
            '记账流水号': 'ledger_no',
            '交易类型': 'transaction_type',
            '交易详情': 'transaction_detail',
            '币种': 'currency',
            '交易金额': 'transaction_amount',
            '总余额': 'balance',
            '付款人名称': 'payer_name',
            '付款银行': 'payer_bank',
            '付款账号': 'payer_account',
            '收款人名称': 'payee_name',
            '收款银行': 'payee_bank',
            '收款账号': 'payee_account',
            '收款人备注名': 'payee_remark_name',
            '附言': 'remark_text',
            '重复': 'duplicate_flag'
        }

        # 转换列名
        df.columns = [column_mapping.get(col, col) for col in df.columns]

        # 转换为字典列表
        rows = []
        for _, row in df.iterrows():
            row_dict = {}
            for col in df.columns:
                value = row[col]
                if pd.isna(value):
                    row_dict[col] = ''
                elif col == 'transaction_amount' or col == 'balance':
                    row_dict[col] = float(value) if value else 0
                else:
                    row_dict[col] = str(value)
            rows.append(row_dict)

        return {
            "success": True,
            "rows": rows,
            "count": len(rows),
            "source_file": file.filename
        }

    except Exception as e:
        return {"success": False, "message": f"解析失败：{str(e)}"}


@app.post("/api/bank/update")
async def api_bank_update(request: Request, current_user: dict = Depends(login_required)):
    """更新流水信息（仅备注）"""
    data = await request.json()
    transaction_id = data.get("transaction_id")
    if not transaction_id:
        return {"success": False, "message": "缺少流水ID"}

    success, message = update_transaction(transaction_id, data)
    return {"success": success, "message": message}


@app.post("/api/bank/delete")
async def api_bank_delete(request: Request, current_user: dict = Depends(login_required)):
    """删除单条流水"""
    data = await request.json()
    transaction_id = data.get("transaction_id")
    if not transaction_id:
        return {"success": False, "message": "缺少流水ID"}

    success, message = delete_transaction(transaction_id)
    return {"success": success, "message": message}


@app.post("/api/bank/batch_delete")
async def api_bank_batch_delete(request: Request, current_user: dict = Depends(login_required)):
    """按批次删除流水"""
    data = await request.json()
    import_batch = data.get("import_batch")
    if not import_batch:
        return {"success": False, "message": "缺少批次号"}

    success, message = batch_delete_by_batch(import_batch)
    return {"success": success, "message": message}


@app.post("/api/bank/batch_delete_selected")
async def api_bank_batch_delete_selected(request: Request, current_user: dict = Depends(login_required)):
    """批量删除勾选的流水"""
    data = await request.json()
    transaction_ids = data.get("transaction_ids", [])
    if not transaction_ids:
        return {"success": False, "message": "未选择任何流水"}

    success, message = batch_delete_selected(transaction_ids)
    return {"success": success, "message": message}


# ============ 台账关联管理 ============

@app.get("/api/ledger/by_transaction/{transaction_id}")
async def api_ledger_by_transaction(transaction_id: str, current_user: dict = Depends(login_required)):
    """查询流水关联的所有订单"""
    ledgers = get_ledger_by_transaction(transaction_id)
    return {"items": ledgers}


@app.get("/api/ledger/by_order/{manager_id}")
async def api_ledger_by_order(manager_id: str, current_user: dict = Depends(login_required)):
    """查询订单关联的所有流水"""
    ledgers, total_received = get_ledger_by_manager(manager_id)
    return {"items": ledgers, "total_received": total_received}


@app.get("/api/ledger/summary/{manager_id}")
async def api_ledger_summary(manager_id: str, current_user: dict = Depends(login_required)):
    """获取订单收款摘要"""
    summary = get_ledger_summary(manager_id)
    return summary


@app.post("/api/ledger/link")
async def api_ledger_link(request: Request, current_user: dict = Depends(login_required)):
    """创建流水与订单的关联"""
    data = await request.json()
    transaction_id = data.get("transaction_id")
    manager_id = data.get("manager_id")
    allocation_amount = data.get("allocation_amount")

    if not transaction_id or not manager_id or not allocation_amount:
        return {"success": False, "message": "缺少必要参数"}

    success, result = create_ledger(
        transaction_id=transaction_id,
        manager_id=manager_id,
        allocation_amount=float(allocation_amount),
        match_type=data.get("match_type", "manual"),
        created_by=current_user.get("emp_id"),
        remark=data.get("remark", "")
    )

    if success:
        return {"success": True, "ledger_id": result.get("ledger_id")}
    else:
        return {"success": False, "message": result}


@app.post("/api/ledger/unlink")
async def api_ledger_unlink(request: Request, current_user: dict = Depends(login_required)):
    """解除流水与订单的关联"""
    data = await request.json()
    ledger_id = data.get("ledger_id")
    if not ledger_id:
        return {"success": False, "message": "缺少台账ID"}

    success, message = delete_ledger(ledger_id)
    return {"success": success, "message": message}


@app.post("/api/ledger/update")
async def api_ledger_update(request: Request, current_user: dict = Depends(login_required)):
    """更新关联信息"""
    data = await request.json()
    ledger_id = data.get("ledger_id")
    if not ledger_id:
        return {"success": False, "message": "缺少台账ID"}

    success, message = update_ledger(
        ledger_id=ledger_id,
        allocation_amount=data.get("allocation_amount"),
        remark=data.get("remark")
    )
    return {"success": success, "message": message}


@app.post("/api/ledger/set_primary")
async def api_ledger_set_primary(request: Request, current_user: dict = Depends(login_required)):
    """设置主要匹配记录"""
    data = await request.json()
    manager_id = data.get("manager_id")
    ledger_id = data.get("ledger_id")
    if not manager_id or not ledger_id:
        return {"success": False, "message": "缺少必要参数"}

    success, message = set_primary_ledger(manager_id, ledger_id)
    return {"success": success, "message": message}


@app.get("/api/ledger/validate/{transaction_id}")
async def api_ledger_validate(transaction_id: str, amount: float, current_user: dict = Depends(login_required)):
    """验证分配金额是否有效"""
    is_valid, remaining = validate_allocation_amount(transaction_id, amount)
    return {"is_valid": is_valid, "remaining": remaining}
