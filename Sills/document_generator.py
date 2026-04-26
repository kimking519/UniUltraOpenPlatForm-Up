"""
Document Generator - 统一的文档生成模块

包含 CI (Commercial Invoice)、PI (Proforma Invoice) 等文档生成功能。
供 main.py 和 skill 脚本共同调用。
"""

import os
import platform
import zipfile
import tempfile
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
except ImportError:
    openpyxl = None

from Sills.base import get_db_connection, get_exchange_rates


# ============================================================
# 样式配置 - 商务蓝风格
# ============================================================

COLOR_HEADER_BG = "1E3A5F"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ODD = "F8FAFC"
COLOR_ROW_EVEN = "FFFFFF"
COLOR_TOTAL_BG = "1E3A5F"
COLOR_TOTAL_FONT = "FFFFFF"
COLOR_TOTAL_AMOUNT = "DC2626"

BORDER_THIN = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

BORDER_MEDIUM = Border(
    left=Side(style='medium', color='1E3A5F'),
    right=Side(style='medium', color='1E3A5F'),
    top=Side(style='medium', color='1E3A5F'),
    bottom=Side(style='medium', color='1E3A5F')
)


# ============================================================
# 数据查询函数
# ============================================================

def _get_default_output_base():
    """获取默认输出目录 - 兼容 Windows 和 WSL2"""
    if platform.system() == "Windows":
        return r"E:\1_Business\1_Auto"
    else:
        # WSL2 或其他 Linux 系统
        return "/mnt/e/1_Business/1_Auto"


def _get_output_base():
    """获取输出基础目录"""
    output_base = os.environ.get('UNIULTRA_OUTPUT_DIR')
    if output_base:
        return output_base
    return _get_default_output_base()


def _generate_unique_invoice_no(output_dir, cli_name, doc_type="PI"):
    """
    生成唯一的发票编号和文件路径

    发票编号格式: UNI%Y%m%d，如果已存在则添加01, 02, 03等后缀
    文件名格式:
      - PI: Proforma Invoice_{cli_name}_{invoice_no}.xlsx
      - CI: COMMERCIAL INVOICE_{cli_name}_{invoice_no}.xlsx

    Args:
        output_dir: 输出目录
        cli_name: 客户名称
        doc_type: 文档类型 (PI, CI_KR, CI_US)

    Returns:
        tuple: (invoice_no, output_filename, output_path)
    """
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    base_invoice_no = f"UNI{date_str}"

    # 确定文档类型前缀
    if doc_type.startswith("CI"):
        doc_prefix = "COMMERCIAL INVOICE"
    else:
        doc_prefix = "Proforma Invoice"

    # 尝试不带后缀
    invoice_no = base_invoice_no
    output_filename = f"{doc_prefix}_{cli_name}_{invoice_no}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    suffix = 1
    while os.path.exists(output_path):
        invoice_no = f"{base_invoice_no}{suffix:02d}"
        output_filename = f"{doc_prefix}_{cli_name}_{invoice_no}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        suffix += 1

    return invoice_no, output_filename, output_path


def _safe_write_cell(ws, row, col, value):
    """安全写入单元格 - 必须先取消合并再获取单元格"""
    # 关键：先取消该单元格所在的所有合并区域，再获取单元格
    # 因为 MergedCell 对象是只读的，必须在获取 cell 之前取消合并
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            ws.unmerge_cells(str(merged_range))

    # 取消合并后获取的一定是普通 Cell 对象
    cell = ws.cell(row=row, column=col)
    cell.value = value


def get_orders_for_document(order_ids):
    """获取订单列表（用于生成CI/PI文档）"""
    if not order_ids:
        return []
    placeholders = ','.join(['?'] * len(order_ids))
    with get_db_connection() as conn:
        rows = conn.execute(f"""
            SELECT o.order_id, o.order_no, o.order_date, o.cli_id,
                   o.inquiry_mpn, o.inquiry_brand, o.price_rmb, o.price_kwr, o.price_usd,
                   o.offer_id, o.cost_price_rmb,
                   c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
                   c.cli_full_name, c.region,
                   off.quoted_qty, off.date_code, off.delivery_date, off.inquiry_qty
            FROM uni_order o
            LEFT JOIN uni_cli c ON o.cli_id = c.cli_id
            LEFT JOIN uni_offer off ON o.offer_id = off.offer_id
            WHERE o.order_id IN ({placeholders})
            ORDER BY o.order_id
        """, order_ids).fetchall()
        return [dict(r) for r in rows]


def get_offers_for_document(offer_ids):
    """获取报价列表（用于生成韩文报价单）"""
    if not offer_ids:
        return []
    placeholders = ','.join(['?'] * len(offer_ids))
    with get_db_connection() as conn:
        rows = conn.execute(f"""
            SELECT o.offer_id, o.offer_date,
                   o.quoted_mpn, o.quoted_brand, o.offer_price_rmb,
                   o.price_kwr, o.price_usd as offer_price_usd, o.price_jpy,
                   o.quoted_qty, o.date_code, o.delivery_date, o.inquiry_qty, o.inquiry_mpn,
                   COALESCE(o.cli_id, q.cli_id) as cli_id,
                   c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
                   c.cli_full_name, c.region
            FROM uni_offer o
            LEFT JOIN uni_quote q ON o.quote_id = q.quote_id
            LEFT JOIN uni_cli c ON COALESCE(o.cli_id, q.cli_id) = c.cli_id
            WHERE o.offer_id IN ({placeholders})
            ORDER BY o.offer_id
        """, offer_ids).fetchall()
        return [dict(r) for r in rows]


# ============================================================
# CI 生成 - 美元版
# ============================================================

def generate_ci_us(order_ids, output_base=None, template_dir=None):
    """
    生成美元版 Commercial Invoice

    Args:
        order_ids: 订单ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not order_ids:
        return False, "未提供订单编号"

    orders = get_orders_for_document(order_ids)
    if not orders:
        return False, f"订单编号 {order_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in orders)
    if len(cli_names) > 1:
        return False, f"订单属于不同客户 ({', '.join(cli_names)})，无法生成同一份CI"

    cli_name = list(cli_names)[0]
    # 获取客户英文名，用于文件名
    cli_name_en = orders[0].get("cli_name_en") or orders[0].get("cli_name") or "Unknown"

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "ci_us")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "CI_US")

    # 生成文档
    return _generate_ci_us_excel(orders, template_dir, output_path, invoice_no)


def _generate_ci_us_excel(orders, template_dir, output_path, invoice_no):
    """生成美元版CI Excel文件 - Header + Footer 简单拼接方式"""
    import shutil

    data_count = len(orders)
    first_order = orders[0]
    now = datetime.now()

    # 只使用 Header + Footer 拼接方式
    header_path = os.path.join(template_dir, "CI_template_header_US.xlsx")
    footer_path = os.path.join(template_dir, "CI_template_footer_US.xlsx")

    if not os.path.exists(header_path):
        return False, f"Header模板不存在: {header_path}"
    if not os.path.exists(footer_path):
        return False, f"Footer模板不存在: {footer_path}"

    wb = openpyxl.load_workbook(header_path)
    ws = wb.active

    # 填写头部信息
    ws['B2'] = invoice_no
    ws['F2'] = now.strftime("%Y-%m-%d")

    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws['B9'] = cli_name_en
    ws['F9'] = first_order.get("contact_name", "") or ""
    ws['B10'] = first_order.get("region", "") or "USA"
    ws['F10'] = first_order.get("phone", "") or ""
    ws['B11'] = first_order.get("address", "") or ""

    # 动态生成数据行
    data_start_row = 13
    template_data_rows = 3

    if data_count > template_data_rows:
        extra_rows = data_count - template_data_rows
        ws.insert_rows(data_start_row + template_data_rows, extra_rows)
        for i in range(extra_rows):
            src_row = data_start_row + template_data_rows - 1
            dst_row = data_start_row + template_data_rows + i
            for col in range(1, 8):
                src_cell = ws.cell(row=src_row, column=col)
                dst_cell = ws.cell(row=dst_row, column=col)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)
            if ws.row_dimensions[src_row].height:
                ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    # 填充数据
    total_qty = 0
    total_amount = 0

    for idx, order in enumerate(orders):
        row = data_start_row + idx
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = "集成电路/IC"
        ws.cell(row=row, column=3).value = order.get("inquiry_mpn", "") or ""
        ws.cell(row=row, column=4).value = "8542399000"

        qty = order.get("quoted_qty") or order.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws.cell(row=row, column=5).value = qty
        ws.cell(row=row, column=5).number_format = '#,##0'
        total_qty += qty

        price_usd = float(order.get("price_usd") or 0)
        ws.cell(row=row, column=6).value = price_usd
        ws.cell(row=row, column=6).number_format = '#,##0.000'

        total = price_usd * qty
        ws.cell(row=row, column=7).value = total if total else 0
        ws.cell(row=row, column=7).number_format = '#,##0.000'
        total_amount += total

    # 完整拼接 Footer（复制值、样式、合并单元格、行高、图片）
    footer_start_row = data_start_row + data_count
    wb_footer = openpyxl.load_workbook(footer_path)
    ws_footer = wb_footer.active

    # 复制footer的所有行
    for src_row in range(1, ws_footer.max_row + 1):
        dst_row = footer_start_row + src_row - 1
        for col in range(1, 8):
            src_cell = ws_footer.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)
        if ws_footer.row_dimensions[src_row].height:
            ws.row_dimensions[dst_row].height = ws_footer.row_dimensions[src_row].height

    # 复制footer的合并单元格
    for merged_range in ws_footer.merged_cells.ranges:
        new_range = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{footer_start_row + merged_range.min_row - 1}:{openpyxl.utils.get_column_letter(merged_range.max_col)}{footer_start_row + merged_range.max_row - 1}"
        try:
            ws.merge_cells(new_range)
        except:
            pass

    # 复制footer中的图片（印章）
    for img in ws_footer._images:
        from copy import deepcopy
        new_img = deepcopy(img)
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            new_img.anchor._from.row = footer_start_row + img.anchor._from.row - 1
            if hasattr(img.anchor, 'to'):
                new_img.anchor.to.row = footer_start_row + img.anchor.to.row - 1
        ws.add_image(new_img)

    # 更新 TOTAL 数据
    ws.cell(row=footer_start_row, column=5).value = total_qty
    ws.cell(row=footer_start_row, column=5).number_format = '#,##0'
    ws.cell(row=footer_start_row + 1, column=7).value = total_amount
    ws.cell(row=footer_start_row + 1, column=7).number_format = '#,##0.000'

    last_row = footer_start_row + ws_footer.max_row - 1
    ws.print_area = f"$A$1:$G${last_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return True, {
        "excel_path": output_path,
        "pdf_path": "",
        "invoice_no": invoice_no,
        "total_qty": total_qty,
        "total_amount": total_amount,
        "cli_name": first_order.get("cli_name", ""),
        "count": data_count
    }


# ============================================================
# PI 生成
# ============================================================

def generate_pi(order_ids, output_base=None, template_dir=None):
    """
    生成 Proforma Invoice (KRW版本) - 双模板拼接方式

    Args:
        order_ids: 订单ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not order_ids:
        return False, "未提供订单编号"

    orders = get_orders_for_document(order_ids)
    if not orders:
        return False, f"订单编号 {order_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in orders)
    if len(cli_names) > 1:
        return False, f"订单属于不同客户 ({', '.join(cli_names)})，无法生成同一份PI"

    cli_name = list(cli_names)[0]
    # 获取客户英文名，用于文件名
    cli_name_en = orders[0].get("cli_name_en") or orders[0].get("cli_name") or "Unknown"

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "pi")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "PI")

    # 获取汇率
    krw_val, _, _ = get_exchange_rates()

    # 计算KWR价格
    for order in orders:
        price_kwr = order.get("price_kwr")
        if not price_kwr or float(price_kwr or 0) == 0:
            price_rmb = order.get("price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                price_kwr = round(float(price_rmb) * krw_val, 1)
            else:
                price_kwr = 0
        else:
            price_kwr = float(price_kwr)
        order["calculated_price_kwr"] = price_kwr

    return _generate_pi_kr_excel(orders, template_dir, output_path, invoice_no)


def _generate_pi_kr_excel(orders, template_dir, output_path, invoice_no):
    """
    生成PI-KR Excel文件 - 双模板拼接方式

    模板1 (KR-1): 头部信息 + 数据表
    模板2 (KR-2): TOTAL AMOUNT + TERMS & CONDITIONS + 付款信息
    """
    from openpyxl.utils import get_column_letter
    from copy import copy

    data_count = len(orders)
    first_order = orders[0]
    now = datetime.now()

    # 模板文件路径
    template1_path = os.path.join(template_dir, "Proforma_Invoice_KR - 1.xlsx")
    template2_path = os.path.join(template_dir, "Proforma_Invoice_KR - 2.xlsx")

    # 检查新模板是否存在
    use_new_template = os.path.exists(template1_path) and os.path.exists(template2_path)

    if not use_new_template:
        # 回退到旧模板
        return _generate_pi_excel_legacy(orders, template_dir, output_path, invoice_no)

    # 加载模板
    wb1 = openpyxl.load_workbook(template1_path)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(template2_path)
    ws2 = wb2.active

    # ---- 1. 填写头部信息 ----
    # Row 8: Invoice No. (D8)
    ws1.cell(8, 4).value = invoice_no
    # Row 9: Date (D9)
    ws1.cell(9, 4).value = now.strftime("%Y-%m-%d")

    # Row 12-16: 客户信息
    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws1.cell(12, 3).value = cli_name_en
    ws1.cell(13, 3).value = first_order.get("contact_name", "") or ""
    ws1.cell(14, 3).value = first_order.get("email", "") or ""
    ws1.cell(15, 3).value = first_order.get("phone", "") or ""
    ws1.cell(16, 3).value = first_order.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    # KR-1 模板: Row 18=表头, Row 19-20=示例数据
    header_row = 18
    first_data_row = 19
    template_data_rows = 2  # 模板中有2行示例数据

    # 调整行数
    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        # 插入新行
        ws1.insert_rows(first_data_row + template_data_rows, rows_diff)
        # 复制样式
        style_row = first_data_row  # 使用第一行作为样式模板
        for i in range(rows_diff):
            new_row = first_data_row + template_data_rows + i
            for col in range(1, 9):
                src_cell = ws1.cell(row=style_row, column=col)
                dest_cell = ws1.cell(row=new_row, column=col)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.alignment = copy(src_cell.alignment)
    elif rows_diff < 0:
        # 删除多余行
        ws1.delete_rows(first_data_row + data_count, -rows_diff)

    # 取消数据行的合并单元格
    merged_to_remove = []
    for merged_range in list(ws1.merged_cells.ranges):
        if first_data_row <= merged_range.min_row < first_data_row + data_count:
            merged_to_remove.append(merged_range)
    for merged_range in merged_to_remove:
        try:
            ws1.unmerge_cells(str(merged_range))
        except:
            pass

    # 写入数据
    for idx, order in enumerate(orders):
        row = first_data_row + idx

        ws1.cell(row, 1).value = idx + 1
        ws1.cell(row, 2).value = order.get("inquiry_mpn", "") or ""
        ws1.cell(row, 3).value = order.get("inquiry_brand", "") or ""
        ws1.cell(row, 4).value = order.get("date_code", "") or ""

        # Qty: 优先使用 quoted_qty，无值时写 0
        qty = order.get("quoted_qty") or order.get("inquiry_qty") or 0
        try:
            qty = int(qty) if qty else 0
        except:
            qty = 0
        ws1.cell(row, 5).value = qty
        ws1.cell(row, 5).number_format = '#,##0'

        ws1.cell(row, 6).value = order.get("delivery_date", "") or ""

        # Unit Price: 无值时写 0，始终设置 number_format
        price_kwr = order.get("calculated_price_kwr") or 0
        try:
            price_kwr = float(price_kwr) if price_kwr else 0
        except:
            price_kwr = 0
        ws1.cell(row, 7).value = price_kwr
        ws1.cell(row, 7).number_format = '#,##0.0'

        # Total Amount: 始终写入公式
        ws1.cell(row, 8).value = f"=G{row}*E{row}"
        ws1.cell(row, 8).number_format = '#,##0.0'

    # ---- 3. 计算 Total 行位置 ----
    last_data_row = first_data_row + data_count - 1

    # ---- 4. 完整拼接 template2 内容（复制值、样式、合并单元格、行高、图片）----
    template2_start_row = 11
    insert_start_row = last_data_row + 2

    for src_row_idx in range(template2_start_row, ws2.max_row + 1):
        dest_row_idx = insert_start_row + (src_row_idx - template2_start_row)
        ws1.insert_rows(dest_row_idx)

        for src_col in range(1, 9):
            src_cell = ws2.cell(row=src_row_idx, column=src_col)
            dest_cell = ws1.cell(row=dest_row_idx, column=src_col)
            if src_cell.value is not None:
                dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = copy(src_cell.alignment)
        if ws2.row_dimensions[src_row_idx].height:
            ws1.row_dimensions[dest_row_idx].height = ws2.row_dimensions[src_row_idx].height

    # 复制template2的合并单元格
    for merged_range in ws2.merged_cells.ranges:
        if merged_range.min_row >= template2_start_row:
            new_min_row = insert_start_row + (merged_range.min_row - template2_start_row)
            new_max_row = insert_start_row + (merged_range.max_row - template2_start_row)
            try:
                ws1.merge_cells(
                    start_row=new_min_row,
                    start_column=merged_range.min_col,
                    end_row=new_max_row,
                    end_column=merged_range.max_col
                )
            except:
                pass

    # 复制template2中的图片（印章）
    from copy import deepcopy
    row_offset = insert_start_row - template2_start_row
    for img in ws2._images:
        new_img = deepcopy(img)
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            original_from_row = img.anchor._from.row
            if original_from_row >= template2_start_row - 1:
                new_img.anchor._from.row = original_from_row + row_offset
            if hasattr(img.anchor, 'to'):
                original_to_row = img.anchor.to.row
                if original_to_row >= template2_start_row - 1:
                    new_img.anchor.to.row = original_to_row + row_offset
        ws1.add_image(new_img)

    # 更新 TOTAL AMOUNT 行的公式
    total_row = insert_start_row
    ws1.cell(total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws1.cell(total_row, 8).number_format = '#,##0.0'

    # 设置打印范围
    thank_you_row = insert_start_row + (ws2.max_row - template2_start_row)
    ws1.print_area = f"$A$1:$H${thank_you_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb1.save(output_path)
    wb1.close()
    wb2.close()

    return True, {
        "excel_path": output_path,
        "invoice_no": now.strftime("UNI%Y%m%d%H"),
        "cli_name": first_order.get("cli_name", ""),
        "count": data_count
    }


def _generate_pi_excel_legacy(orders, template_dir, output_path, invoice_no):
    """生成PI Excel文件"""
    data_count = len(orders)
    first_order = orders[0]
    now = datetime.now()

    # 查找模板文件
    template_path = None
    if os.path.isdir(template_dir):
        # 优先使用正确的模板（头部区域D8/D9是普通Cell，不是MergedCell）
        preferred_template = "Proforma_Invoice_KR.xlsx"
        preferred_path = os.path.join(template_dir, preferred_template)
        if os.path.exists(preferred_path):
            template_path = preferred_path
        else:
            for f in os.listdir(template_dir):
                if f.endswith(".xlsx") and not f.startswith("~"):
                    template_path = os.path.join(template_dir, f)
                    break

    if not template_path or not os.path.exists(template_path):
        return False, f"模板文件不存在于 {template_dir}"

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ---- 1. 填写头部信息 ----
    ws.cell(8, 4).value = invoice_no
    ws.cell(9, 4).value = now.strftime("%Y-%m-%d")

    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws.cell(12, 3).value = cli_name_en
    ws.cell(13, 3).value = first_order.get("contact_name", "") or ""
    ws.cell(14, 3).value = first_order.get("email", "") or ""
    ws.cell(15, 3).value = first_order.get("phone", "") or ""
    ws.cell(16, 3).value = first_order.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    header_row = 18
    first_data_row = 19
    template_data_rows = 2
    total_template_row = 21
    footer_start_row = 22

    # ============================================================
    # 关键修复：在插入行之前保存固定内容区域的所有信息
    # ============================================================

    # 保存固定内容区域的合并单元格
    footer_merged_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= footer_start_row:
            footer_merged_ranges.append({
                "range_str": str(merged_range),
                "min_row": merged_range.min_row,
                "min_col": merged_range.min_col,
                "max_row": merged_range.max_row,
                "max_col": merged_range.max_col,
            })

    # 保存固定内容区域的所有单元格值和样式
    footer_cells_data = {}
    for row in range(footer_start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is not None or cell.font or cell.border or cell.fill:
                footer_cells_data[(row, col)] = {
                    "value": cell.value,
                    "font": copy(cell.font) if cell.font else None,
                    "border": copy(cell.border) if cell.border else None,
                    "fill": copy(cell.fill) if cell.fill else None,
                    "alignment": copy(cell.alignment) if cell.alignment else None,
                    "number_format": cell.number_format,
                }

    footer_row_heights = {}
    for row in range(footer_start_row, ws.max_row + 1):
        if ws.row_dimensions[row].height:
            footer_row_heights[row] = ws.row_dimensions[row].height

    # 调整行数
    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        ws.insert_rows(total_template_row, rows_diff)
        for i in range(rows_diff):
            new_row = total_template_row + i
            ws.row_dimensions[new_row].height = 20.0
    elif rows_diff < 0:
        ws.delete_rows(first_data_row + data_count, -rows_diff)

    actual_total_row = first_data_row + data_count
    actual_footer_start_row = actual_total_row + 1

    # ============================================================
    # 关键修复：恢复固定内容区域的数据
    # ============================================================

    # 先取消所有可能存在的合并单元格
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= actual_footer_start_row:
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass

    # 恢复单元格值和样式（按新的行号偏移）
    for (orig_row, col), cell_data in footer_cells_data.items():
        new_row = orig_row + rows_diff
        if new_row >= actual_footer_start_row:
            cell = ws.cell(new_row, col)
            if cell_data["value"] is not None:
                cell.value = cell_data["value"]
            if cell_data["font"]:
                cell.font = cell_data["font"]
            if cell_data["border"]:
                cell.border = cell_data["border"]
            if cell_data["fill"]:
                cell.fill = cell_data["fill"]
            if cell_data["alignment"]:
                cell.alignment = cell_data["alignment"]
            if cell_data["number_format"]:
                cell.number_format = cell_data["number_format"]

    for orig_row, height in footer_row_heights.items():
        new_row = orig_row + rows_diff
        ws.row_dimensions[new_row].height = height

    # 恢复合并单元格
    for merged_info in footer_merged_ranges:
        new_min_row = merged_info["min_row"] + rows_diff
        new_max_row = merged_info["max_row"] + rows_diff
        if new_min_row <= actual_total_row - 1 and new_max_row >= first_data_row:
            continue
        try:
            ws.merge_cells(
                start_row=new_min_row,
                start_column=merged_info["min_col"],
                end_row=new_max_row,
                end_column=merged_info["max_col"]
            )
        except:
            pass

    # 取消数据行的合并单元格（关键：在写入数据前）
    merged_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        min_row = merged_range.min_row
        if first_data_row <= min_row <= actual_total_row - 1:
            merged_to_remove.append(merged_range)
    for merged_range in merged_to_remove:
        try:
            ws.unmerge_cells(str(merged_range))
        except:
            pass

    # 写入数据 (legacy版本)
    for idx, order in enumerate(orders):
        row = first_data_row + idx

        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = order.get("inquiry_mpn", "") or ""
        ws.cell(row, 3).value = order.get("inquiry_brand", "") or ""
        ws.cell(row, 4).value = order.get("date_code", "") or ""

        # Qty: 无值时写 0，始终设置 number_format
        qty = order.get("quoted_qty") or order.get("inquiry_qty") or 0
        try:
            qty = int(qty) if qty else 0
        except:
            qty = 0
        ws.cell(row, 5).value = qty
        ws.cell(row, 5).number_format = '#,##0'

        ws.cell(row, 6).value = order.get("delivery_date", "") or ""

        # Unit Price: 无值时写 0，始终设置 number_format
        price_kwr = order.get("calculated_price_kwr") or 0
        try:
            price_kwr = float(price_kwr) if price_kwr else 0
        except:
            price_kwr = 0
        ws.cell(row, 7).value = price_kwr
        ws.cell(row, 7).number_format = '#,##0.0'

        # Total Amount: 始终写入公式
        ws.cell(row, 8).value = f"=G{row}*E{row}"
        ws.cell(row, 8).number_format = '#,##0.0'

    # 更新 TOTAL 行
    last_data_row = actual_total_row - 1

    # 只更新公式，保留模板原有的文字
    ws.cell(actual_total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws.cell(actual_total_row, 8).number_format = '#,##0.0'

    ws.print_area = f"$A$1:$H${ws.max_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    # 生成 PDF
    pdf_path = ""
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "excel_to_pdf",
            os.path.join(os.path.dirname(os.path.dirname(__file__)),
                        "openclaw_skills", "order-pi-generator", "scripts", "excel_to_pdf.py")
        )
        if spec and spec.loader:
            pdf_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(pdf_module)
            pdf_success, pdf_result = pdf_module.convert_to_pdf(output_path)
            if pdf_success:
                pdf_path = pdf_result
    except Exception as e:
        pass  # PDF 生成失败不影响 Excel

    return True, {
        "excel_path": output_path,
        "pdf_path": pdf_path,
        "invoice_no": now.strftime("UNI%Y%m%d%H"),
        "cli_name": first_order.get("cli_name", ""),
        "count": data_count
    }


# ============================================================
# PI-US 生成 (USD 价格)
# ============================================================

def generate_pi_us(order_ids, output_base=None, template_dir=None):
    """
    生成 Proforma Invoice (USD版本)

    Args:
        order_ids: 订单ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not order_ids:
        return False, "未提供订单编号"

    orders = get_orders_for_document(order_ids)
    if not orders:
        return False, f"订单编号 {order_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in orders)
    if len(cli_names) > 1:
        return False, f"订单属于不同客户 ({', '.join(cli_names)})，无法生成同一份PI"

    cli_name = list(cli_names)[0]
    # 获取客户英文名，用于文件名
    cli_name_en = orders[0].get("cli_name_en") or orders[0].get("cli_name") or "Unknown"

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "pi")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "PI")

    # 获取汇率 (USD汇率用于计算)
    _, usd_val, _ = get_exchange_rates()

    # 计算USD价格
    for order in orders:
        price_usd = order.get("price_usd")
        if not price_usd or float(price_usd or 0) == 0:
            price_rmb = order.get("price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                # USD汇率表示 1 RMB = ? USD，直接乘
                price_usd = round(float(price_rmb) * usd_val, 3) if usd_val else 0
            else:
                price_usd = 0
        else:
            price_usd = float(price_usd)
        order["calculated_price_usd"] = price_usd

    return _generate_pi_us_excel(orders, template_dir, output_path, invoice_no)


def _generate_pi_us_excel(orders, template_dir, output_path, invoice_no):
    """
    生成PI-US Excel文件 - 双模板拼接方式

    模板1 (US-1): 头部信息 + 数据表
    模板2 (US-2): TOTAL AMOUNT + TERMS & CONDITIONS + 付款信息
    """
    from openpyxl.utils import get_column_letter
    from copy import copy, deepcopy

    data_count = len(orders)
    first_order = orders[0]
    now = datetime.now()

    # 模板文件路径
    template1_path = os.path.join(template_dir, "Proforma_Invoice_US - 1.xlsx")
    template2_path = os.path.join(template_dir, "Proforma_Invoice_US - 2.xlsx")

    if not os.path.exists(template1_path):
        return False, f"US Header模板不存在: {template1_path}"
    if not os.path.exists(template2_path):
        return False, f"US Footer模板不存在: {template2_path}"

    # 加载模板
    wb1 = openpyxl.load_workbook(template1_path)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(template2_path)
    ws2 = wb2.active

    # ---- 1. 填写头部信息 ----
    ws1.cell(8, 4).value = invoice_no
    ws1.cell(9, 4).value = now.strftime("%Y-%m-%d")

    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws1.cell(12, 3).value = cli_name_en
    ws1.cell(13, 3).value = first_order.get("contact_name", "") or ""
    ws1.cell(14, 3).value = first_order.get("email", "") or ""
    ws1.cell(15, 3).value = first_order.get("phone", "") or ""
    ws1.cell(16, 3).value = first_order.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    header_row = 18
    first_data_row = 19
    template_data_rows = 1

    rows_diff = data_count - template_data_rows

    if rows_diff > 0:
        ws1.insert_rows(first_data_row + template_data_rows, rows_diff)
        style_row = first_data_row
        for i in range(rows_diff):
            new_row = first_data_row + template_data_rows + i
            for col in range(1, 9):
                src_cell = ws1.cell(row=style_row, column=col)
                dest_cell = ws1.cell(row=new_row, column=col)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.alignment = copy(src_cell.alignment)
    elif rows_diff < 0:
        ws1.delete_rows(first_data_row + data_count, -rows_diff)

    # 取消数据行的合并单元格
    merged_to_remove = []
    for merged_range in list(ws1.merged_cells.ranges):
        if first_data_row <= merged_range.min_row < first_data_row + data_count:
            merged_to_remove.append(merged_range)
    for merged_range in merged_to_remove:
        try:
            ws1.unmerge_cells(str(merged_range))
        except:
            pass

    # 写入数据 (US版本)
    for idx, order in enumerate(orders):
        row = first_data_row + idx

        ws1.cell(row, 1).value = idx + 1
        ws1.cell(row, 2).value = order.get("inquiry_mpn", "") or ""
        ws1.cell(row, 3).value = order.get("inquiry_brand", "") or ""
        ws1.cell(row, 4).value = order.get("date_code", "") or ""

        # Qty: 无值时写 0，始终设置 number_format
        qty = order.get("quoted_qty") or order.get("inquiry_qty") or 0
        try:
            qty = int(qty) if qty else 0
        except:
            qty = 0
        ws1.cell(row, 5).value = qty
        ws1.cell(row, 5).number_format = '#,##0'

        ws1.cell(row, 6).value = order.get("delivery_date", "") or ""

        # Unit Price USD: 无值时写 0，始终设置 number_format
        price_usd = order.get("calculated_price_usd") or 0
        try:
            price_usd = float(price_usd) if price_usd else 0
        except:
            price_usd = 0
        ws1.cell(row, 7).value = price_usd
        ws1.cell(row, 7).number_format = '#,##0.000'

        # Total Amount: 始终写入公式
        ws1.cell(row, 8).value = f"=G{row}*E{row}"
        ws1.cell(row, 8).number_format = '#,##0.000'

    # ---- 3. 计算 Total 行位置 ----
    last_data_row = first_data_row + data_count - 1

    # ---- 4. 拼接 template2 内容（从 Row 1 开始）----
    insert_start_row = last_data_row + 2

    for src_row_idx in range(1, ws2.max_row + 1):
        dest_row_idx = insert_start_row + src_row_idx - 1

        if src_row_idx > 1:
            ws1.insert_rows(dest_row_idx)

        for src_col in range(1, 9):
            src_cell = ws2.cell(row=src_row_idx, column=src_col)
            dest_cell = ws1.cell(row=dest_row_idx, column=src_col)

            if src_cell.value is not None:
                dest_cell.value = src_cell.value

            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = copy(src_cell.alignment)

    # 处理 template2 的合并单元格
    for merged_range in ws2.merged_cells.ranges:
        new_min_row = merged_range.min_row + insert_start_row - 1
        new_max_row = merged_range.max_row + insert_start_row - 1
        try:
            ws1.merge_cells(
                start_row=new_min_row,
                start_column=merged_range.min_col,
                end_row=new_max_row,
                end_column=merged_range.max_col
            )
        except:
            pass

    # ---- 5. 更新 Total Amount 行的公式 ----
    total_row = insert_start_row

    # 取消 Total 行的合并单元格
    for merged_range in list(ws1.merged_cells.ranges):
        if merged_range.min_row == total_row:
            try:
                ws1.unmerge_cells(str(merged_range))
            except:
                pass

    ws1.cell(total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws1.cell(total_row, 8).number_format = '#,##0.000'

    # 重新合并 Total 行
    try:
        ws1.merge_cells(f"A{total_row}:G{total_row}")
    except:
        pass

    # ---- 5.5 复制 template2 的图片 ----
    row_offset = insert_start_row - 1
    for img in ws2._images:
        new_img = deepcopy(img)
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            original_from_row = img.anchor._from.row
            new_img.anchor._from.row = original_from_row + row_offset
            if hasattr(img.anchor, 'to'):
                original_to_row = img.anchor.to.row
                new_img.anchor.to.row = original_to_row + row_offset
        ws1.add_image(new_img)

    # ---- 6. 保存文件 ----
    thank_you_row = insert_start_row + 18
    ws1.print_area = f"$A$1:$H${thank_you_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb1.save(output_path)
    wb1.close()
    wb2.close()

    return True, {
        "excel_path": output_path,
        "invoice_no": now.strftime("UNI%Y%m%d%H"),
        "cli_name": first_order.get("cli_name", ""),
        "count": data_count
    }


def _generate_pi_us_excel_legacy(orders, template_path, output_path, invoice_no):
    """旧版PI-US生成逻辑 - 兼容旧版本模板"""
    data_count = len(orders)
    first_order = orders[0]
    now = datetime.now()

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 填写头部信息
    ws.cell(8, 4).value = invoice_no
    ws.cell(9, 4).value = now.strftime("%Y-%m-%d")

    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws.cell(12, 3).value = cli_name_en
    ws.cell(13, 3).value = first_order.get("contact_name", "") or ""
    ws.cell(14, 3).value = first_order.get("email", "") or ""
    ws.cell(15, 3).value = first_order.get("phone", "") or ""
    ws.cell(16, 3).value = first_order.get("address", "") or ""

    # 处理数据行
    header_row = 18
    first_data_row = 19
    template_data_rows = 1

    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        ws.insert_rows(20, rows_diff)
        for i in range(rows_diff):
            new_row = 20 + i
            ws.row_dimensions[new_row].height = 20.0

    actual_total_row = first_data_row + data_count

    # 取消数据行的合并单元格
    merged_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        min_row = merged_range.min_row
        if first_data_row <= min_row <= actual_total_row - 1:
            merged_to_remove.append(merged_range)
    for merged_range in merged_to_remove:
        try:
            ws.unmerge_cells(str(merged_range))
        except:
            pass

    # 写入数据
    for idx, order in enumerate(orders):
        row = first_data_row + idx

        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = order.get("inquiry_mpn", "") or ""
        ws.cell(row, 3).value = order.get("inquiry_brand", "") or ""
        ws.cell(row, 4).value = order.get("date_code", "") or ""

        qty = order.get("quoted_qty") or order.get("inquiry_qty") or ""
        ws.cell(row, 5).value = qty
        if qty:
            ws.cell(row, 5).number_format = '#,##0'

        ws.cell(row, 6).value = order.get("delivery_date", "") or ""

        price_usd = order.get("calculated_price_usd", "") or ""
        ws.cell(row, 7).value = price_usd
        if price_usd:
            ws.cell(row, 7).number_format = '#,##0.000'

        if qty and price_usd:
            ws.cell(row, 8).value = f"=G{row}*E{row}"
            ws.cell(row, 8).number_format = '#,##0.000'

    # 更新 TOTAL 行
    last_data_row = actual_total_row - 1
    ws.cell(actual_total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws.cell(actual_total_row, 8).number_format = '#,##0.000'

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()

    return True, {
        "excel_path": output_path,
        "invoice_no": now.strftime("UNI%Y%m%d%H"),
        "cli_name": first_order.get("cli_name", ""),
        "count": data_count
    }


# ============================================================
# 韩文报价单生成 (견적서)
# ============================================================

def generate_koquote(offer_ids, output_base=None, template_dir=None):
    """
    生成韩文报价单 (견적서)

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not offer_ids:
        return False, "未提供报价编号"

    offers = get_offers_for_document(offer_ids)
    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份报价单"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "koquote")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")
    time_str = now.strftime("%Y%m%d%H%M")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    output_filename = f"유니콘_전자부품견적서_{time_str}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # 获取汇率
    krw_val, _, _ = get_exchange_rates()

    return _generate_koquote_excel(offers, template_dir, output_path, krw_val)


def _generate_koquote_excel(offers, template_dir, output_path, exchange_rate_krw):
    """
    生成韩文报价单 Excel 文件 - 使用双模板方案

    模板1 (template1.xlsx): 头部信息 + 动态数据表
    模板2 (template2.xlsx): 固定报价条款 + 特别事项 (直接拼接)
    """
    from openpyxl.utils import get_column_letter
    from copy import copy

    data_count = len(offers)
    first_offer = offers[0]
    now = datetime.now()

    # ---- 1. 加载模板文件 ----
    template1_path = os.path.join(template_dir, "유니콘_전자부품견적서_template1.xlsx")
    template2_path = os.path.join(template_dir, "유니콘_전자부품견적서_template2.xlsx")

    # 检查新模板是否存在
    use_new_template = os.path.exists(template1_path) and os.path.exists(template2_path)

    if not use_new_template:
        # 回退到旧模板
        template_path = None
        if os.path.isdir(template_dir):
            for f in os.listdir(template_dir):
                if f.endswith(".xlsx") and not f.startswith("~") and "template1" not in f and "template2" not in f:
                    template_path = os.path.join(template_dir, f)
                    break
        if not template_path or not os.path.exists(template_path):
            return False, f"模板文件不存在于 {template_dir}"

        # 使用旧模板逻辑
        return _generate_koquote_excel_legacy(offers, template_path, output_path, exchange_rate_krw)

    # ---- 2. 使用新双模板方案 ----
    wb1 = openpyxl.load_workbook(template1_path)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(template2_path)
    ws2 = wb2.active

    # ---- 3. 填写 template1 头部信息 ----
    # 根据新模板结构:
    # Row 3: 수신(客户名) + 공급자
    # Row 4: 견적번호 + 사이트
    # Row 5: 작성일자 + 담당자
    # Row 6: 비고 + 연락처

    cli_full_name = first_offer.get("cli_full_name", "") or first_offer.get("cli_name", "")
    quote_no = now.strftime("%Y%m%d%H%M")

    # Row 3: 수신 (客户名在 C3)
    ws1.cell(3, 3).value = cli_full_name

    # Row 4: 견적번호 (在 C4)
    ws1.cell(4, 3).value = f"제 {quote_no}호"

    # Row 5: 작성일자 (在 C5)
    ws1.cell(5, 3).value = f"{now.year}년 {now.month:02d}월 {now.day:02d}일"

    # ---- 4. 处理数据行 ----
    # template1 数据从 Row 11 开始 (Row 10 是表头)
    # 模板中有5行示例数据 (Row 11-15)
    first_data_row = 11
    template_data_rows = 5  # template1 中有5行示例数据

    # 删除多余的示例行
    if data_count < template_data_rows:
        rows_to_delete = template_data_rows - data_count
        ws1.delete_rows(first_data_row + data_count, rows_to_delete)

    # 如果需要更多行，插入新行
    if data_count > template_data_rows:
        rows_to_insert = data_count - template_data_rows
        insert_at = first_data_row + template_data_rows

        # 使用最后一行作为样式模板
        style_template_row = first_data_row + template_data_rows - 1  # Row 15

        # 插入空白行
        ws1.insert_rows(insert_at, rows_to_insert)

        # 复制样式到新插入的行
        for i in range(rows_to_insert):
            new_row = insert_at + i
            # 设置行高为19.5
            ws1.row_dimensions[new_row].height = 19.5
            for col in range(2, 10):  # B-I列
                src_cell = ws1.cell(row=style_template_row, column=col)
                dest_cell = ws1.cell(row=new_row, column=col)

                # 复制样式
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = src_cell.number_format
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

    # ---- 5. 写入数据 ----
    total_qty = 0
    total_amount = 0

    for idx, offer in enumerate(offers):
        row = first_data_row + idx

        # 列结构 (新模板):
        # B列: No.
        # C列: 모델명 (型号)
        # D列: 메이커 (品牌)
        # E列: 생산일자 (date code)
        # F列: 수량 (数量)
        # G列: 단가 (单价 KRW)
        # H列: 납기 (交期)
        # I列: 비고 (备注)

        ws1.cell(row, 2).value = str(idx + 1)                       # B: No.

        # 型号: 优先使用 quoted_mpn，否则用 inquiry_mpn
        mpn = offer.get("quoted_mpn") or offer.get("inquiry_mpn", "")
        ws1.cell(row, 3).value = mpn                                 # C: 모델명

        # 品牌
        brand = offer.get("quoted_brand") or offer.get("inquiry_brand", "")
        ws1.cell(row, 4).value = brand                               # D: 메이커

        # Date code
        ws1.cell(row, 5).value = offer.get("date_code", "")          # E: 생산일자

        # 数量
        qty = offer.get("quoted_qty") or offer.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws1.cell(row, 6).value = qty                                 # F: 수량
        total_qty += qty

        # 单价: offer_price_rmb * 汇率
        price_rmb = offer.get("offer_price_rmb")
        if price_rmb and float(price_rmb or 0) > 0:
            price_kwr = float(price_rmb) * exchange_rate_krw
        else:
            price_kwr = 0
        price_cell = ws1.cell(row, 7)
        price_cell.value = price_kwr                           # G: 단가

        # 交期
        ws1.cell(row, 8).value = offer.get("delivery_date", "")      # H: 납기

        # 备注
        ws1.cell(row, 9).value = offer.get("remark", "")             # I: 비고

        total_amount += float(price_kwr or 0) * qty

    # ---- 6. 合并 template2 内容到 template1 ----
    # 计算 template1 数据结束后的空行
    last_data_row = first_data_row + data_count - 1
    insert_start_row = last_data_row + 2  # 留一行空行

    # 复制 template2 的所有行到 template1
    for src_row_idx, src_row in enumerate(ws2.iter_rows(min_row=3), start=insert_start_row):
        # 在目标位置插入新行 (第一行除外)
        if src_row_idx > insert_start_row:
            ws1.insert_rows(src_row_idx)

        for src_cell in src_row:
            if src_cell.value is not None:
                dest_cell = ws1.cell(row=src_row_idx, column=src_cell.column)
                dest_cell.value = src_cell.value

                # 复制样式
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = src_cell.number_format
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

    # ---- 设置边框样式 ----
    # 所有数据行的每个单元格都有完整的四边实线边框
    from openpyxl.styles import Border, Side
    full_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx in range(data_count):
        row = first_data_row + idx
        for col in range(2, 10):  # B-I列
            cell = ws1.cell(row, col)
            cell.border = full_border

    # 合并单元格处理
    for merged_range in ws2.merged_cells.ranges:
        # 计算目标范围
        new_min_row = merged_range.min_row - 3 + insert_start_row
        new_max_row = merged_range.max_row - 3 + insert_start_row
        new_range = f"{get_column_letter(merged_range.min_col)}{new_min_row}:{get_column_letter(merged_range.max_col)}{new_max_row}"
        try:
            ws1.merge_cells(new_range)
        except:
            pass  # 忽略合并错误

    # ---- 7. 保存文件 ----
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb1.save(output_path)
    wb1.close()
    wb2.close()

    return True, {
        "excel_path": output_path,
        "quote_no": quote_no,
        "cli_name": first_offer.get("cli_name", ""),
        "count": data_count,
        "total_qty": total_qty,
        "total_amount": total_amount
    }


def _generate_koquote_excel_legacy(offers, template_path, output_path, exchange_rate_krw):
    """旧模板生成逻辑 - 兼容旧版本模板"""
    data_count = len(offers)
    first_offer = offers[0]
    now = datetime.now()

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ---- 1. 填写头部信息 ----
    cli_full_name = first_offer.get("cli_full_name", "") or first_offer.get("cli_name", "")
    ws.cell(5, 2).value = "수신:"
    ws.cell(5, 3).value = cli_full_name

    quote_no = now.strftime("%Y%m%d%H%M")
    ws.cell(7, 2).value = f"견적번호 : 제 {quote_no}호"
    ws.cell(9, 2).value = f"작성일자 :{now.year}년  {now.month:02d}월 {now.day:02d}일"

    # ---- 2. 数据行处理 ----
    first_data_row = 17
    template_data_rows = 3

    last_data_row = first_data_row + data_count - 1

    if data_count < template_data_rows:
        ws.delete_rows(last_data_row + 1, template_data_rows - data_count)

    # ---- 3. 写入数据 ----
    total_qty = 0
    total_amount = 0

    for idx, offer in enumerate(offers):
        row = first_data_row + idx

        ws.cell(row, 2).value = str(idx + 1)
        ws.cell(row, 3).value = offer.get("inquiry_mpn", "")
        ws.cell(row, 4).value = offer.get("quoted_mpn", "")
        ws.cell(row, 5).value = offer.get("inquiry_brand", "")
        ws.cell(row, 6).value = offer.get("date_code", "")

        qty = offer.get("quoted_qty") or offer.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws.cell(row, 7).value = qty
        total_qty += qty

        price_kwr = offer.get("offer_price_rmb")
        if price_kwr and float(price_kwr or 0) > 0:
            price_kwr = round(float(price_kwr) * exchange_rate_krw, 1)
        else:
            price_kwr = 0
        ws.cell(row, 9).value = price_kwr

        ws.cell(row, 10).value = offer.get("delivery_date", "")
        ws.cell(row, 12).value = offer.get("remark", "")

        total_amount += float(price_kwr or 0) * qty

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return True, {
        "excel_path": output_path,
        "quote_no": quote_no,
        "cli_name": first_offer.get("cli_name", ""),
        "count": data_count,
        "total_qty": total_qty,
        "total_amount": total_amount
    }


# ============================================================
# 基于报价的 PI/CI 生成函数
# ============================================================

def generate_pi_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    基于报价生成 Proforma Invoice (KRW版本)

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not offer_ids:
        return False, "未提供报价编号"

    offers = get_offers_for_document(offer_ids)
    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份PI"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "pi")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "PI")

    # 获取汇率
    krw_val, _, _ = get_exchange_rates()

    # 计算KWR价格 - 优先使用已有的price_kwr，否则用offer_price_rmb换算
    for offer in offers:
        price_kwr = float(offer.get("price_kwr") or 0)
        if price_kwr > 0:
            offer["calculated_price_kwr"] = price_kwr
        else:
            price_rmb = offer.get("offer_price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                offer["calculated_price_kwr"] = round(float(price_rmb) * krw_val, 1)
            else:
                offer["calculated_price_kwr"] = 0

    # 复用订单PI生成逻辑，适配报价字段
    return _generate_pi_kr_excel_from_offers(offers, template_dir, output_path, invoice_no)


def _generate_pi_kr_excel_from_offers(offers, template_dir, output_path, invoice_no):
    """基于报价生成PI-KR Excel文件 - Header + Footer 简单拼接方式"""
    # 直接使用 header+footer 拼接方式
    return _generate_pi_excel_legacy_from_offers(offers, template_dir, output_path, invoice_no)


def _generate_pi_excel_legacy_from_offers(offers, template_dir, output_path, invoice_no):
    """基于报价生成PI Excel文件 - Header + Footer模板拼接方式

    模板结构 (KR版本):
    Header模板 (KR - 1):
    - Row 8: Invoice No. (C1=标签, D4=值)
    - Row 9: Date (C1=标签, D4=值)
    - Row 12-16: 客户信息 (C1=标签, C3=值)
    - Row 18: 表头行
    - Row 19-20: 2行示例数据

    Footer模板 (KR - 2):
    - Row 11: TOTAL AMOUNT + 公式
    - Row 13-18: TERMS & CONDITIONS
    - Row 20-27: PAYMENT METHOD
    - Row 29: THANK YOU
    """
    from openpyxl.utils import get_column_letter
    from copy import copy

    data_count = len(offers)
    first_offer = offers[0]
    now = datetime.now()

    # 加载 Header 和 Footer 模板
    header_path = os.path.join(template_dir, "Proforma_Invoice_KR - 1.xlsx")
    footer_path = os.path.join(template_dir, "Proforma_Invoice_KR - 2.xlsx")

    if not os.path.exists(header_path):
        return False, f"Header模板不存在: {header_path}"
    if not os.path.exists(footer_path):
        return False, f"Footer模板不存在: {footer_path}"

    wb = openpyxl.load_workbook(header_path)
    ws = wb.active

    wb_footer = openpyxl.load_workbook(footer_path)
    ws_footer = wb_footer.active

    # ---- 1. 写入头部信息 ----
    # Invoice No.: D8 (Row 8, Col 4)
    ws.cell(8, 4).value = invoice_no
    # Date: D9 (Row 9, Col 4)
    ws.cell(9, 4).value = now.strftime("%Y-%m-%d")

    # 客户信息
    cli_name_en = first_offer.get("cli_name_en", "") or first_offer.get("cli_name", "")
    # Company: C12 (Row 12, Col 3)
    ws.cell(12, 3).value = cli_name_en
    # Contact: C13
    ws.cell(13, 3).value = first_offer.get("contact_name", "") or ""
    # Email: C14
    ws.cell(14, 3).value = first_offer.get("email", "") or ""
    # Phone: C15
    ws.cell(15, 3).value = first_offer.get("phone", "") or ""
    # Address: C16
    ws.cell(16, 3).value = first_offer.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    header_row = 18  # 表头行
    first_data_row = 19  # 第一行数据
    template_data_rows = 2  # 模板中有2行示例数据 (Row 19-20)

    # 调整行数
    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        ws.insert_rows(first_data_row + template_data_rows, rows_diff)
        # 复制样式
        for i in range(rows_diff):
            new_row = first_data_row + template_data_rows + i
            for col in range(1, 9):
                src_cell = ws.cell(row=first_data_row + template_data_rows - 1, column=col)
                dest_cell = ws.cell(row=new_row, column=col)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.alignment = copy(src_cell.alignment)
    elif rows_diff < 0:
        ws.delete_rows(first_data_row + data_count, -rows_diff)

    # 写入报价数据
    for idx, offer in enumerate(offers):
        row = first_data_row + idx

        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = offer.get("inquiry_mpn", "") or offer.get("quoted_mpn", "")
        ws.cell(row, 3).value = offer.get("quoted_brand", "") or offer.get("inquiry_brand", "")
        ws.cell(row, 4).value = offer.get("date_code", "") or ""

        # Col E = Qty (数量)
        qty = offer.get("quoted_qty") or offer.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws.cell(row, 5).value = qty
        ws.cell(row, 5).number_format = '#,##0'

        # Col F = L/T (货期)
        ws.cell(row, 6).value = offer.get("delivery_date", "") or ""

        # Col G = Unit Price (单价 KRW)
        price_kwr = offer.get("calculated_price_kwr", "") or offer.get("offer_price_kwr", "") or 0
        try:
            price_kwr = float(price_kwr)
        except:
            price_kwr = 0
        ws.cell(row, 7).value = price_kwr
        ws.cell(row, 7).number_format = '#,##0.0'

        # Col H = Total Amount (公式: Qty * Unit Price)
        ws.cell(row, 8).value = f"=E{row}*G{row}"
        ws.cell(row, 8).number_format = '#,##0.0'

    # ---- 3. 完整拼接 Footer 内容（复制值、样式、合并单元格、行高、图片）----
    last_data_row = first_data_row + data_count - 1
    footer_start_row = last_data_row + 2
    footer_template_start = 11

    for src_row_idx in range(footer_template_start, ws_footer.max_row + 1):
        dest_row_idx = footer_start_row + (src_row_idx - footer_template_start)
        ws.insert_rows(dest_row_idx)

        for src_col in range(1, 9):
            src_cell = ws_footer.cell(row=src_row_idx, column=src_col)
            dest_cell = ws.cell(row=dest_row_idx, column=src_col)
            if src_cell.value is not None:
                dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = copy(src_cell.alignment)
        if ws_footer.row_dimensions[src_row_idx].height:
            ws.row_dimensions[dest_row_idx].height = ws_footer.row_dimensions[src_row_idx].height

    # 复制Footer的合并单元格
    for merged_range in ws_footer.merged_cells.ranges:
        if merged_range.min_row >= footer_template_start:
            new_min_row = footer_start_row + (merged_range.min_row - footer_template_start)
            new_max_row = footer_start_row + (merged_range.max_row - footer_template_start)
            try:
                ws.merge_cells(
                    start_row=new_min_row,
                    start_column=merged_range.min_col,
                    end_row=new_max_row,
                    end_column=merged_range.max_col
                )
            except:
                pass

    # 复制Footer中的图片（印章）
    from copy import deepcopy
    row_offset = footer_start_row - footer_template_start
    for img in ws_footer._images:
        new_img = deepcopy(img)
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            original_from_row = img.anchor._from.row
            if original_from_row >= footer_template_start - 1:
                new_img.anchor._from.row = original_from_row + row_offset
            if hasattr(img.anchor, 'to'):
                original_to_row = img.anchor.to.row
                if original_to_row >= footer_template_start - 1:
                    new_img.anchor.to.row = original_to_row + row_offset
        ws.add_image(new_img)

    # 更新 TOTAL 公式
    total_row = footer_start_row

    # 取消 Total 行的合并单元格（如果存在）
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == total_row and merged_range.max_col >= 8:
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass

    ws.cell(total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws.cell(total_row, 8).number_format = '#,##0.0'

    # 设置打印区域
    final_row = footer_start_row + (ws_footer.max_row - footer_template_start)
    ws.print_area = f"$A$1:$H${final_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return True, {
        "excel_path": output_path,
        "invoice_no": invoice_no,
        "cli_name": first_offer.get("cli_name", ""),
        "count": data_count
    }


def generate_pi_us_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    基于报价生成 Proforma Invoice (USD版本)

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not offer_ids:
        return False, "未提供报价编号"

    offers = get_offers_for_document(offer_ids)
    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份PI"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "pi")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "PI")

    # 获取汇率
    _, usd_val, _ = get_exchange_rates()

    # 计算USD价格 - 报价使用 offer_price_rmb
    for offer in offers:
        price_usd = offer.get("offer_price_usd")
        if not price_usd or float(price_usd or 0) == 0:
            price_rmb = offer.get("offer_price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                price_usd = round(float(price_rmb) * usd_val, 3) if usd_val else 0
            else:
                price_usd = 0
        else:
            price_usd = float(price_usd)
        offer["calculated_price_usd"] = price_usd

    # 将报价数据适配为订单格式，复用现有模板逻辑
    adapted_offers = []
    for offer in offers:
        adapted = {
            "inquiry_mpn": offer.get("inquiry_mpn") or offer.get("quoted_mpn", ""),
            "inquiry_brand": offer.get("quoted_brand") or offer.get("inquiry_brand", ""),
            "date_code": offer.get("date_code", ""),
            "quoted_qty": offer.get("quoted_qty") or offer.get("inquiry_qty", 0),
            "delivery_date": offer.get("delivery_date", ""),
            "calculated_price_usd": offer.get("calculated_price_usd", 0),
            "cli_name_en": offer.get("cli_name_en", "") or offer.get("cli_name", ""),
            "cli_name": offer.get("cli_name", ""),
            "contact_name": offer.get("contact_name", ""),
            "email": offer.get("email", ""),
            "phone": offer.get("phone", ""),
            "address": offer.get("address", ""),
        }
        adapted_offers.append(adapted)

    return _generate_pi_us_excel(adapted_offers, template_dir, output_path, invoice_no)


def generate_ci_us_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    基于报价生成美元版 Commercial Invoice

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not offer_ids:
        return False, "未提供报价编号"

    offers = get_offers_for_document(offer_ids)
    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份CI"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "ci_us")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "CI_US")

    # 获取汇率
    _, usd_val, _ = get_exchange_rates()

    # 计算USD价格
    for offer in offers:
        price_usd = offer.get("offer_price_usd")
        if not price_usd or float(price_usd or 0) == 0:
            price_rmb = offer.get("offer_price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                price_usd = round(float(price_rmb) * usd_val, 3) if usd_val else 0
            else:
                price_usd = 0
        else:
            price_usd = float(price_usd)
        offer["calculated_price_usd"] = price_usd

    # 适配报价数据为订单格式
    adapted_offers = []
    for offer in offers:
        adapted = {
            "inquiry_mpn": offer.get("inquiry_mpn") or offer.get("quoted_mpn", ""),
            "inquiry_brand": offer.get("quoted_brand") or offer.get("inquiry_brand", ""),
            "date_code": offer.get("date_code", ""),
            "quoted_qty": offer.get("quoted_qty") or offer.get("inquiry_qty", 0),
            "inquiry_qty": offer.get("inquiry_qty", 0),
            "delivery_date": offer.get("delivery_date", ""),
            "price_usd": offer.get("calculated_price_usd", 0),
            "cli_name_en": offer.get("cli_name_en", "") or offer.get("cli_name", ""),
            "cli_name": offer.get("cli_name", ""),
            "contact_name": offer.get("contact_name", ""),
            "email": offer.get("email", ""),
            "phone": offer.get("phone", ""),
            "address": offer.get("address", ""),
            "region": offer.get("region", "USA"),
        }
        adapted_offers.append(adapted)

    return _generate_ci_us_excel(adapted_offers, template_dir, output_path, invoice_no)


def generate_pi_jp_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    基于报价生成 Proforma Invoice (JPY日元版本)

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    if not offer_ids:
        return False, "未提供报价编号"

    offers = get_offers_for_document(offer_ids)
    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份PI"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    if not output_base:
        output_base = _get_output_base()

    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "pi")

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "PI_JP")

    # 使用日元价格字段
    for offer in offers:
        price_jpy = offer.get("price_jpy")
        if price_jpy and float(price_jpy or 0) > 0:
            offer["calculated_price_jpy"] = float(price_jpy)
        else:
            offer["calculated_price_jpy"] = 0

    # 将报价数据适配为订单格式，复用现有模板逻辑
    adapted_offers = []
    for offer in offers:
        adapted = {
            "inquiry_mpn": offer.get("inquiry_mpn") or offer.get("quoted_mpn", ""),
            "inquiry_brand": offer.get("quoted_brand") or offer.get("inquiry_brand", ""),
            "date_code": offer.get("date_code", ""),
            "quoted_qty": offer.get("quoted_qty") or offer.get("inquiry_qty", 0),
            "delivery_date": offer.get("delivery_date", ""),
            "calculated_price_jpy": offer.get("calculated_price_jpy", 0),
            "cli_name_en": offer.get("cli_name_en", "") or offer.get("cli_name", ""),
            "cli_name": offer.get("cli_name", ""),
            "contact_name": offer.get("contact_name", ""),
            "email": offer.get("email", ""),
            "phone": offer.get("phone", ""),
            "address": offer.get("address", ""),
        }
        adapted_offers.append(adapted)

    return _generate_pi_jp_excel(adapted_offers, template_dir, output_path, invoice_no)


def _generate_pi_jp_excel(offers, template_dir, output_path, invoice_no):
    """基于报价生成PI-JP Excel文件 - Header + Footer 模板拼接方式

    模板结构 (JP版本):
    Header模板 (JP - 1):
    - Row 8: Invoice No. (D4=值)
    - Row 9: Date (D4=值)
    - Row 12-16: 客户信息 (C3=值)
    - Row 18: 表头行
    - Row 19-20: 2行示例数据

    Footer模板 (JP - 2):
    - Row 11: TOTAL AMOUNT + 公式
    - Row 13-18: TERMS & CONDITIONS
    - Row 20-27: PAYMENT METHOD
    - Row 29: THANK YOU
    """
    from openpyxl.utils import get_column_letter
    from copy import copy, deepcopy

    data_count = len(offers)
    first_offer = offers[0]
    now = datetime.now()

    # 加载 Header 和 Footer 模板
    header_path = os.path.join(template_dir, "Proforma_Invoice_JP - 1.xlsx")
    footer_path = os.path.join(template_dir, "Proforma_Invoice_JP - 2.xlsx")

    if not os.path.exists(header_path):
        return False, f"Header模板不存在: {header_path}"
    if not os.path.exists(footer_path):
        return False, f"Footer模板不存在: {footer_path}"

    wb = openpyxl.load_workbook(header_path)
    ws = wb.active

    wb_footer = openpyxl.load_workbook(footer_path)
    ws_footer = wb_footer.active

    # ---- 1. 写入头部信息 ----
    ws.cell(8, 4).value = invoice_no
    ws.cell(9, 4).value = now.strftime("%Y-%m-%d")

    cli_name_en = first_offer.get("cli_name_en", "") or first_offer.get("cli_name", "")
    ws.cell(12, 3).value = cli_name_en
    ws.cell(13, 3).value = first_offer.get("contact_name", "") or ""
    ws.cell(14, 3).value = first_offer.get("email", "") or ""
    ws.cell(15, 3).value = first_offer.get("phone", "") or ""
    ws.cell(16, 3).value = first_offer.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    header_row = 18
    first_data_row = 19
    template_data_rows = 2

    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        ws.insert_rows(first_data_row + template_data_rows, rows_diff)
        for i in range(rows_diff):
            new_row = first_data_row + template_data_rows + i
            for col in range(1, 9):
                src_cell = ws.cell(row=first_data_row + template_data_rows - 1, column=col)
                dest_cell = ws.cell(row=new_row, column=col)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.alignment = copy(src_cell.alignment)
    elif rows_diff < 0:
        ws.delete_rows(first_data_row + data_count, -rows_diff)

    # 写入报价数据
    for idx, offer in enumerate(offers):
        row = first_data_row + idx

        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = offer.get("inquiry_mpn", "") or offer.get("quoted_mpn", "")
        ws.cell(row, 3).value = offer.get("quoted_brand", "") or offer.get("inquiry_brand", "")
        ws.cell(row, 4).value = offer.get("date_code", "") or ""

        # Col E = Qty (数量)
        qty = offer.get("quoted_qty") or offer.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws.cell(row, 5).value = qty
        ws.cell(row, 5).number_format = '#,##0'

        # Col F = L/T (货期)
        ws.cell(row, 6).value = offer.get("delivery_date", "") or ""

        # Col G = Unit Price (单价 JPY)
        price_jpy = offer.get("calculated_price_jpy", "") or offer.get("price_jpy", "") or 0
        try:
            price_jpy = float(price_jpy)
        except:
            price_jpy = 0
        ws.cell(row, 7).value = price_jpy
        ws.cell(row, 7).number_format = '#,##0.0'

        # Col H = Total Amount (公式: Qty * Unit Price)
        ws.cell(row, 8).value = f"=E{row}*G{row}"
        ws.cell(row, 8).number_format = '#,##0.0'

    # ---- 3. 完整拼接 Footer 内容（复制值、样式、合并单元格、行高、图片）----
    # JP-2 模板结构: Row 1=Total Amount, Row 3-8=TERMS, Row 10-11=BANK, Row 19=THANK YOU
    last_data_row = first_data_row + data_count - 1
    footer_start_row = last_data_row + 2
    footer_template_start = 1  # JP模板从Row 1开始（Total Amount行）

    for src_row_idx in range(footer_template_start, ws_footer.max_row + 1):
        dest_row_idx = footer_start_row + (src_row_idx - footer_template_start)
        ws.insert_rows(dest_row_idx)

        for src_col in range(1, 9):
            src_cell = ws_footer.cell(row=src_row_idx, column=src_col)
            dest_cell = ws.cell(row=dest_row_idx, column=src_col)
            if src_cell.value is not None:
                dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = copy(src_cell.alignment)
        if ws_footer.row_dimensions[src_row_idx].height:
            ws.row_dimensions[dest_row_idx].height = ws_footer.row_dimensions[src_row_idx].height

    # 复制Footer的合并单元格
    for merged_range in ws_footer.merged_cells.ranges:
        if merged_range.min_row >= footer_template_start:
            new_min_row = footer_start_row + (merged_range.min_row - footer_template_start)
            new_max_row = footer_start_row + (merged_range.max_row - footer_template_start)
            try:
                ws.merge_cells(
                    start_row=new_min_row,
                    start_column=merged_range.min_col,
                    end_row=new_max_row,
                    end_column=merged_range.max_col
                )
            except:
                pass

    # 复制Footer中的图片（印章）
    row_offset = footer_start_row - footer_template_start
    for img in ws_footer._images:
        new_img = deepcopy(img)
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            original_from_row = img.anchor._from.row
            if original_from_row >= footer_template_start - 1:
                new_img.anchor._from.row = original_from_row + row_offset
            if hasattr(img.anchor, 'to'):
                original_to_row = img.anchor.to.row
                if original_to_row >= footer_template_start - 1:
                    new_img.anchor.to.row = original_to_row + row_offset
        ws.add_image(new_img)

    # 更新 TOTAL 公式
    total_row = footer_start_row

    # 取消 Total 行的合并单元格（如果存在）
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == total_row and merged_range.max_col >= 8:
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass

    ws.cell(total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws.cell(total_row, 8).number_format = '#,##0.0'

    # 设置打印区域
    final_row = footer_start_row + (ws_footer.max_row - footer_template_start)
    ws.print_area = f"$A$1:$H${final_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    wb.close()
    wb_footer.close()

    return True, {
        "excel_path": output_path,
        "invoice_no": invoice_no,
        "cli_name": first_offer.get("cli_name", ""),
        "count": data_count
    }


def generate_pi_auto_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    智能判断币种并生成PI

    根据报价中的 price_kwr, price_usd, price_jpy 字段自动选择模板：
    - 优先级: price_kwr > price_usd > price_jpy（选择第一个非零值）

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if not offer_ids:
        return False, "未提供报价编号"

    # 获取报价数据
    offers = get_offers_for_document(offer_ids)

    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    # 检查币种：选择第一个报价中第一个非零的币种字段
    currency_type = None
    for offer in offers:
        price_kwr = float(offer.get("price_kwr") or 0)
        price_usd = float(offer.get("price_usd") or 0)
        price_jpy = float(offer.get("price_jpy") or 0)

        if price_kwr > 0:
            currency_type = "KRW"
            break
        elif price_usd > 0:
            currency_type = "USD"
            break
        elif price_jpy > 0:
            currency_type = "JPY"
            break

    # 默认使用KRW（如果没有币种字段）
    if not currency_type:
        currency_type = "KRW"

    # 根据币种调用对应的生成函数
    if currency_type == "KRW":
        return generate_pi_from_offers(offer_ids, output_base, template_dir)
    elif currency_type == "USD":
        return generate_pi_us_from_offers(offer_ids, output_base, template_dir)
    elif currency_type == "JPY":
        return generate_pi_jp_from_offers(offer_ids, output_base, template_dir)