"""
CI Generator - Commercial Invoice 生成模块

提供 CI 文件生成功能，供 main.py 和 skill 脚本共同调用。
"""

import os
import platform
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
except ImportError:
    openpyxl = None

from Sills.base import get_db_connection, get_exchange_rates


# ============================================================
# 样式配置 - 商务蓝风格
# ============================================================

COLOR_HEADER_BG = "1E3A5F"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ODD = "F8FAFC"
COLOR_ROW_EVEN = "FFFFFF"
COLOR_TOTAL_BG = "1E3A5F"
COLOR_TOTAL_FONT = "FFFFFF"

BORDER_THIN = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

BORDER_MEDIUM = Border(
    left=Side(style='medium', color='1E3A5F'),
    right=Side(style='medium', color='1E3A5F'),
    top=Side(style='medium', color='1E3A5F'),
    bottom=Side(style='medium', color='1E3A5F')
)


def _get_default_output_base():
    """获取默认输出目录 - 兼容 Windows 和 WSL2"""
    if platform.system() == "Windows":
        return r"E:\1_Business\1_Auto"
    else:
        # WSL2 或其他 Linux 系统
        return "/mnt/e/1_Business/1_Auto"


def _get_output_base():
    """获取输出基础目录"""
    output_base = os.environ.get('UNIULTRA_OUTPUT_DIR')
    if output_base:
        return output_base
    return _get_default_output_base()


def _generate_unique_invoice_no(output_dir, cli_name, doc_type="CI"):
    """
    生成唯一的发票编号和文件路径

    发票编号格式: UNI%Y%m%d，如果已存在则添加01, 02, 03等后缀
    文件名格式: COMMERCIAL INVOICE_{cli_name}_{invoice_no}.xlsx

    Args:
        output_dir: 输出目录
        cli_name: 客户名称
        doc_type: 文档类型 (CI)

    Returns:
        tuple: (invoice_no, output_filename, output_path)
    """
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    base_invoice_no = f"UNI{date_str}"

    # CI固定使用 COMMERCIAL INVOICE 前缀
    doc_prefix = "COMMERCIAL INVOICE"

    # 尝试不带后缀
    invoice_no = base_invoice_no
    output_filename = f"{doc_prefix}_{cli_name}_{invoice_no}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    suffix = 1
    while os.path.exists(output_path):
        invoice_no = f"{base_invoice_no}{suffix:02d}"
        output_filename = f"{doc_prefix}_{cli_name}_{invoice_no}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        suffix += 1

    return invoice_no, output_filename, output_path


def get_header_style():
    return {
        "font": Font(name='Arial', size=8, bold=True, color=COLOR_HEADER_FONT),
        "fill": PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid'),
        "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True),
        "border": BORDER_THIN
    }


def get_data_row_style(row_index):
    bg_color = COLOR_ROW_ODD if row_index % 2 == 1 else COLOR_ROW_EVEN
    return {
        "font": Font(name='Arial', size=9),
        "fill": PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid'),
        "alignment": Alignment(horizontal='center', vertical='center'),
        "border": BORDER_THIN
    }


def get_total_style():
    return {
        "font": Font(name='Arial', size=9, bold=True, color=COLOR_TOTAL_FONT),
        "fill": PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type='solid'),
        "alignment": Alignment(horizontal='right', vertical='center'),
        "border": BORDER_MEDIUM
    }


# ============================================================
# 核心函数
# ============================================================

def calculate_price_kwr(order, exchange_rate_krw):
    """计算韩元价格"""
    price_kwr = order.get("price_kwr")
    if not price_kwr or float(price_kwr or 0) == 0:
        price_rmb = order.get("price_rmb")
        if price_rmb and float(price_rmb or 0) > 0:
            price_kwr = round(float(price_rmb) * exchange_rate_krw, 1)
        else:
            price_kwr = 0
    else:
        price_kwr = float(price_kwr)
    return price_kwr


def get_orders_for_ci(order_ids):
    """获取订单列表（用于生成CI文档）"""
    if not order_ids:
        return []
    placeholders = ','.join(['?'] * len(order_ids))
    with get_db_connection() as conn:
        rows = conn.execute(f"""
            SELECT o.order_id, o.order_no, o.order_date, o.cli_id,
                   o.inquiry_mpn, o.inquiry_brand, o.price_rmb, o.price_kwr, o.price_usd,
                   o.offer_id, o.cost_price_rmb,
                   c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
                   c.cli_full_name, c.region,
                   off.quoted_qty, off.date_code, off.delivery_date, off.inquiry_qty
            FROM uni_order o
            LEFT JOIN uni_cli c ON o.cli_id = c.cli_id
            LEFT JOIN uni_offer off ON o.offer_id = off.offer_id
            WHERE o.order_id IN ({placeholders})
            ORDER BY o.order_id
        """, order_ids).fetchall()
        return [dict(r) for r in rows]


def generate_ci_excel(orders, template_dir, output_path, invoice_no, header_filename="CI_template_header.xlsx", footer_filename="CI_template_footer.xlsx"):
    """
    使用切分模板模式生成CI Excel文件

    Args:
        orders: 订单数据列表
        template_dir: 模板目录路径
        output_path: 输出文件路径
        invoice_no: 发票编号
        header_filename: Header模板文件名（默认 CI_template_header.xlsx）
        footer_filename: Footer模板文件名（默认 CI_template_footer.xlsx）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if openpyxl is None:
        return False, "缺少依赖: 请安装 openpyxl -> pip install openpyxl"

    data_count = len(orders)
    if data_count == 0:
        return False, "没有订单数据"

    first_order = orders[0]
    now = datetime.now()

    # 加载模板（支持自定义文件名）
    header_path = os.path.join(template_dir, header_filename)
    footer_path = os.path.join(template_dir, footer_filename)

    if not os.path.exists(header_path):
        return False, f"Header模板不存在: {header_path}"
    if not os.path.exists(footer_path):
        return False, f"Footer模板不存在: {footer_path}"

    wb = openpyxl.load_workbook(header_path)
    ws = wb.active

    # 填写头部信息
    invoice_date = now.strftime("%Y-%m-%d")

    ws['B2'] = invoice_no
    ws['F2'] = invoice_date

    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws['B9'] = cli_name_en
    ws['F9'] = first_order.get("contact_name", "") or ""
    ws['B10'] = first_order.get("region", "") or "韩国"
    ws['F10'] = first_order.get("phone", "") or ""
    ws['B11'] = first_order.get("address", "") or ""

    # 动态生成数据行
    data_start_row = 13
    template_data_rows = 3

    if data_count > template_data_rows:
        extra_rows = data_count - template_data_rows
        ws.insert_rows(data_start_row + template_data_rows, extra_rows)

        for i in range(extra_rows):
            src_row = data_start_row + template_data_rows - 1
            dst_row = data_start_row + template_data_rows + i
            for col in range(1, 8):
                src_cell = ws.cell(row=src_row, column=col)
                dst_cell = ws.cell(row=dst_row, column=col)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)
            if ws.row_dimensions[src_row].height:
                ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    # 获取汇率
    krw_val, _, _ = get_exchange_rates()

    # 填充数据
    total_qty = 0
    total_amount = 0

    for idx, order in enumerate(orders):
        row = data_start_row + idx

        # 先取消该行可能存在的合并单元格（关键：必须在获取cell之前取消）
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row == row:
                ws.unmerge_cells(str(merged_range))

        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = "集成电路/IC"
        ws.cell(row=row, column=3).value = order.get("inquiry_mpn", "") or ""
        ws.cell(row=row, column=4).value = "8542399000"

        qty = order.get("quoted_qty") or order.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws.cell(row=row, column=5).value = qty
        ws.cell(row=row, column=5).number_format = '#,##0'
        total_qty += qty

        price_kwr = calculate_price_kwr(order, krw_val)
        ws.cell(row=row, column=6).value = price_kwr
        ws.cell(row=row, column=6).number_format = '#,##0'

        total = price_kwr * qty
        ws.cell(row=row, column=7).value = total if total else 0
        ws.cell(row=row, column=7).number_format = '#,##0'
        total_amount += total

    # 追加 Footer
    footer_start_row = data_start_row + data_count

    wb_footer = openpyxl.load_workbook(footer_path)
    ws_footer = wb_footer.active

    for src_row in range(1, 4):
        dst_row = footer_start_row + src_row - 1
        for col in range(1, 8):
            src_cell = ws_footer.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)
        if ws_footer.row_dimensions[src_row].height:
            ws.row_dimensions[dst_row].height = ws_footer.row_dimensions[src_row].height

    # 复制 footer 中的合并单元格
    for merged_range in ws_footer.merged_cells.ranges:
        new_range = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{footer_start_row + merged_range.min_row - 1}:{openpyxl.utils.get_column_letter(merged_range.max_col)}{footer_start_row + merged_range.max_row - 1}"
        try:
            ws.merge_cells(new_range)
        except:
            pass

    ws.cell(row=footer_start_row, column=5).value = total_qty
    ws.cell(row=footer_start_row, column=5).number_format = '#,##0'

    ws.cell(row=footer_start_row + 1, column=7).value = total_amount
    ws.cell(row=footer_start_row + 1, column=7).number_format = '#,##0'

    # 复制印章图片
    if ws_footer._images:
        img = ws_footer._images[0]
        new_img = Image(img.ref)
        import zipfile
        import re
        try:
            with zipfile.ZipFile(footer_path, 'r') as z:
                drawing_content = z.read('xl/drawings/drawing1.xml').decode('utf-8')
                cx_match = re.search(r'cx="(\d+)"', drawing_content)
                cy_match = re.search(r'cy="(\d+)"', drawing_content)
                if cx_match and cy_match:
                    new_img.width = int(cx_match.group(1)) / 9525
                    new_img.height = int(cy_match.group(1)) / 9525
        except:
            new_img.width = img.width
            new_img.height = img.height

        target_row = footer_start_row + 2
        ws.add_image(new_img, f'C{target_row}')

    last_row = footer_start_row + 2
    ws.print_area = f"$A$1:$G${last_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    # 生成 PDF
    pdf_path = ""
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "excel_to_pdf",
            os.path.join(os.path.dirname(os.path.dirname(__file__)),
                        "openclaw_skills", "order-ci-generator-kr", "scripts", "excel_to_pdf.py")
        )
        if spec and spec.loader:
            pdf_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(pdf_module)
            pdf_success, pdf_result = pdf_module.convert_to_pdf(output_path)
            if pdf_success:
                pdf_path = pdf_result
    except Exception as e:
        pass  # PDF 生成失败不影响 Excel

    return True, {
        "excel_path": output_path,
        "pdf_path": pdf_path,
        "invoice_no": invoice_no,
        "total_qty": total_qty,
        "total_amount": total_amount,
        "cli_name": first_order.get("cli_name", ""),
        "count": data_count
    }


def generate_ci_kr(order_ids, output_base=None, template_dir=None):
    """
    生成韩国版 Commercial Invoice

    Args:
        order_ids: 订单ID列表
        output_base: 输出基础目录（可选，默认使用环境变量或项目 Trans 目录）
        template_dir: 模板目录（可选，默认使用项目 templates 目录）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if not order_ids:
        return False, "未提供订单编号"

    # 获取订单数据
    orders = get_orders_for_ci(order_ids)

    if not orders:
        return False, f"订单编号 {order_ids} 不存在"

    # 检查客户一致性
    cli_names = set(o.get("cli_name") or "未知客户" for o in orders)
    if len(cli_names) > 1:
        return False, f"订单属于不同客户 ({', '.join(cli_names)})，无法生成同一份CI"

    cli_name = list(cli_names)[0]
    # 获取客户英文名，用于文件名
    cli_name_en = orders[0].get("cli_name_en") or orders[0].get("cli_name") or "Unknown"

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if not output_base:
        output_base = _get_output_base()

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "CI")

    # 确定模板目录
    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "ci_kr")

    # 生成 CI
    return generate_ci_excel(orders, template_dir, output_path, invoice_no)


def generate_ci_kr_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    基于报价生成韩国版 Commercial Invoice

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if not offer_ids:
        return False, "未提供报价编号"

    # 获取报价数据
    offers = get_offers_for_ci(offer_ids)

    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    # 检查客户一致性
    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份CI"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if not output_base:
        output_base = _get_output_base()

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "CI")

    # 确定模板目录
    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "ci_kr")

    # 获取汇率
    krw_val, _, _ = get_exchange_rates()

    # 计算韩元价格
    for offer in offers:
        price_rmb = offer.get("offer_price_rmb")
        if price_rmb and float(price_rmb or 0) > 0:
            offer["price_kwr"] = round(float(price_rmb) * krw_val, 1)
        else:
            offer["price_kwr"] = 0

    # 适配报价数据为订单格式
    adapted_offers = []
    for offer in offers:
        adapted = {
            "inquiry_mpn": offer.get("inquiry_mpn") or offer.get("quoted_mpn", ""),
            "inquiry_brand": offer.get("quoted_brand") or offer.get("inquiry_brand", ""),
            "date_code": offer.get("date_code", ""),
            "quoted_qty": offer.get("quoted_qty") or offer.get("inquiry_qty", 0),
            "inquiry_qty": offer.get("inquiry_qty", 0),
            "delivery_date": offer.get("delivery_date", ""),
            "price_kwr": offer.get("price_kwr", 0),
            "price_rmb": offer.get("offer_price_rmb", 0),
            "cli_name_en": offer.get("cli_name_en", "") or offer.get("cli_name", ""),
            "cli_name": offer.get("cli_name", ""),
            "contact_name": offer.get("contact_name", ""),
            "email": offer.get("email", ""),
            "phone": offer.get("phone", ""),
            "address": offer.get("address", ""),
        }
        adapted_offers.append(adapted)

    # 生成 CI
    return generate_ci_excel(adapted_offers, template_dir, output_path, invoice_no)


def get_offers_for_ci(offer_ids):
    """获取报价列表（用于生成CI文档）"""
    if not offer_ids:
        return []
    placeholders = ','.join(['?'] * len(offer_ids))
    with get_db_connection() as conn:
        rows = conn.execute(f"""
            SELECT o.offer_id, o.offer_date,
                   o.quoted_mpn, o.quoted_brand, o.offer_price_rmb, o.price_usd as offer_price_usd,
                   o.quoted_qty, o.date_code, o.delivery_date, o.inquiry_qty, o.inquiry_mpn,
                   COALESCE(o.cli_id, q.cli_id) as cli_id,
                   c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
                   c.cli_full_name, c.region
            FROM uni_offer o
            LEFT JOIN uni_quote q ON o.quote_id = q.quote_id
            LEFT JOIN uni_cli c ON COALESCE(o.cli_id, q.cli_id) = c.cli_id
            WHERE o.offer_id IN ({placeholders})
            ORDER BY o.offer_id
        """, offer_ids).fetchall()
        return [dict(r) for r in rows]


def get_offers_for_ci_with_currency(offer_ids):
    """获取报价列表（用于生成CI文档，包含币种字段）"""
    if not offer_ids:
        return []
    placeholders = ','.join(['?'] * len(offer_ids))
    with get_db_connection() as conn:
        rows = conn.execute(f"""
            SELECT o.offer_id, o.offer_date,
                   o.quoted_mpn, o.quoted_brand, o.offer_price_rmb,
                   o.price_kwr, o.price_usd, o.price_jpy,
                   o.quoted_qty, o.date_code, o.delivery_date, o.inquiry_qty, o.inquiry_mpn,
                   COALESCE(o.cli_id, q.cli_id) as cli_id,
                   c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
                   c.cli_full_name, c.region
            FROM uni_offer o
            LEFT JOIN uni_quote q ON o.quote_id = q.quote_id
            LEFT JOIN uni_cli c ON COALESCE(o.cli_id, q.cli_id) = c.cli_id
            WHERE o.offer_id IN ({placeholders})
            ORDER BY o.offer_id
        """, offer_ids).fetchall()
        return [dict(r) for r in rows]


def generate_ci_jp_from_offers(offer_ids, output_base=None, template_dir=None):
    """
    基于报价生成日元版 Commercial Invoice

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）
        template_dir: 模板目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if not offer_ids:
        return False, "未提供报价编号"

    # 获取报价数据
    offers = get_offers_for_ci_with_currency(offer_ids)

    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    # 检查客户一致性
    cli_names = set(o.get("cli_name") or "未知客户" for o in offers)
    if len(cli_names) > 1:
        return False, f"报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份CI"

    cli_name = list(cli_names)[0]

    # 确定输出目录
    project_root = os.environ.get('UNIULTRA_PROJECT_ROOT',
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if not output_base:
        output_base = _get_output_base()

    now = datetime.now()
    date_dir = now.strftime("%Y%m%d")

    output_dir = os.path.join(output_base, cli_name, date_dir)
    invoice_no, output_filename, output_path = _generate_unique_invoice_no(output_dir, cli_name, "CI_JP")

    # 确定模板目录（JP模板文件名带_JP后缀）
    if not template_dir:
        template_dir = os.path.join(project_root, "templates", "ci_jp")

    # 使用日元价格字段
    for offer in offers:
        price_jpy = offer.get("price_jpy")
        if price_jpy and float(price_jpy or 0) > 0:
            offer["price_jpy"] = float(price_jpy)
        else:
            offer["price_jpy"] = 0

    # 适配报价数据为订单格式（将 price_jpy 映射为 price_kwr，因为模板结构相同）
    adapted_offers = []
    for offer in offers:
        adapted = {
            "inquiry_mpn": offer.get("inquiry_mpn") or offer.get("quoted_mpn", ""),
            "inquiry_brand": offer.get("quoted_brand") or offer.get("inquiry_brand", ""),
            "date_code": offer.get("date_code", ""),
            "quoted_qty": offer.get("quoted_qty") or offer.get("inquiry_qty", 0),
            "inquiry_qty": offer.get("inquiry_qty", 0),
            "delivery_date": offer.get("delivery_date", ""),
            "price_kwr": offer.get("price_jpy", 0),  # JPY 价格映射到 kwr 字段
            "price_rmb": offer.get("offer_price_rmb", 0),
            "cli_name_en": offer.get("cli_name_en", "") or offer.get("cli_name", ""),
            "cli_name": offer.get("cli_name", ""),
            "contact_name": offer.get("contact_name", ""),
            "email": offer.get("email", ""),
            "phone": offer.get("phone", ""),
            "address": offer.get("address", ""),
        }
        adapted_offers.append(adapted)

    # 生成 CI-JP（使用JP模板文件名）
    return generate_ci_excel(adapted_offers, template_dir, output_path, invoice_no,
                             header_filename="CI_template_header_JP.xlsx",
                             footer_filename="CI_template_footer_JP.xlsx")


def generate_ci_auto_from_offers(offer_ids, output_base=None):
    """
    智能判断币种并生成CI

    根据报价中的 price_kwr, price_usd, price_jpy 字段自动选择模板：
    - 优先级: price_kwr > price_usd > price_jpy（选择第一个非零值）

    Args:
        offer_ids: 报价ID列表
        output_base: 输出基础目录（可选）

    Returns:
        tuple: (success: bool, result: dict or error_message: str)
    """
    if not offer_ids:
        return False, "未提供报价编号"

    # 获取报价数据（包含币种字段）
    offers = get_offers_for_ci_with_currency(offer_ids)

    if not offers:
        return False, f"报价编号 {offer_ids} 不存在"

    # 检查币种：选择第一个报价中第一个非零的币种字段
    currency_type = None
    for offer in offers:
        price_kwr = float(offer.get("price_kwr") or 0)
        price_usd = float(offer.get("price_usd") or 0)
        price_jpy = float(offer.get("price_jpy") or 0)

        if price_kwr > 0:
            currency_type = "KRW"
            break
        elif price_usd > 0:
            currency_type = "USD"
            break
        elif price_jpy > 0:
            currency_type = "JPY"
            break

    # 默认使用KRW（如果没有币种字段，使用RMB换算）
    if not currency_type:
        currency_type = "KRW"

    # 根据币种调用对应的生成函数
    if currency_type == "KRW":
        return generate_ci_kr_from_offers(offer_ids, output_base)
    elif currency_type == "USD":
        from Sills.document_generator import generate_ci_us_from_offers
        return generate_ci_us_from_offers(offer_ids, output_base)
    elif currency_type == "JPY":
        return generate_ci_jp_from_offers(offer_ids, output_base)

    return False, "未知币种类型"