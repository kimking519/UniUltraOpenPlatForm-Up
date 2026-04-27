import sqlite3
from Sills.base import get_db_connection


def sync_cli_marketing_status(cli_id=None):
    """
    同步客户营销状态
    - is_contacted: 有联系人记录
    - has_inquiry: 有询价记录
    - has_order: 有订单记录

    cli_id: 指定客户ID，为None则同步所有客户
    """
    try:
        with get_db_connection() as conn:
            if cli_id:
                # 同步单个客户
                # 检查是否有联系人
                has_contact = conn.execute(
                    "SELECT 1 FROM uni_contact WHERE cli_id = ? LIMIT 1",
                    (cli_id,)
                ).fetchone()

                # 检查是否有询价
                has_inquiry = conn.execute(
                    "SELECT 1 FROM uni_quote WHERE cli_id = ? LIMIT 1",
                    (cli_id,)
                ).fetchone()

                # 检查是否有订单
                has_order = conn.execute(
                    "SELECT 1 FROM uni_order WHERE cli_id = ? LIMIT 1",
                    (cli_id,)
                ).fetchone()

                conn.execute("""
                    UPDATE uni_cli
                    SET is_contacted = ?, has_inquiry = ?, has_order = ?
                    WHERE cli_id = ?
                """, (
                    1 if has_contact else 0,
                    1 if has_inquiry else 0,
                    1 if has_order else 0,
                    cli_id
                ))
            else:
                # 同步所有客户
                # 设置 is_contacted
                conn.execute("""
                    UPDATE uni_cli
                    SET is_contacted = CASE
                        WHEN EXISTS (SELECT 1 FROM uni_contact WHERE uni_contact.cli_id = uni_cli.cli_id)
                        THEN 1 ELSE 0
                    END
                """)

                # 设置 has_inquiry
                conn.execute("""
                    UPDATE uni_cli
                    SET has_inquiry = CASE
                        WHEN EXISTS (SELECT 1 FROM uni_quote WHERE uni_quote.cli_id = uni_cli.cli_id)
                        THEN 1 ELSE 0
                    END
                """)

                # 设置 has_order
                conn.execute("""
                    UPDATE uni_cli
                    SET has_order = CASE
                        WHEN EXISTS (SELECT 1 FROM uni_order WHERE uni_order.cli_id = uni_cli.cli_id)
                        THEN 1 ELSE 0
                    END
                """)

            conn.commit()
            return True, "状态同步成功"
    except Exception as e:
        return False, str(e)


def update_cli_domain_from_email():
    """
    根据邮箱自动提取域名并更新客户表
    """
    try:
        with get_db_connection() as conn:
            # 从邮箱字段提取域名
            rows = conn.execute("""
                SELECT cli_id, email FROM uni_cli
                WHERE email IS NOT NULL AND email != '' AND (domain IS NULL OR domain = '')
            """).fetchall()

            updated = 0
            for row in rows:
                email = row[1]
                if email and '@' in email:
                    domain = email.split('@')[-1].lower().strip()
                    conn.execute("UPDATE uni_cli SET domain = ? WHERE cli_id = ?", (domain, row[0]))
                    updated += 1

            conn.commit()
            return True, f"更新了 {updated} 个客户的域名"
    except Exception as e:
        return False, str(e)


def get_cli_list(page=1, page_size=10, search_kw=""):
    offset = (page - 1) * page_size
    query = """
    SELECT * FROM uni_cli 
    WHERE cli_name LIKE ? OR region LIKE ? OR cli_id LIKE ?
    ORDER BY created_at DESC
    LIMIT ? OFFSET ?
    """
    count_query = "SELECT COUNT(*) FROM uni_cli WHERE cli_name LIKE ? OR region LIKE ? OR cli_id LIKE ?"
    params = (f"%{search_kw}%", f"%{search_kw}%", f"%{search_kw}%")
    
    with get_db_connection() as conn:
        total = conn.execute(count_query, params).fetchone()[0]
        items = conn.execute(query, params + (page_size, offset)).fetchall()
        
        results = [
            {k: ("" if v is None else v) for k, v in dict(row).items()}
            for row in items
        ]
        return results, total

def get_next_cli_id():
    with get_db_connection() as conn:
        row = conn.execute("SELECT MAX(cli_id) FROM uni_cli").fetchone()
        if row and row[0]:
            # Assuming format like C001
            num_part = ''.join(filter(str.isdigit, row[0]))
            next_num = int(num_part) + 1 if num_part else 1
            return f"C{next_num:03d}"
        return "C001"

def add_cli(data):
    try:
        if 'cli_id' not in data or not data['cli_id']:
            data['cli_id'] = get_next_cli_id()

        # 检查客户名称是否已存在
        cli_name = data.get('cli_name', '')
        if cli_name:
            with get_db_connection() as conn:
                existing = conn.execute(
                    "SELECT cli_id FROM uni_cli WHERE cli_name = ?",
                    (cli_name,)
                ).fetchone()
                if existing:
                    return False, f"客户名称 '{cli_name}' 已存在"

        # Defaults
        if 'region' not in data or not data['region']:
            data['region'] = '韩国'
        if 'credit_level' not in data or not data['credit_level']:
            data['credit_level'] = 'A'
        if 'margin_rate' not in data or not str(data['margin_rate']).strip():
            data['margin_rate'] = 10.0

        columns = ', '.join(data.keys())
        placeholders = ', '.join(['?'] * len(data))
        sql = f"INSERT INTO uni_cli ({columns}) VALUES ({placeholders})"

        with get_db_connection() as conn:
            conn.execute(sql, list(data.values()))
            conn.commit()
            return True, f"客户 {data['cli_id']} 添加成功"
    except Exception as e:
        return False, str(e)

def batch_import_cli_text(text):
    lines = text.strip().split('\n')
    success_count = 0
    errors = []
    for line in lines:
        parts = [p.strip() for p in line.split(',')]
        if len(parts) < 1: continue
        
        data = {
            "cli_name": parts[0],
            "region": parts[1] if len(parts) > 1 else "韩国",
            "credit_level": parts[2] if len(parts) > 2 else "A",
            "margin_rate": parts[3] if len(parts) > 3 else 10.0,
            "emp_id": parts[4] if len(parts) > 4 else "000",
            "website": parts[5] if len(parts) > 5 else "",
            "payment_terms": parts[6] if len(parts) > 6 else "",
            "email": parts[7] if len(parts) > 7 else "",
            "phone": parts[8] if len(parts) > 8 else "",
            "remark": parts[9] if len(parts) > 9 else ""
        }
        ok, msg = add_cli(data)
        if ok: success_count += 1
        else: errors.append(f"{parts[0]}: {msg}")
    
    return success_count, errors

def update_cli(cli_id, data):
    try:
        set_cols = []
        params = []
        for k, v in data.items():
            set_cols.append(f"{k} = ?")
            params.append(v)
        
        params.append(cli_id)
        sql = f"UPDATE uni_cli SET {(', '.join(set_cols))} WHERE cli_id = ?"
        
        with get_db_connection() as conn:
            conn.execute(sql, params)
            conn.commit()
            return True, "更新成功"
    except Exception as e:
        return False, str(e)

def delete_cli(cli_id):
    try:
        with get_db_connection() as conn:
            conn.execute("DELETE FROM uni_cli WHERE cli_id = ?", (cli_id,))
            conn.commit()
            return True, "删除成功"
    except Exception as e:
        return False, str(e)

def batch_delete_cli(cli_ids):
    """批量删除客户，返回成功和失败数量"""
    if not cli_ids:
        return 0, 0, "未选择记录"

    deleted_count = 0
    failed_count = 0

    with get_db_connection() as conn:
        for cli_id in cli_ids:
            try:
                conn.execute("DELETE FROM uni_cli WHERE cli_id = ?", (cli_id,))
                deleted_count += 1
            except Exception:
                # 外键约束导致删除失败
                failed_count += 1

        conn.commit()

    return deleted_count, failed_count, "批量删除完成"


def export_cli_to_excel(output_path=None):
    """导出客户数据到Excel

    Args:
        output_path: 输出文件路径或BytesIO对象（可选，默认自动生成）

    Returns:
        (success, file_path_or_bytesio) tuple
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        import io

        with get_db_connection() as conn:
            clients = conn.execute("""
                SELECT cli_id, cli_name, cli_full_name, cli_name_en, contact_name,
                       address, region, credit_level, margin_rate, emp_id, website,
                       payment_terms, email, phone, remark, domain,
                       is_contacted, has_inquiry, has_order, created_at
                FROM uni_cli
                ORDER BY created_at DESC
            """).fetchall()

        if not clients:
            return False, "没有客户数据"

        wb = Workbook()
        ws = wb.active
        ws.title = "客户数据"

        # 设置表头样式
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            '客户编号', '客户名称', '公司全名', '公司英文名', '联系人',
            '地址', '所属区域', '信用等级', '利润率(%)', '负责员工',
            '网站', '账期', '邮箱', '电话', '备注', '域名',
            '有联系人', '有询价', '有订单', '创建时间'
        ]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, client in enumerate(clients, 2):
            data = [
                client[0] or '',   # cli_id
                client[1] or '',   # cli_name
                client[2] or '',   # cli_full_name
                client[3] or '',   # cli_name_en
                client[4] or '',   # contact_name
                client[5] or '',   # address
                client[6] or '',   # region
                client[7] or '',   # credit_level
                client[8] or 0,    # margin_rate
                client[9] or '',   # emp_id
                client[10] or '',  # website
                client[11] or '',  # payment_terms
                client[12] or '',  # email
                client[13] or '',  # phone
                client[14] or '',  # remark
                client[15] or '',  # domain
                '是' if client[16] else '否',  # is_contacted
                '是' if client[17] else '否',  # has_inquiry
                '是' if client[18] else '否',  # has_order
                client[19] or ''   # created_at
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

        # 调整列宽
        column_widths = [12, 20, 30, 25, 15, 30, 10, 10, 12, 12, 25, 15, 30, 20, 30, 25, 10, 10, 10, 20]
        for col, width in enumerate(column_widths, 1):
            col_letter = chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)
            ws.column_dimensions[col_letter].width = width

        # 判断输出类型
        if isinstance(output_path, io.BytesIO):
            wb.save(output_path)
            return True, output_path
        else:
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                output_path = f"cli_export_{timestamp}.xlsx"
            wb.save(output_path)
            return True, output_path

    except Exception as e:
        return False, str(e)
