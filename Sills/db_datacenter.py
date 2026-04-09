"""
数据中心数据库操作模块

提供表结构查询、SQL 执行、查询保存等功能
"""

from typing import List, Dict, Any, Optional, Tuple
from Sills.base import get_db_connection


def init_saved_queries_table():
    """初始化 saved_queries 表"""
    with get_db_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS saved_queries (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) UNIQUE NOT NULL,
                sql TEXT NOT NULL,
                created_by VARCHAR(50),
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()


def get_all_tables() -> List[Dict[str, Any]]:
    """
    获取所有用户表

    Returns:
        表名列表，包含表名和注释
    """
    with get_db_connection() as conn:
        rows = conn.execute("""
            SELECT
                table_name,
                obj_description((table_schema || '.' || table_name)::regclass, 'pg_class') as table_comment
            FROM information_schema.tables
            WHERE table_schema = 'public'
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """).fetchall()
        return [dict(row) for row in rows]


def get_table_structure(table_name: str) -> List[Dict[str, Any]]:
    """
    获取表结构

    Args:
        table_name: 表名

    Returns:
        列信息列表（列名、数据类型、是否可空、默认值、注释）
    """
    with get_db_connection() as conn:
        rows = conn.execute("""
            SELECT
                column_name,
                data_type,
                character_maximum_length,
                is_nullable,
                column_default,
                col_description((table_schema || '.' || table_name)::regclass, ordinal_position) as column_comment
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            ORDER BY ordinal_position
        """, (table_name,)).fetchall()
        return [dict(row) for row in rows]


def get_table_data(table_name: str, limit: int = 100) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    获取表数据

    Args:
        table_name: 表名
        limit: 最大返回行数

    Returns:
        (数据列表, 列名列表)
    """
    # 安全检查：防止 SQL 注入
    if not table_name.replace('_', '').isalnum():
        return [], []

    with get_db_connection() as conn:
        # 先获取列名
        structure = get_table_structure(table_name)
        columns = [col['column_name'] for col in structure]

        # 获取数据
        rows = conn.execute(f"SELECT * FROM {table_name} LIMIT %s", (limit,)).fetchall()
        data = [dict(row) for row in rows]

        return data, columns


def execute_sql(sql: str, timeout: int = 30) -> Tuple[bool, Any, str]:
    """
    执行 SQL 语句

    Args:
        sql: SQL 语句
        timeout: 超时时间（秒）

    Returns:
        (是否成功, 结果/错误信息, 执行信息)
    """
    import time
    start_time = time.time()

    try:
        with get_db_connection() as conn:
            # 设置超时
            conn.execute(f"SET statement_timeout = '{timeout}s'")

            # 判断 SQL 类型
            sql_upper = sql.strip().upper()
            is_query = sql_upper.startswith('SELECT') or sql_upper.startswith('SHOW') or sql_upper.startswith('EXPLAIN') or sql_upper.startswith('WITH')

            if is_query:
                # 查询语句，返回结果
                rows = conn.execute(sql).fetchall()
                elapsed = time.time() - start_time

                if rows:
                    result = [dict(row) for row in rows]
                    return True, result, f"返回 {len(result)} 行，耗时 {elapsed:.2f} 秒"
                else:
                    return True, [], f"返回 0 行，耗时 {elapsed:.2f} 秒"
            else:
                # 修改语句（INSERT/UPDATE/DELETE），执行并提交
                result = conn.execute(sql)
                conn.commit()
                elapsed = time.time() - start_time
                affected = result.rowcount if hasattr(result, 'rowcount') else 0
                return True, [], f"执行成功，影响 {affected} 行，耗时 {elapsed:.2f} 秒"

    except Exception as e:
        elapsed = time.time() - start_time
        return False, str(e), f"执行失败，耗时 {elapsed:.2f} 秒"


def save_query(name: str, sql: str, created_by: str = None) -> Tuple[bool, str]:
    """
    保存查询

    Args:
        name: 查询名称（必须唯一）
        sql: SQL 语句
        created_by: 创建者

    Returns:
        (是否成功, 消息)
    """
    try:
        with get_db_connection() as conn:
            conn.execute("""
                INSERT INTO saved_queries (name, sql, created_by)
                VALUES (%s, %s, %s)
            """, (name, sql, created_by))
            conn.commit()
            return True, f"查询 '{name}' 保存成功"
    except Exception as e:
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return False, f"查询名称 '{name}' 已存在，请使用其他名称"
        return False, f"保存失败：{str(e)}"


def get_saved_queries() -> List[Dict[str, Any]]:
    """
    获取所有保存的查询

    Returns:
        查询列表
    """
    with get_db_connection() as conn:
        rows = conn.execute("""
            SELECT id, name, sql, created_by, created_at
            FROM saved_queries
            ORDER BY created_at DESC
        """).fetchall()
        return [dict(row) for row in rows]


def delete_saved_query(query_id: int) -> Tuple[bool, str]:
    """
    删除保存的查询

    Args:
        query_id: 查询 ID

    Returns:
        (是否成功, 消息)
    """
    try:
        with get_db_connection() as conn:
            result = conn.execute("DELETE FROM saved_queries WHERE id = %s", (query_id,))
            conn.commit()
            if result.rowcount > 0:
                return True, "删除成功"
            return False, "查询不存在"
    except Exception as e:
        return False, f"删除失败：{str(e)}"


def get_query_by_id(query_id: int) -> Optional[Dict[str, Any]]:
    """
    根据 ID 获取查询

    Args:
        query_id: 查询 ID

    Returns:
        查询详情
    """
    with get_db_connection() as conn:
        row = conn.execute("""
            SELECT id, name, sql, created_by, created_at
            FROM saved_queries
            WHERE id = %s
        """, (query_id,)).fetchone()
        return dict(row) if row else None


def export_to_excel(data: List[Dict[str, Any]], columns: List[str], selected_fields: List[str] = None) -> bytes:
    """
    导出数据到 Excel

    Args:
        data: 数据列表
        columns: 所有列名
        selected_fields: 选中的列名（可选）

    Returns:
        Excel 文件字节流
    """
    import io
    from openpyxl import Workbook

    # 确定要导出的列
    export_columns = selected_fields if selected_fields else columns

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "数据导出"

    # 写入表头
    for col_idx, col_name in enumerate(export_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(export_columns, 1):
            value = row_data.get(col_name, '')
            # 处理特殊类型
            if isinstance(value, (list, dict)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 调整列宽
    for col_idx, col_name in enumerate(export_columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(col_name) + 2)

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


# 初始化表
init_saved_queries_table()