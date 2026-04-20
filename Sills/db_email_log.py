"""
邮件发送日志数据库操作模块
用于记录和查询邮件发送状态
支持导出任务联系人发送状态
"""
import sqlite3
from datetime import datetime
from Sills.base import get_db_connection


def add_log(task_id, contact_id, email, company_name="", status="sent", error_message=None):
    """添加发送日志

    Args:
        task_id: 任务ID
        contact_id: 联系人ID (空字符串转为NULL以跳过外键约束)
        email: 收件人邮箱
        company_name: 公司名称
        status: 发送状态 (sent/failed)
        error_message: 错误信息(失败时)

    Returns:
        log_id
    """
    # 空字符串转为None,跳过PostgreSQL外键约束检查
    contact_id_value = contact_id if contact_id and contact_id.strip() else None

    with get_db_connection() as conn:
        result = conn.execute("""
            INSERT INTO uni_email_log (task_id, contact_id, email, company_name, status, error_message)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (task_id, contact_id_value, email, company_name, status, error_message or ""))
        conn.commit()
        return result.lastrowid


def get_task_logs(task_id, page=1, page_size=50):
    """获取任务的发送日志"""
    offset = (page - 1) * page_size

    query = """
    SELECT log_id, task_id, contact_id, email, company_name, sent_at, status, error_message
    FROM uni_email_log
    WHERE task_id = ?
    ORDER BY sent_at DESC
    LIMIT ? OFFSET ?
    """

    count_query = "SELECT COUNT(*) FROM uni_email_log WHERE task_id = ?"

    with get_db_connection() as conn:
        total = conn.execute(count_query, (task_id,)).fetchone()[0]
        items = conn.execute(query, (task_id, page_size, offset)).fetchall()
        results = [
            {k: ("" if v is None else v) for k, v in dict(row).items()}
            for row in items
        ]
        return results, total


def get_sent_emails_for_task(task_id):
    """获取任务已发送成功的邮件列表（用于跳过已发送联系人）

    Args:
        task_id: 任务ID

    Returns:
        list 已发送成功的邮箱列表
    """
    with get_db_connection() as conn:
        items = conn.execute("""
            SELECT email
            FROM uni_email_log
            WHERE task_id = ? AND status = 'sent'
        """, (task_id,)).fetchall()
        return [row.get('email') if isinstance(row, dict) else row[0] for row in items]


def get_failed_logs(task_id):
    """获取任务的失败日志"""
    with get_db_connection() as conn:
        items = conn.execute("""
            SELECT log_id, contact_id, email, company_name, sent_at, error_message
            FROM uni_email_log
            WHERE task_id = ? AND status = 'failed'
            ORDER BY sent_at DESC
        """, (task_id,)).fetchall()
        results = [
            {k: ("" if v is None else v) for k, v in dict(row).items()}
            for row in items
        ]
        return results


def get_sent_count(task_id):
    """获取任务已发送数量"""
    with get_db_connection() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM uni_email_log WHERE task_id = ? AND status = 'sent'",
            (task_id,)
        ).fetchone()[0]
        return count


def get_failed_count(task_id):
    """获取任务失败数量"""
    with get_db_connection() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM uni_email_log WHERE task_id = ? AND status = 'failed'",
            (task_id,)
        ).fetchone()[0]
        return count


def get_task_stats(task_id):
    """获取任务统计"""
    with get_db_connection() as conn:
        sent = conn.execute(
            "SELECT COUNT(*) FROM uni_email_log WHERE task_id = ? AND status = 'sent'",
            (task_id,)
        ).fetchone()[0]
        failed = conn.execute(
            "SELECT COUNT(*) FROM uni_email_log WHERE task_id = ? AND status = 'failed'",
            (task_id,)
        ).fetchone()[0]
        return {
            'sent': sent,
            'failed': failed,
            'total': sent + failed
        }


def delete_task_logs(task_id):
    """删除任务的所有日志"""
    with get_db_connection() as conn:
        conn.execute("DELETE FROM uni_email_log WHERE task_id = ?", (task_id,))
        conn.commit()


def update_log_status(task_id, email, status, error_message=None):
    """更新日志状态（用于重试模式）

    Args:
        task_id: 任务ID
        email: 邮箱地址
        status: 新状态 (sent/failed)
        error_message: 错误信息（失败时）

    Returns:
        bool 是否更新成功
    """
    with get_db_connection() as conn:
        conn.execute("""
            UPDATE uni_email_log
            SET status = ?, error_message = ?, sent_at = NOW()
            WHERE task_id = ? AND email = ?
        """, (status, error_message or "", task_id, email.lower()))
        conn.commit()
        return True


def get_recent_logs(limit=100):
    """获取最近的发送日志"""
    with get_db_connection() as conn:
        items = conn.execute("""
            SELECT l.*, t.task_name, t.subject
            FROM uni_email_log l
            LEFT JOIN uni_email_task t ON l.task_id = t.task_id
            ORDER BY l.sent_at DESC
            LIMIT ?
        """, (limit,)).fetchall()
        results = [
            {k: ("" if v is None else v) for k, v in dict(row).items()}
            for row in items
        ]
        return results


def get_all_task_contacts_with_status(task_id):
    """获取任务所有联系人及其发送状态（用于导出）

    Returns:
        list 联系人列表，包含所有字段和发送状态
    """
    from Sills.db_email_task import get_task_contacts

    # 获取任务的所有联系人
    contacts = get_task_contacts(task_id)

    # 获取已发送的日志
    with get_db_connection() as conn:
        sent_logs = conn.execute("""
            SELECT email, status, sent_at, error_message
            FROM uni_email_log
            WHERE task_id = ?
        """, (task_id,)).fetchall()

        # 构建email->状态的映射
        status_map = {}
        for row in sent_logs:
            log = dict(row)
            email = log.get('email', '').lower()
            status_map[email] = {
                'send_status': log.get('status', ''),
                'sent_at': log.get('sent_at', ''),
                'error_message': log.get('error_message', '') or ''
            }

    # 合并联系人信息和发送状态
    results = []
    for contact in contacts:
        email = contact.get('email', '').lower()
        status_info = status_map.get(email, {})

        results.append({
            'contact_id': contact.get('contact_id', ''),
            'email': email,
            'company': contact.get('company', ''),
            'contact_name': contact.get('contact_name', ''),
            'country': contact.get('country', ''),
            'domain': contact.get('domain', ''),
            'position': contact.get('position', ''),
            'phone': contact.get('phone', ''),
            'is_bounced': contact.get('is_bounced', 0),
            'send_count': contact.get('send_count', 0),
            'bounce_count': contact.get('bounce_count', 0),
            'read_count': contact.get('read_count', 0),
            'last_sent_at': contact.get('last_sent_at', ''),
            'remark': contact.get('remark', ''),
            'send_status': status_info.get('send_status', 'pending'),
            'sent_at': status_info.get('sent_at', ''),
            'error_message': status_info.get('error_message', '')
        })

    return results


def export_task_contacts_to_excel(task_id, output_path=None):
    """导出任务联系人发送状态到Excel

    Args:
        task_id: 任务ID
        output_path: 输出文件路径（可选，默认自动生成）

    Returns:
        (success, file_path_or_error) tuple
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime

        # 获取联系人数据
        contacts = get_all_task_contacts_with_status(task_id)
        if not contacts:
            return False, "没有联系人数据"

        # 生成文件路径
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            output_path = f"task_{task_id}_contacts_{timestamp}.xlsx"

        # 创建Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "联系人发送状态"

        # 定义表头
        headers = [
            '联系人ID', '邮箱', '公司', '联系人姓名', '国家', '域名',
            '职位', '电话', '是否退信', '发送次数', '退信次数', '已读次数',
            '最后发送时间', '备注', '本次发送状态', '发送时间', '错误信息'
        ]

        # 设置表头样式
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 状态颜色映射
        status_fills = {
            'sent': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # 绿色-成功
            'failed': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # 红色-失败
            'pending': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色-待发送
        }

        # 写入数据
        for row_idx, contact in enumerate(contacts, 2):
            data = [
                contact.get('contact_id', ''),
                contact.get('email', ''),
                contact.get('company', ''),
                contact.get('contact_name', ''),
                contact.get('country', ''),
                contact.get('domain', ''),
                contact.get('position', ''),
                contact.get('phone', ''),
                contact.get('is_bounced', 0),
                contact.get('send_count', 0),
                contact.get('bounce_count', 0),
                contact.get('read_count', 0),
                contact.get('last_sent_at', ''),
                contact.get('remark', ''),
                contact.get('send_status', 'pending'),
                contact.get('sent_at', ''),
                contact.get('error_message', '')
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')

                # 根据发送状态设置背景色
                if col == 15:  # 发送状态列
                    status = str(value)
                    if status in status_fills:
                        cell.fill = status_fills[status]

        # 调整列宽
        column_widths = [15, 30, 25, 15, 10, 30, 15, 15, 10, 10, 10, 10, 20, 30, 12, 20, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

        # 保存文件
        wb.save(output_path)

        return True, output_path
    except Exception as e:
        return False, str(e)