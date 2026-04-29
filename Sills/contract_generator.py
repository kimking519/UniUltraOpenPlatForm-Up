"""
Contract Generator - 合同生成模块

生成采购合同 Excel 文件，使用 Header + Footer 模板拼接方式。
"""

import os
import platform
import random
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from Sills.base import get_db_connection


# ============================================================
# 模板路径配置
# ============================================================

def _get_template_dir():
    """获取模板目录路径"""
    # 相对于项目根目录
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_dir, "templates", "ht_cn")


def _get_header_template():
    """获取 Header 模板路径"""
    return os.path.join(_get_template_dir(), "hetong-Header.xlsx")


def _get_footer_template():
    """获取 Footer 模板路径"""
    return os.path.join(_get_template_dir(), "hetong-Footer.xlsx")


def _get_default_output_dir():
    """获取默认输出目录"""
    if platform.system() == "Windows":
        return r"E:\1_Business\1_unicorn\3_合规\审计资料2026\3_data\合同"
    else:
        return "/mnt/e/1_Business/1_unicorn/3_合规/审计资料2026/3_data/合同"


# ============================================================
# 合同生成函数
# ============================================================

def generate_contract_from_manager(manager_id, output_dir=None):
    """
    从客户订单生成采购合同

    Args:
        manager_id: 客户订单管理器ID
        output_dir: 输出目录，默认使用 _get_default_output_dir()

    Returns:
        (success, result): 成功时 result 包含 excel_path，失败时 result 包含错误信息
    """
    if openpyxl is None:
        return False, "openpyxl 未安装"

    # 1. 获取客户订单信息
    manager = get_manager_by_id(manager_id)
    if not manager:
        return False, "客户订单不存在"

    customer_order_no = manager.get('customer_order_no', 'UNKNOWN')
    order_date = manager.get('order_date', '')
    cli_name = manager.get('cli_name', 'Unknown')

    # 2. 获取关联的报价数据
    offers = get_manager_offers(manager_id)
    if not offers:
        return False, "没有关联的报价订单"

    # 3. 设置输出目录
    if not output_dir:
        output_dir = _get_default_output_dir()
        # 单独生成时，按日期创建子目录
        date_dir = datetime.now().strftime("%Y%m%d")
        final_output_dir = os.path.join(output_dir, cli_name, date_dir)
        os.makedirs(final_output_dir, exist_ok=True)
    else:
        # 批量生成时，直接使用传入的目录（已由 batch 函数创建）
        final_output_dir = output_dir

    # 4. 加载模板并生成合同
    try:
        excel_path = _generate_contract_excel(
            customer_order_no=customer_order_no,
            order_date=order_date,
            offers=offers,
            output_dir=final_output_dir,
            cli_name=cli_name
        )
        return True, {"excel_path": excel_path, "customer_order_no": customer_order_no}
    except Exception as e:
        return False, str(e)


def _generate_contract_excel(customer_order_no, order_date, offers, output_dir, cli_name):
    """
    生成合同 Excel 文件

    使用 Header + Footer 模板拼接：
    - Header: 第1-26行（合同头部信息）
    - Footer: 从第27行开始（付款条款、签名栏、印章）

    数据填充：
    - C4: 客户订单号
    - G4: 客户订单日期
    - 第22行开始: 动态数据行
    - Footer F1: 金额总和
    """
    # 加载 Header 模板
    header_wb = openpyxl.load_workbook(_get_header_template())
    header_ws = header_wb.active

    # 加载 Footer 模板
    footer_wb = openpyxl.load_workbook(_get_footer_template())
    footer_ws = footer_wb.active

    # 填充 Header 信息
    header_ws['C4'] = customer_order_no
    header_ws['G4'] = order_date

    # 计算需要的数据行数
    data_rows_needed = len(offers)
    header_data_start_row = 22  # Header模板数据开始行
    header_template_rows = 26   # Header模板总行数

    # 删除 Header 模板中的示例数据行（22-26行），保留表头21行
    # 先复制第22行（数据行示例）的样式，用于数据行（白色背景）
    header_row_22_styles = _copy_row_styles(header_ws, 22)

    # 删除22-26行
    for _ in range(5):  # 删除5行（22-26）
        header_ws.delete_rows(22)

    # 插入数据行
    total_amount = 0.0
    for idx, offer in enumerate(offers, start=1):
        insert_row = 21 + idx  # 从第22行开始插入

        # 插入新行
        header_ws.insert_rows(insert_row)

        # 应用样式（从第22行复制，白色背景）
        _apply_row_styles(header_ws, insert_row, header_row_22_styles)

        # 填充数据
        quoted_mpn = offer.get('quoted_mpn') or offer.get('inquiry_mpn', '')
        quoted_brand = offer.get('quoted_brand') or offer.get('inquiry_brand', '')
        actual_qty = offer.get('actual_qty') or offer.get('quoted_qty') or offer.get('inquiry_qty', 0)
        # 合同使用成本价（cost_price_rmb），而不是报价（offer_price_rmb）
        cost_price_rmb = offer.get('cost_price_rmb') or 0

        # 金额 = 数量 × 含税单价（成本价）
        amount = actual_qty * cost_price_rmb if actual_qty and cost_price_rmb else 0
        total_amount += amount

        # 填入各列 (F列=含税单价, G列=金额=数量×含税单价)
        header_ws.cell(row=insert_row, column=1, value=idx)  # A: 序号
        header_ws.cell(row=insert_row, column=2, value=quoted_mpn)  # B: 型号
        header_ws.cell(row=insert_row, column=3, value=quoted_brand)  # C: 品牌
        header_ws.cell(row=insert_row, column=4, value='pcs')  # D: 单位
        header_ws.cell(row=insert_row, column=5, value=actual_qty)  # E: 数量
        header_ws.cell(row=insert_row, column=6, value=round(cost_price_rmb, 1))  # F: 含税单价(成本价，精度1位)
        header_ws.cell(row=insert_row, column=7, value=round(amount, 1))  # G: 金额=数量×含税单价
        header_ws.cell(row=insert_row, column=8, value='')  # H: 备注

    # Header部分的最后一行（数据结束行）
    header_end_row = 21 + data_rows_needed

    # 合并 Footer 模板到 Header（参考PI生成代码）
    # Footer 模板的第1行对应最终文件的第 header_end_row + 1 行
    footer_start_row = header_end_row + 1

    # 复制 Footer 的所有行（值、样式）
    for src_row in range(1, footer_ws.max_row + 1):
        dst_row = footer_start_row + src_row - 1
        for col in range(1, 9):
            src_cell = footer_ws.cell(row=src_row, column=col)
            dst_cell = header_ws.cell(row=dst_row, column=col)

            # 复制值
            dst_cell.value = src_cell.value

            # 复制样式
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)

        # 复制行高
        if footer_ws.row_dimensions[src_row].height:
            header_ws.row_dimensions[dst_row].height = footer_ws.row_dimensions[src_row].height

    # 复制 Footer 的合并单元格（调整行号）
    for merged_range in footer_ws.merged_cells.ranges:
        new_min_row = footer_start_row + merged_range.min_row - 1
        new_max_row = footer_start_row + merged_range.max_row - 1
        new_range = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{new_min_row}:{openpyxl.utils.get_column_letter(merged_range.max_col)}{new_max_row}"
        try:
            header_ws.merge_cells(new_range)
        except:
            pass

    # 复制 Footer 中的图片（印章）
    for img in footer_ws._images:
        from copy import deepcopy
        new_img = deepcopy(img)
        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
            new_img.anchor._from.row = footer_start_row + img.anchor._from.row - 1
            if hasattr(img.anchor, 'to'):
                new_img.anchor.to.row = footer_start_row + img.anchor.to.row - 1
        header_ws.add_image(new_img)

    # 更新数据（只修改值，不改变样式）
    # Footer第1行 = footer_start_row
    # F列：合计金额公式
    sum_formula = f"=SUM(G22:G{header_end_row})"
    header_ws.cell(row=footer_start_row, column=6).value = sum_formula

    # Footer第2行 = footer_start_row + 1
    # D列：大写金额
    chinese_amount = _number_to_chinese(total_amount)
    header_ws.cell(row=footer_start_row + 1, column=4).value = chinese_amount

    # Footer中的订单日期更新
    # B29和F29填写订单日期（不是当前日期）
    header_ws.cell(row=footer_start_row + 28, column=2).value = order_date  # B29
    header_ws.cell(row=footer_start_row + 28, column=6).value = order_date  # F29

    # 保存文件
    random_num = random.randint(0, 999999)
    filename = f"采购合同_UnicornTech_{customer_order_no}_{random_num:06d}.xlsx"
    excel_path = os.path.join(output_dir, filename)
    header_wb.save(excel_path)
    header_wb.close()
    footer_wb.close()

    return excel_path


def _copy_row_styles(ws, row_num):
    """
    复制一行中所有单元格的样式
    """
    styles = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        style = {}
        if cell.font:
            style['font'] = copy(cell.font)
        if cell.alignment:
            style['alignment'] = copy(cell.alignment)
        if cell.border:
            style['border'] = copy(cell.border)
        if cell.fill:
            style['fill'] = copy(cell.fill)
        if cell.number_format and cell.number_format != 'General':
            style['number_format'] = cell.number_format
        styles[col_idx] = style
    return styles


def _apply_row_styles(ws, row_num, styles):
    """
    应用样式到一行
    """
    for col_idx, style in styles.items():
        cell = ws.cell(row=row_num, column=col_idx)
        if 'font' in style:
            cell.font = style['font']
        if 'alignment' in style:
            cell.alignment = style['alignment']
        if 'border' in style:
            cell.border = style['border']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'number_format' in style:
            cell.number_format = style['number_format']


def _number_to_chinese(num):
    """
    将数字转换为中文大写金额
    例如: 10080 -> "壹万零捌拾元整"
    """
    if num is None or num == 0:
        return "零元整"

    # 中文数字字符
    chinese_nums = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    chinese_units = ['', '拾', '佰', '仟']
    chinese_big_units = ['', '万', '亿']

    # 处理小数部分（角分）
    num_float = float(num)
    integer_part = int(num_float)
    decimal_part = round(num_float - integer_part, 2)

    # 转换整数部分
    result = ""

    if integer_part > 0:
        # 分组处理（每4位一组，从低位到高位）
        groups = []
        temp = integer_part
        while temp > 0:
            groups.append(temp % 10000)
            temp = temp // 10000

        for i, group in enumerate(groups):
            if group == 0:
                continue

            group_str = ""
            # 处理四位数字：仟、佰、拾、个
            digits = [group // 1000, (group // 100) % 10, (group // 10) % 10, group % 10]

            for j, digit in enumerate(digits):
                if digit > 0:
                    group_str += chinese_nums[digit] + chinese_units[3 - j]
                elif group_str and not group_str.endswith('零'):
                    group_str += '零'

            # 移除末尾多余的零
            while group_str.endswith('零'):
                group_str = group_str[:-1]

            # 高位组（万位、亿位）与低位组之间需要零的情况
            # 当低位组（groups[更小的索引]）的最高位为零时（即低位组 < 1000）
            if i > 0:
                # 检查所有更低位组是否有非零值且最高位为零
                need_zero = False
                for lower_i in range(i):
                    if groups[lower_i] > 0 and groups[lower_i] < 1000:
                        need_zero = True
                        break
                if need_zero and result and not result.startswith('零'):
                    result = '零' + result

            result = group_str + chinese_big_units[i] + result

        # 移除开头多余的零
        while result.startswith('零'):
            result = result[1:]

        # 添加"元"
        result += "元"

    # 处理小数部分
    if decimal_part > 0:
        jiao = int(decimal_part * 10)
        fen = int(round(decimal_part * 100) % 10)

        if jiao > 0:
            result += chinese_nums[jiao] + "角"
        if fen > 0:
            result += chinese_nums[fen] + "分"
    else:
        # 没有小数部分，整数末尾加"整"
        result += "整"

    return result


# ============================================================
# 数据查询函数（从 db_order_manager 导入或本地实现）
# ============================================================

def get_manager_by_id(manager_id):
    """
    获取客户订单管理器信息
    """
    with get_db_connection() as conn:
        row = conn.execute("""
            SELECT m.*, c.cli_name
            FROM uni_order_manager m
            LEFT JOIN uni_cli c ON m.cli_id = c.cli_id
            WHERE m.manager_id = ?
        """, (manager_id,)).fetchone()
        if row:
            return dict(row)
        return None


def get_manager_offers(manager_id):
    """
    获取客户订单关联的所有报价
    """
    with get_db_connection() as conn:
        rows = conn.execute("""
            SELECT o.*
            FROM uni_order_manager_rel r
            JOIN uni_offer o ON r.offer_id = o.offer_id
            WHERE r.manager_id = ?
            ORDER BY o.offer_id
        """, (manager_id,)).fetchall()
        return [dict(row) for row in rows]


# ============================================================
# 批量生成函数
# ============================================================

def generate_contract_batch(manager_ids, output_dir=None):
    """
    批量生成合同

    Args:
        manager_ids: 客户订单ID列表
        output_dir: 输出目录

    Returns:
        (success_count, errors, generated_files)
    """
    if not output_dir:
        output_dir = _get_default_output_dir()

    # 创建时间戳子目录
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    batch_output_dir = os.path.join(output_dir, f"合同批量-{timestamp}")
    os.makedirs(batch_output_dir, exist_ok=True)

    success_count = 0
    errors = []
    generated_files = []

    for manager_id in manager_ids:
        try:
            success, result = generate_contract_from_manager(manager_id, batch_output_dir)
            if success:
                success_count += 1
                generated_files.append(os.path.basename(result['excel_path']))
            else:
                # 获取订单号用于错误信息
                manager = get_manager_by_id(manager_id)
                order_no = manager.get('customer_order_no', manager_id) if manager else manager_id
                errors.append(f"{order_no}: {result}")
        except Exception as e:
            errors.append(f"{manager_id}: {str(e)}")

    return success_count, errors, generated_files, batch_output_dir